#!/usr/bin/env python3
"""
Genera GANTT.html — "Carga de taller por trabajador"

Usa los MISMOS exports de Asana que descarga exportar_taller.py.
No lee el Excel de nadie: quien hace cada cosa sale del campo Assignee,
y las fechas de Start Date / Due Date.

Los datos quedan incrustados en el HTML (igual que generar_kpi.py),
asi que el archivo funciona solo, sin fetch ni CORS ni servidor.

USO
---
  python3 generar_gantt.py <carpeta_con_excels> <archivo_salida_html>

Por defecto:
  carpeta   = data-taller
  salida    = GANTT.HTML   (raiz del repo, junto a KPI.HTML)
"""

import os
import re
import sys
import glob
import json
import datetime
import warnings
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------
# CONFIGURACION
# ---------------------------------------------------------------------

# Cuantas semanas hacia atras mostrar (tareas ya cerradas quedan en blanco)
SEMANAS_ATRAS = 2

# Este tablero es independiente del KPI. Por eso, por defecto:
#  - toma TODAS las tareas, no solo las que cuelgan de una tarea padre
#  - no aplica las exclusiones de despacho/costos/cierre del KPI
# Ponlos en True si quieres que se comporte igual que generar_kpi.py
SOLO_NIVEL_2 = False
APLICAR_EXCLUSIONES = False

# Areas -> color de linea. Tonos apagados, legibles sobre papel blanco.
COLOR_AREA = {
    "Ingeniería":         "#1B4F8C",
    "Compras":            "#A8620F",
    "Producción":         "#2E7D52",
    "Bodega":             "#6A3F8F",
    "Logística":          "#A63E5A",
    "Control de calidad": "#1D6E78",
    "Servicios":          "#6B7233",
    "Equipo Proyecto":    "#54606E",
}

ASSIGNEE_AREA = {
    "Ignacio García": "Servicios", "Ignacio Garcia": "Servicios",
    "IGNACIO GARCIA MARTINEZ": "Servicios", "Ignacio Garcia Martinez": "Servicios",
    "Sergio Saavedra": "Servicios", "sergio saavedra fernandez": "Servicios",
    "Sergio Saavedra Fernandez": "Servicios",
    "Carlos Pérez": "Equipo Proyecto", "Carlos Perez": "Equipo Proyecto",
    "Francisca Ramos": "Equipo Proyecto",
    "Francisca Alejandra Ramos Aravales": "Equipo Proyecto",
    "Sergio de la Fuente": "Equipo Proyecto",
    "Sergio de la Fuente Fernandez": "Equipo Proyecto",
    "David Blazquez": "Equipo Proyecto", "DAVID BLAZQUEZ": "Equipo Proyecto",
    "Benjamin Umaña": "Producción", "Nicolás Mol": "Producción",
    "Nicolas Mol": "Producción", "nespinoza@zitron.com": "Producción",
    "Natalia Espinoza": "Producción", "NATALYA ESPINOZA": "Producción",
    "Eliana": "Compras", "Yerlia": "Compras",
    "Yerlia Ayleen Castillo Diaz": "Compras",
    "Rose": "Compras", "Rosemary Singh": "Compras",
    "Hernán Gutierrez": "Ingeniería",
    "Hernan Roberto Gutierrez Barrientos": "Ingeniería",
    "Hernan Gutierrez": "Ingeniería",
    "Francisca González": "Ingeniería", "Francisca González Cornejo": "Ingeniería",
    "Francisca Gonzalez": "Ingeniería",
    "Gonzalo Davila": "Ingeniería", "Gonzalo Dávila": "Ingeniería",
    "Gabriel Venegas": "Ingeniería", "Gabriel Venega": "Ingeniería",
    "Nicolás López": "Control de calidad", "Nicolas Lopez": "Control de calidad",
    "Víctor Muñoz": "Bodega", "Victor Muñoz": "Bodega", "Victor Munoz": "Bodega",
    "Karin Pinto": "Logística",
}

SECTION_AREA_KEYWORDS = [
    ("produccion", "Producción"), ("producción", "Producción"),
    ("ingenier", "Ingeniería"),
    ("compra", "Compras"),
    ("bodega", "Bodega"),
    ("logist", "Logística"), ("logíst", "Logística"),
    ("montaje", "Producción"),
]

EXCLUIR_NOMBRE_CONTIENE = ["despacho"]
EXCLUIR_NOMBRE_EXACTO = ["costos"]
EXCLUIR_SECCION_CONTIENE = ["cierre de proyecto"]
EXCLUIR_ASSIGNEE_EXACTO = ["nicolas"]


# ---------------------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------------------

def _norm(s):
    s = str(s or "").replace("\u00a0", " ").lower().strip()
    for a, b in {"á": "a", "é": "e", "í": "i", "ó": "o",
                 "ú": "u", "ü": "u", "ñ": "n"}.items():
        s = s.replace(a, b)
    return " ".join(s.split())


def to_date(val):
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def area_for(assignee, section):
    sec = _norm(section)
    for kw, area in SECTION_AREA_KEYWORDS:
        if kw in sec:
            return area
    if assignee:
        a_norm = _norm(assignee)
        a_words = set(a_norm.split())
        for k, v in ASSIGNEE_AREA.items():
            k_norm = _norm(k)
            if k_norm == a_norm or k_norm in a_norm:
                return v
            k_words = set(k_norm.split())
            if k_words and k_words.issubset(a_words):
                return v
    return "Equipo Proyecto"


def find_header_row(ws, max_scan=12):
    """Devuelve (fila, mapa_canonico). Acepta cabeceras en ingles o
    espanol, porque el idioma del export depende de como tenga puesta
    Asana la persona que genero la sesion."""
    for r in range(1, max_scan + 1):
        valores = [c.value for c in ws[r]]
        mapa = {}
        for i, celda in enumerate(valores):
            n = _norm(celda)
            if not n:
                continue
            for canon, alias in ALIAS.items():
                if canon in mapa:
                    continue
                if any(n == _norm(a) for a in alias):
                    mapa[canon] = i
                    break
        if "Name" in mapa and ("Due Date" in mapa or "Start Date" in mapa):
            return r, mapa, valores
    return None, None, None


ALIAS = {
    "Task ID":      ["Task ID", "ID de la tarea", "ID de tarea"],
    "Name":         ["Name", "Nombre", "Nombre de la tarea"],
    "Section/Column": ["Section/Column", "Seccion/Columna", "Sección/Columna",
                       "Seccion", "Sección"],
    "Assignee":     ["Assignee", "Responsable", "Asignado a", "Encargado"],
    "Start Date":   ["Start Date", "Fecha de inicio", "Fecha inicio"],
    "Due Date":     ["Due Date", "Fecha de vencimiento", "Fecha de entrega",
                     "Fecha limite", "Fecha límite"],
    "Completed At": ["Completed At", "Completada el", "Fecha de finalizacion",
                     "Fecha de finalización", "Completado el"],
    "Parent task":  ["Parent task", "Tarea principal", "Tarea padre"],
}


def extraer_ot(texto):
    m = re.search(r"OT\s*-?\s*(\d{3,5})", str(texto or ""), re.I)
    return "OT-" + m.group(1) if m else None


def extraer_pedido(nombre_proyecto):
    m = re.match(r"\s*(\d{7,9})", str(nombre_proyecto or ""))
    return m.group(1) if m else ""


# ---------------------------------------------------------------------
# LECTURA
# ---------------------------------------------------------------------

def process_file(path, hoy, desde):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    proyecto = Path(path).stem.strip()

    header_row, col, cabeceras = find_header_row(ws)
    if header_row is None:
        print(f"  [{proyecto}] No se reconocio la fila de cabeceras.")
        print("  Primeras filas del archivo:")
        for r in range(1, 4):
            fila = [str(c.value)[:22] for c in ws[r] if c.value is not None]
            print(f"    fila {r}: {fila}")
        wb.close()
        return []

    print(f"  [{proyecto}] cabeceras en fila {header_row}: "
          + ", ".join(sorted(col.keys())))
    faltan = [c for c in ("Assignee", "Start Date", "Due Date") if c not in col]
    if faltan:
        print(f"  [{proyecto}] AVISO: el export no trae {', '.join(faltan)}. "
              "Agrega esas columnas al exportar desde Asana.")

    def get(row, key):
        i = col.get(key)
        return row[i] if i is not None and i < len(row) else None

    # Contadores para saber por que se descarta cada fila
    n = {"filas": 0, "cabecera": 0, "nivel1": 0, "excluida": 0,
         "sin_responsable": 0, "sin_fechas": 0, "fuera_ventana": 0, "ok": 0}

    tareas = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or get(row, "Name") in (None, ""):
            continue
        n["filas"] += 1

        if SOLO_NIVEL_2 and get(row, "Parent task") in (None, ""):
            n["nivel1"] += 1
            continue

        name = str(get(row, "Name")).strip()
        section = get(row, "Section/Column") or ""
        assignee = get(row, "Assignee") or ""

        n_name, n_sec, n_asg = _norm(name), _norm(section), _norm(assignee)

        # Cabeceras de seccion: nombre terminado en ":" y sin responsable
        if n_name.endswith(":") and not str(assignee).strip():
            n["cabecera"] += 1
            continue

        if APLICAR_EXCLUSIONES:
            if (any(k in n_name for k in EXCLUIR_NOMBRE_CONTIENE)
                    or n_name in EXCLUIR_NOMBRE_EXACTO
                    or any(k in n_sec for k in EXCLUIR_SECCION_CONTIENE)
                    or n_asg in EXCLUIR_ASSIGNEE_EXACTO):
                n["excluida"] += 1
                continue

        if not str(assignee).strip():
            n["sin_responsable"] += 1
            assignee = ""          # se decide en main que hacer con estas

        due = to_date(get(row, "Due Date"))
        start = to_date(get(row, "Start Date")) or due
        if not due or not start:
            n["sin_fechas"] += 1
            continue
        if due < start:
            start, due = due, start

        completed = to_date(get(row, "Completed At"))

        if due < desde and (completed is None or completed < desde):
            n["fuera_ventana"] += 1
            continue

        if completed:
            estado = "cerrada_tarde" if completed > due else "cerrada"
        else:
            estado = "vencida" if due < hoy else "curso"

        task_id = get(row, "Task ID")
        n["ok"] += 1
        tareas.append({
            "n": name,
            "p": str(assignee).strip(),
            "a": area_for(assignee, section),
            "s": section or "",
            "pr": proyecto,
            "ped": extraer_pedido(proyecto),
            "ot": extraer_ot(proyecto) or extraer_ot(name) or "",
            "i": start.isoformat(),
            "f": due.isoformat(),
            "c": completed.isoformat() if completed else None,
            "e": estado,
            "u": f"https://app.asana.com/0/0/{int(task_id)}/f" if task_id else None,
        })

    print(f"  [{proyecto}] {n['filas']} filas leidas -> {n['ok']} al Gantt")
    for etiqueta, clave in (("cabeceras de seccion", "cabecera"),
                            ("sin responsable", "sin_responsable"),
                            ("sin fecha de inicio o vencimiento", "sin_fechas"),
                            ("terminadas hace mas de "
                             f"{SEMANAS_ATRAS} semanas", "fuera_ventana"),
                            ("nivel 1 (SOLO_NIVEL_2)", "nivel1"),
                            ("excluidas por reglas", "excluida")):
        if n[clave]:
            print(f"      descartadas por {etiqueta}: {n[clave]}")

    wb.close()
    return tareas


# ---------------------------------------------------------------------
# PLANTILLA — lamina de plano tecnico sobre papel blanco
# ---------------------------------------------------------------------

TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<title>Carga de taller — Gantt por trabajador</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=IBM+Plex+Sans+Condensed:wght@500;600;700&family=IBM+Plex+Sans:wght@400;500;600&display=swap" rel="stylesheet">
<style>
:root{
  --papel:#FFFFFF;
  --reticula:#E9EDF2;
  --regla:#C6CFDA;
  --regla-fuerte:#9AA7B6;
  --tinta:#0E1620;
  --tinta-2:#56626F;
  --tinta-3:#8C98A5;
  --rojo:#C0392B;
  --azul:#1B4F8C;
  --dia:26px;
  --cond:"IBM Plex Sans Condensed",system-ui,sans-serif;
  --sans:"IBM Plex Sans",system-ui,-apple-system,sans-serif;
  --mono:"IBM Plex Mono",ui-monospace,Menlo,monospace;
}
*{box-sizing:border-box}
html,body{margin:0;padding:0}
body{
  background:var(--papel);
  color:var(--tinta);
  font-family:var(--sans);
  font-size:14px;
  -webkit-font-smoothing:antialiased;
}
button,input{font-family:inherit;font-size:inherit;color:inherit}
button{cursor:pointer}

/* ---------- Encabezado tipo lamina ---------- */
.lamina{
  border-bottom:2px solid var(--tinta);
  padding:22px 26px 0;
  display:flex;gap:28px;align-items:flex-start;flex-wrap:wrap;
}
.rotulo h1{
  font-family:var(--cond);
  font-size:34px;font-weight:700;letter-spacing:.06em;
  text-transform:uppercase;margin:0;line-height:.95;
}
.rotulo .bajada{
  font-family:var(--mono);font-size:11px;color:var(--tinta-2);
  letter-spacing:.08em;text-transform:uppercase;margin-top:7px;
}

/* Cajetin: la firma de la lamina */
.cajetin{
  margin-left:auto;
  border:1.5px solid var(--tinta);
  display:grid;grid-template-columns:repeat(4,minmax(78px,auto));
}
.cajetin div{
  border-right:1px solid var(--regla);
  padding:6px 12px 7px;
}
.cajetin div:last-child{border-right:0}
.cajetin .et{
  font-family:var(--mono);font-size:8.5px;letter-spacing:.14em;
  text-transform:uppercase;color:var(--tinta-3);
}
.cajetin .dt{
  font-family:var(--cond);font-size:16px;font-weight:600;
  margin-top:2px;white-space:nowrap;
}

/* ---------- Cotas / indicadores ---------- */
.cotas{
  display:flex;flex-wrap:wrap;
  border-bottom:1px solid var(--regla);
  padding:0 26px;
}
.cota{
  padding:14px 30px 13px 0;margin-right:30px;
  border-right:1px solid var(--reticula);
}
.cota:last-child{border-right:0}
.cota .et{
  font-family:var(--mono);font-size:9px;letter-spacing:.15em;
  text-transform:uppercase;color:var(--tinta-3);
}
.cota .val{
  font-family:var(--mono);font-size:27px;font-weight:500;
  line-height:1.15;margin-top:3px;font-variant-numeric:tabular-nums;
}
.cota.alerta .val{color:var(--rojo)}

/* ---------- Controles ---------- */
.controles{
  display:flex;gap:10px;align-items:center;flex-wrap:wrap;
  padding:14px 26px;border-bottom:1px solid var(--regla);
}
.grupo{display:flex;border:1px solid var(--regla-fuerte)}
.grupo button{
  padding:6px 13px;background:var(--papel);border:0;
  border-right:1px solid var(--regla);
  font-family:var(--mono);font-size:11px;letter-spacing:.06em;
  text-transform:uppercase;color:var(--tinta-2);
}
.grupo button:last-child{border-right:0}
.grupo button[aria-pressed="true"]{background:var(--tinta);color:var(--papel)}
.buscar{
  padding:6px 11px;border:1px solid var(--regla-fuerte);
  background:var(--papel);min-width:210px;font-size:13px;
}
.buscar::placeholder{color:var(--tinta-3)}
.leyenda{display:flex;gap:15px;margin-left:auto;flex-wrap:wrap}
.leyenda span{
  display:flex;align-items:center;gap:6px;
  font-family:var(--mono);font-size:10px;letter-spacing:.05em;
  text-transform:uppercase;color:var(--tinta-2);
}
.tramo{width:16px;height:4px;flex:none}

/* ---------- Tablero ---------- */
.scroll{overflow-x:auto}
.rejilla{min-width:max-content;position:relative}
.fila{display:flex;border-bottom:1px solid var(--reticula)}
.fila.cabecera{
  position:sticky;top:0;z-index:6;background:var(--papel);
  border-bottom:1.5px solid var(--tinta);
}
.izq{
  width:246px;flex:none;padding:11px 16px 10px;
  border-right:1.5px solid var(--tinta);
  position:sticky;left:0;z-index:5;background:var(--papel);
}
.fila.cabecera .izq{z-index:7}
.nombre{
  font-family:var(--cond);font-size:17px;font-weight:600;
  letter-spacing:.03em;text-transform:uppercase;line-height:1.15;
}
.rol{
  font-family:var(--mono);font-size:9.5px;letter-spacing:.1em;
  text-transform:uppercase;color:var(--tinta-3);margin-top:3px;
}
/* Cota de carga: linea de cota, no barra de progreso */
.cotaLinea{
  margin-top:9px;height:7px;position:relative;
  border-left:1px solid var(--regla-fuerte);
  border-right:1px solid var(--regla-fuerte);
}
.cotaLinea::before{
  content:"";position:absolute;top:3px;left:0;right:0;
  border-top:1px solid var(--regla);
}
.cotaLinea i{
  position:absolute;top:2px;left:0;height:3px;display:block;
  background:var(--tinta);
}
.cotaLinea i.alerta{background:var(--rojo)}
.cifras{
  font-family:var(--mono);font-size:9.5px;color:var(--tinta-3);
  margin-top:5px;display:flex;justify-content:space-between;
}
.cifras b{font-weight:500;color:var(--rojo)}

.pista{position:relative;flex:1;min-height:50px}
.fila.cabecera .pista{min-height:0}

.meses{display:flex;height:21px;border-bottom:1px solid var(--reticula)}
.mes{
  font-family:var(--mono);font-size:9.5px;font-weight:500;
  letter-spacing:.16em;text-transform:uppercase;color:var(--tinta-2);
  padding:5px 0 0 8px;border-left:1px solid var(--regla);
  overflow:hidden;white-space:nowrap;
}
.dias{display:flex;height:25px}
.dia{
  width:var(--dia);flex:none;text-align:center;
  font-family:var(--mono);font-size:9.5px;color:var(--tinta-3);
  padding-top:6px;border-left:1px solid var(--reticula);
  font-variant-numeric:tabular-nums;
}
.dia.lunes{border-left:1px solid var(--regla)}
.dia.finde{color:#B4BEC9}

.trama{position:absolute;inset:0;display:flex;pointer-events:none}
.celda{width:var(--dia);flex:none;border-left:1px solid var(--reticula)}
.celda.lunes{border-left:1px solid var(--regla)}
/* Achurado de fin de semana, como zona fuera de alcance en un plano */
.celda.finde{
  background:repeating-linear-gradient(45deg,
    transparent 0 3px, #EFF2F6 3px 4px);
}

/* Linea de hoy: eje de referencia */
.hoy{
  position:absolute;top:0;bottom:0;width:0;
  border-left:1px dashed var(--rojo);z-index:4;pointer-events:none;
}
.hoy::after{
  content:"HOY";position:absolute;top:1px;left:4px;
  font-family:var(--mono);font-size:8.5px;font-weight:600;
  letter-spacing:.16em;color:var(--rojo);background:var(--papel);
  padding:1px 3px;
}

/* Barras: elementos dibujados, no pastillas */
.barra{
  position:absolute;height:23px;border-radius:1px;
  padding:0 7px;display:flex;align-items:center;gap:6px;
  font-size:11px;font-weight:500;color:#fff;
  overflow:hidden;white-space:nowrap;text-decoration:none;
  border:1px solid rgba(0,0,0,.32);
}
.barra:hover{filter:brightness(1.12)}
.barra .ot{
  font-family:var(--mono);font-size:9.5px;font-weight:600;
  letter-spacing:.04em;opacity:.85;
}
.barra .txt{overflow:hidden;text-overflow:ellipsis}
/* Cerrada: contorno, sin relleno — el trabajo ya no ocupa capacidad */
.barra.cerrada{background:var(--papel)!important;border-width:1px}
.barra.cerrada .txt,.barra.cerrada .ot{opacity:.9}
.barra.vencida::before{
  content:"";position:absolute;inset:0;pointer-events:none;
  background:repeating-linear-gradient(45deg,
    transparent 0 4px, rgba(255,255,255,.4) 4px 8px);
}
.barra.vencida{border:1.5px solid var(--rojo)}

.vacio{padding:56px 26px;text-align:center}
.vacio .t{
  font-family:var(--cond);font-size:19px;letter-spacing:.06em;
  text-transform:uppercase;color:var(--tinta-2);
}
.vacio p{color:var(--tinta-3);font-size:13px}

/* Aviso cuando no hay responsables asignados */
#aviso{
  display:none;margin:0;padding:12px 26px;
  border-bottom:1px solid var(--regla);
  border-left:4px solid var(--rojo);
  background:#FDF4F3;font-size:13px;color:var(--tinta-2);line-height:1.5;
}
#aviso.on{display:block}
#aviso b{color:var(--tinta);font-weight:600}

.pie{
  padding:14px 26px 30px;border-top:1px solid var(--regla);
  font-family:var(--mono);font-size:10px;letter-spacing:.05em;
  color:var(--tinta-3);display:flex;gap:26px;flex-wrap:wrap;
  text-transform:uppercase;
}
/* ---------- Portada de acceso ---------- */
#gate{
  position:fixed;inset:0;background:var(--papel);z-index:9999;
  display:flex;align-items:center;justify-content:center;
}
#gate.hidden{display:none}
#gate .caja{
  border:1.5px solid var(--tinta);width:330px;padding:0;
}
#gate .caja .franja{
  border-bottom:1px solid var(--regla);padding:14px 20px;
}
#gate .caja h2{
  font-family:var(--cond);font-size:23px;font-weight:700;
  letter-spacing:.05em;text-transform:uppercase;margin:0;line-height:1;
}
#gate .caja .sub{
  font-family:var(--mono);font-size:9.5px;letter-spacing:.13em;
  text-transform:uppercase;color:var(--tinta-3);margin-top:6px;
}
#gate .caja .cuerpo{padding:20px}
#gate input{
  padding:9px 12px;border:1px solid var(--regla-fuerte);
  font-size:14px;width:100%;background:var(--papel);
}
#gate button{
  margin-top:10px;padding:10px 16px;border:0;width:100%;
  background:var(--tinta);color:var(--papel);
  font-family:var(--mono);font-size:11px;letter-spacing:.14em;
  text-transform:uppercase;
}
#gate .error{
  font-family:var(--mono);font-size:10px;letter-spacing:.08em;
  text-transform:uppercase;color:var(--rojo);margin-top:9px;height:13px;
}
#app{display:none}
#app.show{display:block}

@media (max-width:760px){
  .izq{width:158px}
  .rotulo h1{font-size:26px}
  .cajetin{margin-left:0;margin-top:14px}
  .leyenda{margin-left:0;width:100%}
}
@media print{
  .controles{display:none}
  body{font-size:11px}
}
</style>
</head>
<body>

<div id="gate">
  <div class="caja">
    <div class="franja">
      <h2>Carga de taller</h2>
      <div class="sub">Acceso restringido</div>
    </div>
    <div class="cuerpo">
      <input type="password" id="gatePass" placeholder="Contraseña"
             onkeydown="if(event.key==='Enter')checkPass()">
      <button onclick="checkPass()">Ingresar</button>
      <div class="error" id="gateError"></div>
    </div>
  </div>
</div>

<div id="app">

<header class="lamina">
  <div class="rotulo">
    <h1>Carga de taller</h1>
    <div class="bajada">Programa por trabajador · origen Asana</div>
  </div>
  <div class="cajetin">
    <div><div class="et">Emitido</div><div class="dt" id="cbFecha">—</div></div>
    <div><div class="et">Proyectos</div><div class="dt" id="cbProy">—</div></div>
    <div><div class="et">Tareas</div><div class="dt" id="cbTareas">—</div></div>
    <div><div class="et">Escala</div><div class="dt" id="cbEscala">8 sem</div></div>
  </div>
</header>

<div id="aviso"></div>

<section class="cotas">
  <div class="cota"><div class="et">Personas con carga</div><div class="val" id="kPers">—</div></div>
  <div class="cota"><div class="et">Tareas abiertas</div><div class="val" id="kAbiertas">—</div></div>
  <div class="cota"><div class="et">Media por persona</div><div class="val" id="kMedia">—</div></div>
  <div class="cota alerta"><div class="et">Vencidas</div><div class="val" id="kVenc">—</div></div>
</section>

<div class="controles">
  <div class="grupo" id="gVista">
    <button data-v="persona" aria-pressed="true">Por persona</button>
    <button data-v="area" aria-pressed="false">Por área</button>
  </div>
  <div class="grupo" id="gRango">
    <button data-s="4" aria-pressed="false">4 sem</button>
    <button data-s="8" aria-pressed="true">8 sem</button>
    <button data-s="16" aria-pressed="false">16 sem</button>
  </div>
  <div class="grupo" id="gCerradas">
    <button data-c="0" aria-pressed="true">Solo abiertas</button>
    <button data-c="1" aria-pressed="false">Con cerradas</button>
  </div>
  <input class="buscar" id="buscar" placeholder="Filtrar por persona, OT o tarea">
  <div class="leyenda" id="leyenda"></div>
</div>

<div class="scroll"><div class="rejilla" id="rejilla"></div></div>

<div class="pie">
  <span>Barra = Start Date → Due Date en Asana</span>
  <span>Contorno sin relleno = tarea cerrada</span>
  <span>Achurado rojo = vencida sin completar</span>
  <span>Clic en la barra abre la tarea</span>
</div>

</div><!-- /#app -->

<script>
// Misma clave y misma sesion que KPI.HTML: si ya entraste alli en esta
// pestana, este tablero se abre directo.
const CLAVE = "zitron2026!";
function checkPass(){
  if(document.getElementById('gatePass').value === CLAVE){
    sessionStorage.setItem('zitron_ok','1');
    document.getElementById('gate').classList.add('hidden');
    document.getElementById('app').classList.add('show');
  } else {
    document.getElementById('gateError').textContent = 'Contraseña incorrecta';
  }
}
if(sessionStorage.getItem('zitron_ok') === '1'){
  document.getElementById('gate').classList.add('hidden');
  document.getElementById('app').classList.add('show');
}

const TAREAS = __DATA__;
const COLOR  = __COLORES__;
const FECHA  = __FECHA_CORTA__;
const AVISO  = __AVISO__;
const POR_ETAPA = AVISO !== "";

if(POR_ETAPA){
  const a = document.getElementById("aviso");
  a.className = "on";
  a.innerHTML = "<b>Vista por etapa.</b> " + AVISO;
  document.querySelector('#gVista button[data-v="persona"]').textContent = "Por etapa";
}

let vista = "persona", semanas = 8, verCerradas = false, filtro = "";
const $ = s => document.querySelector(s);
const hoy = (()=>{const d=new Date();d.setHours(0,0,0,0);return d;})();
const diaMs = 86400000;
const D = s => { const [a,m,d] = s.split("-").map(Number); return new Date(a,m-1,d); };
const sumaDias = (d,n)=>{const x=new Date(d);x.setDate(x.getDate()+n);return x;};
const difDias = (a,b)=>Math.round((b-a)/diaMs);
const esFinde = d => d.getDay()===0 || d.getDay()===6;

TAREAS.forEach(t => { t._i = D(t.i); t._f = D(t.f); });

$("#leyenda").innerHTML = Object.entries(COLOR)
  .map(([a,c])=>`<span><i class="tramo" style="background:${c}"></i>${a}</span>`).join("");
$("#cbFecha").textContent  = FECHA;
$("#cbProy").textContent   = new Set(TAREAS.map(t=>t.pr)).size;
$("#cbTareas").textContent = TAREAS.length;

function visibles(){
  let out = TAREAS;
  if(!verCerradas) out = out.filter(t => t.e === "curso" || t.e === "vencida");
  const f = filtro.toLowerCase();
  if(f) out = out.filter(t =>
    (t.p+" "+t.pr+" "+t.ot+" "+t.n+" "+t.a+" "+t.ped).toLowerCase().includes(f));
  return out;
}

function pintar(){
  const datos = visibles();
  cotas(datos);
  $("#cbEscala").textContent = semanas + " sem";

  const rejilla = $("#rejilla");
  rejilla.innerHTML = "";

  if(!datos.length){
    rejilla.innerHTML = '<div class="vacio"><div class="t">Sin tareas en este corte</div>' +
      '<p>Ajusta el filtro o incluye las cerradas.</p></div>';
    return;
  }

  let ini = new Date(Math.min(...datos.map(t=>+t._i), +hoy));
  ini = sumaDias(ini, -3);
  const finMax = new Date(Math.max(...datos.map(t=>+t._f)));
  let fin = sumaDias(ini, semanas*7);
  if(fin > finMax) fin = sumaDias(finMax, 3);
  const total = Math.max(14, difDias(ini, fin) + 1);
  const ancho = total * 26;

  rejilla.appendChild(cabecera(ini, total, ancho));

  const grupos = {};
  datos.forEach(t => {
    const k = vista === "persona" ? t.p : t.a;
    (grupos[k] = grupos[k] || []).push(t);
  });

  Object.keys(grupos)
    .sort((a,b)=>grupos[b].length - grupos[a].length || a.localeCompare(b,"es"))
    .forEach(k => rejilla.appendChild(fila(k, grupos[k], ini, total, ancho)));

  const off = difDias(ini, hoy);
  if(off >= 0 && off < total){
    const l = document.createElement("div");
    l.className = "hoy";
    l.style.left = (246 + off*26 + 13) + "px";
    rejilla.appendChild(l);
  }
}

function cabecera(ini, total, ancho){
  const f = document.createElement("div");
  f.className = "fila cabecera";
  const izq = document.createElement("div");
  izq.className = "izq";
  izq.innerHTML = '<div class="rol" style="margin:0">' +
    (vista === "persona"
      ? (POR_ETAPA ? "Etapa · carga abierta" : "Trabajador · carga abierta")
      : "Área · carga abierta") + '</div>';
  f.appendChild(izq);

  const pista = document.createElement("div");
  pista.className = "pista";
  pista.style.width = ancho + "px";

  const meses = document.createElement("div");
  meses.className = "meses";
  let d = new Date(ini), i = 0;
  while(i < total){
    const desde = i, m = d.getMonth();
    while(i < total && d.getMonth() === m){ d = sumaDias(d,1); i++; }
    const c = document.createElement("div");
    c.className = "mes";
    c.style.width = ((i - desde) * 26) + "px";
    const et = sumaDias(ini, desde);
    c.textContent = et.toLocaleDateString("es-CL",{month:"long"}) + " " + et.getFullYear();
    meses.appendChild(c);
  }
  pista.appendChild(meses);

  const dias = document.createElement("div");
  dias.className = "dias";
  for(let j=0;j<total;j++){
    const x = sumaDias(ini,j);
    const c = document.createElement("div");
    c.className = "dia" + (esFinde(x)?" finde":"") + (x.getDay()===1?" lunes":"");
    c.textContent = x.getDate();
    dias.appendChild(c);
  }
  pista.appendChild(dias);
  f.appendChild(pista);
  return f;
}

function fila(nombre, tareas, ini, total, ancho){
  const f = document.createElement("div");
  f.className = "fila";

  const izq = document.createElement("div");
  izq.className = "izq";
  izq.innerHTML = '<div class="nombre">' + nombre + '</div>';

  const abiertas = tareas.filter(t => t.e==="curso" || t.e==="vencida").length;
  const vencidas = tareas.filter(t => t.e==="vencida").length;

  const r = document.createElement("div");
  r.className = "rol";
  r.textContent = vista === "persona"
    ? [...new Set(tareas.map(t=>t.a))].join(" · ")
    : [...new Set(tareas.map(t=>t.p))].length + " personas";
  izq.appendChild(r);

  // La cota se llena respecto de la persona mas cargada del tablero
  const tope = Math.max(1, window.__topeCarga || 1);
  const cota = document.createElement("div");
  cota.className = "cotaLinea";
  const marca = document.createElement("i");
  marca.style.width = Math.min(100, Math.round(abiertas/tope*100)) + "%";
  if(vencidas) marca.className = "alerta";
  cota.appendChild(marca);
  izq.appendChild(cota);

  const c = document.createElement("div");
  c.className = "cifras";
  c.innerHTML = "<span>" + abiertas + " abiertas</span>" +
                (vencidas ? "<b>" + vencidas + " vencidas</b>" : "<span>—</span>");
  izq.appendChild(c);
  f.appendChild(izq);

  const pista = document.createElement("div");
  pista.className = "pista";
  pista.style.width = ancho + "px";

  const trama = document.createElement("div");
  trama.className = "trama";
  for(let j=0;j<total;j++){
    const x = sumaDias(ini,j);
    const cc = document.createElement("div");
    cc.className = "celda" + (esFinde(x)?" finde":"") + (x.getDay()===1?" lunes":"");
    trama.appendChild(cc);
  }
  pista.appendChild(trama);

  const carriles = [];
  tareas.sort((a,b)=>a._i-b._i).forEach(t => {
    const x0 = difDias(ini, t._i), x1 = difDias(ini, t._f);
    if(x1 < 0 || x0 > total) return;
    let k = carriles.findIndex(fin => fin < x0);
    if(k === -1){ carriles.push(x1); k = carriles.length-1; } else carriles[k] = x1;

    const col = COLOR[t.a] || "#54606E";
    const cerrada = t.e.startsWith("cerrada");
    const el = document.createElement(t.u ? "a" : "div");
    if(t.u){ el.href = t.u; el.target = "_blank"; el.rel = "noopener"; }
    el.className = "barra" + (t.e === "vencida" ? " vencida" : "") + (cerrada ? " cerrada" : "");
    el.style.left  = Math.max(0,x0)*26 + 2 + "px";
    el.style.width = Math.max(24,(Math.min(total,x1+1) - Math.max(0,x0))*26 - 4) + "px";
    el.style.top   = (6 + k*29) + "px";
    if(cerrada){ el.style.borderColor = col; el.style.color = col; }
    else { el.style.background = col; }
    el.innerHTML = (t.ot ? '<span class="ot">'+t.ot+'</span>' : '') +
                   '<span class="txt">'+t.n+'</span>';
    el.title = t.n + "\n" + t.pr + "\n" + t.p + " · " + t.a +
      (t.s ? " · " + t.s : "") + "\n" +
      t._i.toLocaleDateString("es-CL") + " → " + t._f.toLocaleDateString("es-CL") +
      (t.c ? "\nCompletada " + D(t.c).toLocaleDateString("es-CL") : "");
    pista.appendChild(el);
  });

  pista.style.minHeight = (12 + Math.max(1,carriles.length)*29) + "px";
  f.appendChild(pista);
  return f;
}

function cotas(datos){
  const abiertas = datos.filter(t => t.e==="curso" || t.e==="vencida");
  const personas = {};
  abiertas.forEach(t => personas[t.p] = (personas[t.p]||0) + 1);
  const nombres = Object.keys(personas);
  window.__topeCarga = nombres.length ? Math.max(...Object.values(personas)) : 1;

  $("#kPers").textContent = nombres.length;
  $("#kAbiertas").textContent = abiertas.length;
  $("#kMedia").textContent = nombres.length
    ? (abiertas.length/nombres.length).toFixed(1) : "0";
  $("#kVenc").textContent = abiertas.filter(t=>t.e==="vencida").length;
}

$("#gVista").addEventListener("click", e=>{
  const b = e.target.closest("button"); if(!b) return;
  vista = b.dataset.v;
  [...e.currentTarget.children].forEach(x=>x.setAttribute("aria-pressed", x===b));
  pintar();
});
$("#gRango").addEventListener("click", e=>{
  const b = e.target.closest("button"); if(!b) return;
  semanas = +b.dataset.s;
  [...e.currentTarget.children].forEach(x=>x.setAttribute("aria-pressed", x===b));
  pintar();
});
$("#gCerradas").addEventListener("click", e=>{
  const b = e.target.closest("button"); if(!b) return;
  verCerradas = b.dataset.c === "1";
  [...e.currentTarget.children].forEach(x=>x.setAttribute("aria-pressed", x===b));
  pintar();
});
let tec;
$("#buscar").addEventListener("input", e=>{
  clearTimeout(tec);
  tec = setTimeout(()=>{ filtro = e.target.value.trim(); pintar(); }, 200);
});

cotas(visibles());   // fija el tope de carga antes del primer dibujo
pintar();
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():
    carpeta = sys.argv[1] if len(sys.argv) > 1 else "data-taller"
    salida = sys.argv[2] if len(sys.argv) > 2 else "GANTT.HTML"

    if not os.path.isdir(carpeta):
        print(f"[ERROR] No existe la carpeta {os.path.abspath(carpeta)}")
        sys.exit(1)

    archivos = sorted(glob.glob(os.path.join(carpeta, "*.xlsx")))
    archivos = [f for f in archivos if not os.path.basename(f).startswith("~$")]
    print(f"Archivos .xlsx encontrados: {len(archivos)}")
    if not archivos:
        print("Nada que procesar. ¿Corrió antes exportar_taller.py?")
        sys.exit(1)

    hoy = datetime.date.today()
    desde = hoy - datetime.timedelta(weeks=SEMANAS_ATRAS)

    tareas, errores = [], []
    for f in archivos:
        try:
            tareas.extend(process_file(f, hoy, desde))
        except Exception as e:
            errores.append((os.path.basename(f), str(e)))

    if not tareas:
        print("No se encontró ninguna tarea con fechas de inicio o vencimiento.")
        print("Revisa que el export de Asana incluya Start Date y Due Date.")
        sys.exit(1)

    # ¿Hay responsables? Si los hay, las tareas sin asignar se descartan,
    # como corresponde a un Gantt por trabajador.
    con_resp = [t for t in tareas if t["p"]]
    aviso = ""
    if con_resp:
        descartadas = len(tareas) - len(con_resp)
        tareas = con_resp
        if descartadas:
            print(f"Descartadas {descartadas} tareas sin responsable asignado.")
    else:
        # Nadie asignado en todo el proyecto: en vez de fallar, se agrupa
        # por etapa. No es el Gantt por trabajador, y el HTML lo dice.
        print("=" * 62)
        print("AVISO: ninguna tarea del proyecto tiene responsable en Asana.")
        print("El Gantt se genera agrupado por ETAPA, no por persona.")
        print("Para el Gantt por trabajador hay que asignar las tareas en Asana.")
        print("=" * 62)
        for t in tareas:
            etapa = (t["s"] or "Sin etapa").strip().rstrip(":")
            t["p"] = etapa
        aviso = ("Ninguna tarea de este proyecto tiene responsable asignado en "
                 "Asana, así que las filas son etapas, no personas. En cuanto "
                 "se asignen responsables, el tablero pasa solo a mostrarlos.")

    ahora = datetime.datetime.now(ZoneInfo("America/Santiago"))
    fecha_corta = ahora.strftime("%d/%m %H:%M")

    html = (TEMPLATE
            .replace("__DATA__", json.dumps(tareas, ensure_ascii=False))
            .replace("__COLORES__", json.dumps(COLOR_AREA, ensure_ascii=False))
            .replace("__AVISO__", json.dumps(aviso, ensure_ascii=False))
            .replace("__FECHA_CORTA__", json.dumps(fecha_corta, ensure_ascii=False)))

    Path(salida).parent.mkdir(parents=True, exist_ok=True)
    Path(salida).write_text(html, encoding="utf-8")

    personas = sorted({t["p"] for t in tareas})
    abiertas = [t for t in tareas if t["e"] in ("curso", "vencida")]
    vencidas = [t for t in tareas if t["e"] == "vencida"]

    print(f"Tareas con responsable y fechas: {len(tareas)}")
    print(f"  abiertas: {len(abiertas)} · vencidas: {len(vencidas)}")
    print(f"Personas en el Gantt: {len(personas)}")
    for p in personas:
        n = len([t for t in abiertas if t["p"] == p])
        if n:
            print(f"  - {p}: {n} abiertas")
    for f, e in errores:
        print(f"  [ERROR] {f}: {e}")
    print(f"Gantt generado en: {salida}")


if __name__ == "__main__":
    main()
