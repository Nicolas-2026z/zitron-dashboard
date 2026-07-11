#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GENERA index.html - "Portafolio Proyectos Zitron"
Lee los .xlsx de cada proyecto (carpeta data_excels/) + el .xlsx del
board "Servicios y Mantencion" (carpeta data_servicios/) + el Excel de
fechas EXW (carpeta data_exw/) y genera el HTML del portafolio.

REGLAS DE CALCULO (avance nivel 1, sin cabeceras, ponderado por nivel 2/3)
---------------------------------------------------------------------------
- Nivel 1 (FASE)   = tarea sin "Parent task"
- Nivel 2          = tarea cuyo "Parent task" == nombre de una fase nivel1
- Nivel 3          = tarea cuyo "Parent task" == nombre de una tarea nivel2
- Avance FASE      = suma(completadas n2/n3) / suma(total n2/n3)
- Avance PROYECTO  = suma(d fases) / suma(t fases) * 100

REGLAS DE ESTADO / COLOR
---------------------------------------------------------------------------
- Color: 0-49% rojo, 50-75% amarillo patito (#EAB308), 76-100% verde
- "Completado" SOLO si fase Logistica/Despacho esta al 100%
- Fecha EXW: se lee del Excel de fabricacion, fecha maxima por pedido
- Descripcion del pedido: se lee de la tarea "Apertura pedido" (columna
  Notes/Notas del export de Asana) dentro de la fase "Kick Off meeting"

USO
---
  python3 generar_portafolio.py [data_excels] [data_servicios] [salida.html] [data_exw]

Por defecto:
  data_excels    = /mnt/user-data/uploads
  data_servicios = /mnt/user-data/uploads/servicios
  salida.html    = /mnt/user-data/outputs/index.html
  data_exw       = /mnt/user-data/uploads/exw  (opcional)
"""

import sys
import os
import glob
import json
import datetime
import warnings
import re as _re
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------
# CONFIGURACION EDITABLE
# ---------------------------------------------------------------------

COUNTRY_KEYWORDS = [
    ("colombia", "Colombia"),
    ("peru", "Peru"),
    ("mexico", "Mexico"), ("xemortiz", "Mexico"), ("mapim", "Mexico"),
    ("chile", "Chile"), ("gesvial", "Chile"), ("atacama", "Chile"),
    ("dominicana", "Rep. Dominicana"), ("calabresse", "Rep. Dominicana"),
    ("brasil", "Brasil"),
    ("nicaragua", "Nicaragua"), ("triton", "Nicaragua"),
]

WORKSPACE_GID = "402967058777498"
SERVICIOS_PROJECT_GID = "1213595645392940"

SECTION_COLORS = {
    "finalizado": "#6B7280",
    "finalizada": "#6B7280",
    "ejecucion": "#1D9E75",
    "ejecución": "#1D9E75",
    "proyecto": "#185FA5",
    "revisado": "#EAB308",
}
DEFAULT_SECTION_COLOR = "#378ADD"

PASSWORD = "zitron2026!"

# Nombre (normalizado) de la tarea cuya descripcion queremos mostrar en el
# modal de detalle del portafolio.
DESC_TASK_NAME = "apertura pedido"

# ---------------------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------------------

def _norm(s):
    s = str(s or "").lower().strip()
    repl = {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n"}
    for a, b in repl.items():
        s = s.replace(a, b)
    return s

def to_date(val):
    if val is None: return None
    if isinstance(val, datetime.datetime): return val.date()
    if isinstance(val, datetime.date): return val
    return None

def find_header_row(ws, max_scan=10):
    for r in range(1, max_scan + 1):
        values = [c.value for c in ws[r]]
        if "Task ID" in values and "Name" in values:
            return r, values
    return None, None

def country_for(project_name):
    n = _norm(project_name)
    for kw, country in COUNTRY_KEYWORDS:
        if kw in n:
            return country
    return "Sin pais"

# ---------------------------------------------------------------------
# LEER EXCEL DE FECHAS EXW
# Devuelve dict {pedido: "DD/MM/YYYY"} con la fecha maxima por pedido
# ---------------------------------------------------------------------

def leer_fechas_exw(path):
    """
    Lee el Excel de fabricacion y devuelve un dict
    {numero_pedido_str: fecha_str_dd/mm/yyyy}
    Toma la fecha maxima (mas tardia) cuando un pedido tiene varias OTs.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Buscar fila de encabezado: busca "PEDIDO" y fecha
    header_row = None
    col_pedido = col_fecha = None
    for r in range(1, 15):
        values = [str(c.value or "").strip() for c in ws[r]]
        for i, v in enumerate(values):
            vn = _norm(v)
            if "pedido" in vn:
                col_pedido = i
            if "entrega" in vn or "exw" in vn or "prevista" in vn:
                col_fecha = i
        if col_pedido is not None and col_fecha is not None:
            header_row = r
            break

    if header_row is None:
        print("  [AVISO] No se encontro encabezado en el Excel EXW")
        return {}

    fechas = {}
    pedido_actual = None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Leer pedido (puede estar vacio si es continuacion del anterior)
        raw_pedido = row[col_pedido] if col_pedido < len(row) else None
        raw_fecha  = row[col_fecha]  if col_fecha  < len(row) else None

        if raw_pedido not in (None, ""):
            p = str(raw_pedido).strip()
            # Solo acepta pedidos que parezcan numeros de 8 digitos
            if _re.match(r'^\d{7,8}$', p):
                pedido_actual = p

        if pedido_actual is None:
            continue

        # Parsear fecha
        fecha_date = None
        if raw_fecha not in (None, ""):
            if isinstance(raw_fecha, (datetime.datetime, datetime.date)):
                fecha_date = to_date(raw_fecha)
            else:
                s = str(raw_fecha).strip()
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        fecha_date = datetime.datetime.strptime(s, fmt).date()
                        break
                    except Exception:
                        pass

        if fecha_date is None:
            continue

        # Guardar la fecha maxima por pedido
        if pedido_actual not in fechas or fecha_date > fechas[pedido_actual]:
            fechas[pedido_actual] = fecha_date

    result = {k: v.strftime("%d/%m/%Y") for k, v in fechas.items()}
    print(f"  Fechas EXW leidas: {len(result)} pedidos")
    return result


# ---------------------------------------------------------------------
# LEER EXCEL DE FECHAS DE ENTREGA DEL PORTAFOLIO
# (exportado desde la vista de portafolio de Asana, no desde el excel de
# fabricacion). Devuelve {pedido: [(nombre_fila, "DD/MM/YYYY"), ...]}
# porque un mismo pedido puede tener varios proyectos (variantes A/B, OTs).
# ---------------------------------------------------------------------

# ---------------------------------------------------------------------
# LEER FECHAS DE ENTREGA DEL PORTAFOLIO
# (exportado desde la vista de portafolio de Asana). OJO: a diferencia de
# los proyectos, Asana exporta portafolios solo en CSV (no ofrece XLSX),
# asi que esta lectura soporta ambos formatos.
# Devuelve {pedido: [(nombre_fila, "DD/MM/YYYY"), ...]} porque un mismo
# pedido puede tener varios proyectos (variantes A/B, distintas OT).
# ---------------------------------------------------------------------

def _cargar_filas_tabla(path):
    """Devuelve una lista de filas (cada una lista de valores), leyendo
    .xlsx con openpyxl o .csv con el modulo csv, sin importar la extension."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        import csv as _csv
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(path, newline="", encoding=enc) as fh:
                    sample = fh.read(4096)
                    fh.seek(0)
                    try:
                        dialecto = _csv.Sniffer().sniff(sample, delimiters=",;\t")
                    except Exception:
                        dialecto = _csv.excel
                    return [row for row in _csv.reader(fh, dialecto)]
            except UnicodeDecodeError:
                continue
        return []
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        return [[c for c in row] for row in ws.iter_rows(values_only=True)]


def leer_fechas_entrega_portafolio(path):
    filas = _cargar_filas_tabla(path)

    header_idx = None
    col_name = col_fecha = None
    for ridx in range(min(15, len(filas))):
        values = [str(v or "").strip() for v in filas[ridx]]
        for i, v in enumerate(values):
            vn = _norm(v)
            if vn in ("name", "nombre", "project name", "nombre del proyecto"):
                col_name = i
            if "entrega" in vn or "vencimiento" in vn or "due date" in vn or vn == "due":
                col_fecha = i
        if col_name is not None and col_fecha is not None:
            header_idx = ridx
            break

    if header_idx is None:
        print("  [AVISO] No se encontro encabezado (Name + Fecha de entrega) en el archivo del portafolio")
        return {}

    resultado = {}
    for row in filas[header_idx + 1:]:
        raw_name = row[col_name] if col_name < len(row) else None
        raw_fecha = row[col_fecha] if col_fecha < len(row) else None
        if raw_name in (None, ""):
            continue
        nombre_fila = str(raw_name).strip()
        m = _re.search(r'(\d{7,8})', nombre_fila)
        if not m:
            continue
        pedido = m.group(1)

        fecha_date = None
        if raw_fecha not in (None, ""):
            if isinstance(raw_fecha, (datetime.datetime, datetime.date)):
                fecha_date = to_date(raw_fecha)
            else:
                s = str(raw_fecha).strip()
                # CSV de Asana suele traer fecha ISO (YYYY-MM-DD); XLSX puede
                # traer DD/MM/YYYY. Se prueban ambos formatos.
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        fecha_date = datetime.datetime.strptime(s, fmt).date()
                        break
                    except Exception:
                        pass
        if fecha_date is None:
            continue

        resultado.setdefault(pedido, []).append((nombre_fila, fecha_date.strftime("%d/%m/%Y")))

    total_filas = sum(len(v) for v in resultado.values())
    print(f"  Fechas de entrega (portafolio) leidas: {total_filas} filas, {len(resultado)} pedidos distintos")
    return resultado


def emparejar_fecha_entrega(nombre_proyecto, entrega_por_pedido):
    """
    Empareja un proyecto del dashboard con su fecha de entrega leida del
    portafolio. Un mismo numero de pedido puede repetirse en varios
    proyectos (variantes A/B, distintas OT); se desambigua por palabras en
    comun con el nombre de fila del portafolio. Si sigue habiendo empate y
    las fechas candidatas difieren, se prefiere dejarlo vacio antes que
    arriesgar una fecha incorrecta.
    """
    m = _re.search(r'(\d{7,8})', nombre_proyecto)
    if not m:
        return ""
    candidatas = entrega_por_pedido.get(m.group(1))
    if not candidatas:
        return ""
    if len(candidatas) == 1:
        return candidatas[0][1]

    def tokens(s):
        return {t for t in _re.split(r'[\s\-_/]+', _norm(s)) if len(t) >= 2 and not t.isdigit()}

    tok_proyecto = tokens(nombre_proyecto)
    mejor = None
    mejor_score = -1
    empate = False
    for nombre_fila, fecha in candidatas:
        score = len(tok_proyecto & tokens(nombre_fila))
        if score > mejor_score:
            mejor_score, mejor, empate = score, fecha, False
        elif score == mejor_score:
            empate = True

    if empate:
        fechas_unicas = {f for _, f in candidatas}
        if len(fechas_unicas) == 1:
            return fechas_unicas.pop()
        return ""
    return mejor


# ---------------------------------------------------------------------
# PROCESAR UN PROYECTO (xlsx) -> entrada del array P
# ---------------------------------------------------------------------

def process_project(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    project_name = Path(path).stem.strip()

    header_row, headers = find_header_row(ws)
    if header_row is None:
        return None

    col = {name: idx for idx, name in enumerate(headers) if name is not None}
    required = ["Name", "Parent task", "Completed At"]
    if not all(r in col for r in required):
        return None

    # Columna de descripcion/notas: Asana la exporta como "Notes" (o "Notas"
    # si el export quedo localizado). Es opcional: si no existe, desc queda "".
    # Deteccion tolerante a mayusculas/acentos/espacios extra en el encabezado.
    headers_norm = {_norm(h): idx for idx, h in enumerate(headers) if h is not None}
    col_notes = None
    for cand in ("notes", "notas", "description", "descripcion"):
        if cand in headers_norm:
            col_notes = headers_norm[cand]
            break

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = row[col["Name"]]
        if name in (None, ""):
            continue
        notes = ""
        if col_notes is not None and col_notes < len(row) and row[col_notes] not in (None, ""):
            notes = str(row[col_notes]).strip()
        rows.append({
            "name": str(name).strip(),
            "parent": (str(row[col["Parent task"]]).strip()
                       if row[col["Parent task"]] not in (None, "") else None),
            "completed": to_date(row[col["Completed At"]]) is not None,
            "notes": notes,
        })

    # --- Descripcion del pedido: buscar la tarea "Apertura pedido" en
    # cualquier nivel y tomar su columna de notas/descripcion. ---
    desc = ""
    for r in rows:
        if DESC_TASK_NAME in _norm(r["name"]):
            if r["notes"]:
                desc = r["notes"]
                break

    nivel1 = [r for r in rows if r["parent"] is None]
    nivel1_names = {r["name"] for r in nivel1}
    nivel2 = [r for r in rows if r["parent"] in nivel1_names]
    nivel2_names = {r["name"] for r in nivel2}
    nivel3 = [r for r in rows if r["parent"] in nivel2_names]

    n2_por_fase = {}
    for r in nivel2:
        n2_por_fase.setdefault(r["parent"], []).append(r)

    n3_por_n2 = {}
    for r in nivel3:
        n3_por_n2.setdefault(r["parent"], []).append(r)

    ph = []
    total_t = total_d = 0
    nombre_conteo = {}
    bodega_t = bodega_d = 0
    bodega_incluida = False

    for f in nivel1:
        hijos2 = n2_por_fase.get(f["name"], [])
        if not hijos2:
            continue
        fase_t = fase_d = 0
        for h2 in hijos2:
            hijos3 = n3_por_n2.get(h2["name"], [])
            if hijos3:
                fase_t += len(hijos3)
                fase_d += sum(1 for x in hijos3 if x["completed"])
            else:
                fase_t += 1
                fase_d += 1 if h2["completed"] else 0

        if "bodega" in f["name"].lower():
            bodega_t += fase_t; bodega_d += fase_d
            total_t  += fase_t; total_d  += fase_d
            if not bodega_incluida:
                ph.append(["__BODEGA__", 0, 0])
                bodega_incluida = True
            continue

        nombre_fase = f["name"]
        nombre_conteo[nombre_fase] = nombre_conteo.get(nombre_fase, 0) + 1
        if nombre_conteo[nombre_fase] > 1:
            sufijos = ["", " B", " C", " D", " E"]
            nombre_fase += sufijos[min(nombre_conteo[nombre_fase]-1, 4)]

        ph.append([nombre_fase, fase_t, fase_d])
        total_t += fase_t; total_d += fase_d

    for entry in ph:
        if entry[0] == "__BODEGA__":
            entry[0] = "Bodega"; entry[1] = bodega_t; entry[2] = bodega_d

    pct = round(total_d / total_t * 100) if total_t else 0

    return {
        "n": project_name,
        "p": country_for(project_name),
        "t": total_t,
        "d": total_d,
        "pct": pct,
        "ph": ph,
        "exw": "",   # se rellena despues con el cruce
        "entrega": "",  # fecha de entrega del portafolio, se rellena despues
        "desc": desc,  # descripcion de la tarea "Apertura pedido"
    }


# ---------------------------------------------------------------------
# PROCESAR SERVICIOS
# ---------------------------------------------------------------------

def process_servicios(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row, headers = find_header_row(ws)
    if header_row is None:
        return None

    col = {name: idx for idx, name in enumerate(headers) if name is not None}

    def getcol(*names):
        for n in names:
            if n in col: return col[n]
        return None

    c_name     = col.get("Name")
    c_section  = getcol("Section/Column", "Section")
    c_assignee = col.get("Assignee")
    c_due      = col.get("Due Date")
    c_completed= col.get("Completed At")
    c_taskid   = col.get("Task ID")
    c_country  = getcol("Pais", "País", "Country")
    c_priority = getcol("Prioridad", "Priority")

    total = completadas = 0
    tasks_activas = []
    secciones = {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if c_name is None or row[c_name] in (None, ""):
            continue
        total += 1
        name    = str(row[c_name]).strip()
        section = str(row[c_section]).strip() if c_section is not None and row[c_section] else "Sin seccion"
        secciones[section] = secciones.get(section, 0) + 1

        if c_completed is not None and to_date(row[c_completed]) is not None:
            completadas += 1
            continue

        assignee = str(row[c_assignee]).strip() if c_assignee is not None and row[c_assignee] else "(sin asignar)"
        due      = to_date(row[c_due]) if c_due is not None else None
        country  = str(row[c_country]).strip() if c_country is not None and row[c_country] else ""
        priority = str(row[c_priority]).strip() if c_priority is not None and row[c_priority] else ""
        taskid   = str(row[c_taskid]).strip() if c_taskid is not None and row[c_taskid] else ""
        url      = (f"https://app.asana.com/1/{WORKSPACE_GID}/project/{SERVICIOS_PROJECT_GID}/task/{taskid}"
                    if taskid else "")

        tasks_activas.append({
            "name": name, "section": section, "assignee": assignee,
            "due": due.strftime("%Y-%m-%d") if due else "",
            "pais": country, "prioridad": priority, "url": url,
        })

    return {
        "total": total,
        "en_curso": total - completadas,
        "finalizadas": completadas,
        "avance": round(completadas / total * 100) if total else 0,
        "secciones": secciones,
        "tasks": tasks_activas,
    }


# ---------------------------------------------------------------------
# TEMPLATE HTML
# ---------------------------------------------------------------------

TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Portafolio Proyectos — Zitron</title>
<style>
:root{--hdr:#1A3A5C;--grn:#1D9E75;--blu:#378ADD;--amb:#EAB308;--red:#E24B4A;--bg:#F3F6FA;--border:#E2E8F0;--text:#1A202C;--sub:#6B7280;}
*{box-sizing:border-box;margin:0;padding:0;}
body{font-family:'Segoe UI',system-ui,Arial,sans-serif;background:var(--bg);color:var(--text);}
.header{background:var(--hdr);color:#fff;padding:0 24px;height:58px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;box-shadow:0 2px 10px rgba(0,0,0,.25);}
.hleft{display:flex;align-items:center;gap:12px;}
.hlogo{width:34px;height:34px;background:rgba(255,255,255,.15);border-radius:7px;display:flex;align-items:center;justify-content:center;font-size:18px;font-weight:800;}
.htitle{font-size:15px;font-weight:700;}
.hsub{font-size:10px;opacity:.6;margin-top:1px;}
.hright{font-size:11px;opacity:.7;text-align:right;line-height:1.7;}
.kpis{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:20px;}
.kpi{background:#fff;border:1px solid var(--border);border-radius:10px;padding:14px 16px;flex:1;min-width:110px;text-align:center;}
.kl{font-size:10px;color:var(--sub);margin-bottom:5px;}
.kv{font-size:26px;font-weight:700;}
.ctrl{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:12px;}
.ctrl input,.ctrl select{padding:8px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px;background:#fff;outline:none;}
.ctrl input{width:220px;}
.pills{display:flex;gap:7px;flex-wrap:wrap;margin-bottom:16px;}
.pill{padding:5px 15px;border-radius:20px;border:1.5px solid var(--border);background:#fff;font-size:12px;cursor:pointer;color:var(--sub);}
.pill.act{background:var(--hdr);color:#fff;border-color:var(--hdr);}
.plist{display:flex;flex-direction:column;gap:8px;}
.pcard{background:#fff;border:1px solid var(--border);border-radius:12px;padding:12px 16px;cursor:pointer;transition:border-color .15s;}
.pcard:hover{border-color:#94A3B8;}
.ptop{display:flex;align-items:center;gap:8px;margin-bottom:9px;flex-wrap:wrap;}
.pname{flex:1;min-width:0;font-size:13px;font-weight:600;white-space:normal;}
.pcountry{font-size:11px;color:var(--sub);padding:2px 8px;border-radius:12px;background:var(--bg);}
.pbadge{font-size:11px;font-weight:700;padding:3px 10px;border-radius:12px;}
.ppct{font-size:15px;font-weight:700;min-width:42px;text-align:right;}
.pchev{font-size:12px;color:var(--sub);transition:transform .2s;}
.pcard.open .pchev{transform:rotate(180deg);}
.pbar-row{display:flex;align-items:center;gap:10px;}
.pbar-bg{height:7px;background:#E8EDF3;border-radius:4px;flex:1;overflow:hidden;}
.pbar-fill{height:100%;border-radius:4px;}
.pcnt{font-size:11px;color:var(--sub);white-space:nowrap;min-width:95px;text-align:right;}
.pdet{display:none;border-top:1px solid var(--border);margin-top:11px;padding-top:11px;}
.phrow{display:flex;align-items:center;gap:8px;padding:3px 0;}
.phname{font-size:11px;color:var(--sub);min-width:160px;}
.phbg{height:4px;background:#E8EDF3;border-radius:2px;flex:1;overflow:hidden;}
.phfill{height:100%;border-radius:2px;}
.phpct{font-size:11px;font-weight:600;min-width:34px;text-align:right;}
.phct{font-size:10px;color:var(--sub);min-width:40px;text-align:right;}
.footer{text-align:center;font-size:11px;color:#94A3B8;padding:24px 0 16px;}
.sec-title{font-size:15px;font-weight:700;color:var(--hdr);margin:32px 0 6px;display:flex;align-items:center;gap:8px;}
.sec-title::before{content:'';display:inline-block;width:4px;height:18px;background:var(--hdr);border-radius:2px;}
.sec-sub{font-size:11px;color:var(--sub);margin-bottom:14px;}
.svc-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:16px;}
@media(max-width:640px){.svc-grid{grid-template-columns:1fr;}}
.svc-card{background:#fff;border:1px solid var(--border);border-radius:12px;padding:16px 18px;}
.svc-card h3{font-size:11px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.05em;margin-bottom:12px;}
.sbar-row{display:flex;align-items:center;gap:8px;margin-bottom:7px;}
.sbar-label{font-size:12px;color:var(--text);min-width:150px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.sbar-track{flex:1;background:#E8EDF3;border-radius:3px;height:8px;overflow:hidden;}
.sbar-fill{height:100%;border-radius:3px;}
.sbar-count{font-size:12px;font-weight:600;color:var(--text);min-width:22px;text-align:right;}
.svc-kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:14px;}
@media(max-width:640px){.svc-kpis{grid-template-columns:repeat(2,1fr);}}
.skpi{background:#fff;border:1px solid var(--border);border-radius:10px;padding:12px 14px;text-align:center;}
.skpi .kl{font-size:10px;color:var(--sub);margin-bottom:4px;}
.skpi .kv{font-size:22px;font-weight:700;}
.svc-tasks{background:#fff;border:1px solid var(--border);border-radius:12px;padding:16px 18px;margin-bottom:10px;}
.svc-tasks h3{font-size:11px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.05em;margin-bottom:10px;}
.stabs{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:10px;}
.stab{font-size:12px;padding:4px 12px;border-radius:20px;border:1.5px solid var(--border);cursor:pointer;background:#fff;color:var(--sub);}
.stab.act{background:var(--hdr);color:#fff;border-color:var(--hdr);}
.strow{display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid var(--border);font-size:12px;}
.strow:last-child{border-bottom:none;}
.stname{flex:1;color:var(--text);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.stname a{color:var(--blu);text-decoration:none;}
.stname a:hover{text-decoration:underline;}
.stdate{font-size:11px;color:var(--sub);white-space:nowrap;}
.stdate.ov{color:var(--red);font-weight:600;}
.stassign{font-size:11px;color:var(--sub);white-space:nowrap;max-width:120px;overflow:hidden;text-overflow:ellipsis;}
.stpill{font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px;white-space:nowrap;}
.p-alta{background:#FEE2E2;color:#991B1B;}
.p-media{background:#FEF3C7;color:#92400E;}
.p-baja{background:#D1FAE5;color:#065F46;}
.exw-tag{font-size:10px;color:#1A3A5C;padding:2px 7px;border-radius:8px;background:#EEF4FF;border:1px solid #C7D9F5;white-space:nowrap;font-weight:600;}
.exw-tag.vencida{background:#FEE2E2;border-color:#FECACA;color:#991B1B;}
.entrega-tag{font-size:10px;color:#065F46;padding:2px 7px;border-radius:8px;background:#ECFDF5;border:1px solid #A7F3D0;white-space:nowrap;font-weight:600;}
.entrega-tag.vencida{background:#FEE2E2;border-color:#FECACA;color:#991B1B;}
.info-btn{font-size:11px;font-weight:700;padding:3px 9px;border-radius:12px;border:1.5px solid #C7D9F5;background:#EEF4FF;color:#1A3A5C;cursor:pointer;white-space:nowrap;display:inline-flex;align-items:center;gap:4px;}
.info-btn:hover{background:#DCE9FB;}
.modal-overlay{display:none;position:fixed;inset:0;background:rgba(15,25,40,.55);z-index:2000;align-items:flex-start;justify-content:center;padding:40px 16px;overflow-y:auto;}
.modal-overlay.show{display:flex;}
.modal-box{background:#fff;border-radius:14px;max-width:640px;width:100%;box-shadow:0 24px 60px rgba(0,0,0,.35);overflow:hidden;}
.modal-head{background:var(--hdr);color:#fff;padding:16px 22px;display:flex;align-items:center;justify-content:space-between;gap:12px;}
.modal-head-title{font-size:14px;font-weight:700;}
.modal-head-sub{font-size:11px;opacity:.7;margin-top:2px;}
.modal-close{background:rgba(255,255,255,.15);border:none;color:#fff;width:28px;height:28px;border-radius:8px;font-size:16px;cursor:pointer;flex-shrink:0;}
.modal-close:hover{background:rgba(255,255,255,.28);}
.modal-body{padding:20px 22px;max-height:65vh;overflow-y:auto;}
.modal-label{font-size:10px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px;}
.modal-desc{font-size:13px;line-height:1.65;color:var(--text);white-space:pre-wrap;word-wrap:break-word;}
.modal-desc.empty{color:var(--sub);font-style:italic;}
.modal-foot{padding:12px 22px 18px;font-size:11px;color:var(--sub);border-top:1px solid var(--border);}
</style>
</head>
<body>
<div id="loginScreen" style="display:flex;min-height:100vh;align-items:center;justify-content:center;background:linear-gradient(135deg,#0F2540 0%,#1A3A5C 60%,#244E82 100%);position:fixed;top:0;left:0;width:100%;z-index:9999">
  <div style="background:white;border-radius:16px;padding:44px 40px;width:360px;text-align:center;box-shadow:0 24px 64px rgba(0,0,0,.35)">
    <div style="width:64px;height:64px;background:#1A3A5C;border-radius:14px;display:flex;align-items:center;justify-content:center;font-size:30px;font-weight:800;color:white;margin:0 auto 18px">Z</div>
    <div style="font-size:20px;font-weight:700;color:#1A3A5C;margin-bottom:4px">Portafolio Proyectos</div>
    <div style="font-size:13px;color:#6B7280;margin-bottom:28px">Zitron Acceso privado</div>
    <input id="pwInput" type="password" placeholder="Contrasena" maxlength="40"
      style="width:100%;padding:12px 16px;border:2px solid #E2E8F0;border-radius:8px;font-size:16px;text-align:center;letter-spacing:4px;margin-bottom:14px;outline:none;box-sizing:border-box"
      onkeydown="if(event.key==='Enter')chkPw()">
    <button onclick="chkPw()" style="width:100%;padding:13px;background:#1A3A5C;color:white;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer">Ingresar</button>
    <div id="loginErr" style="color:#E24B4A;font-size:13px;margin-top:10px;min-height:18px"></div>
    <div style="font-size:10px;color:#CBD5E0;margin-top:16px">Solicita la contrasena al administrador</div>
  </div>
</div>
<script>
function chkPw(){var v=document.getElementById('pwInput').value;if(v==='__PASSWORD__'){document.getElementById('loginScreen').style.display='none';sessionStorage.setItem('zpw','ok');}else{document.getElementById('loginErr').textContent='Contrasena incorrecta';document.getElementById('pwInput').value='';document.getElementById('pwInput').focus();}}
if(sessionStorage.getItem('zpw')==='ok'){document.addEventListener('DOMContentLoaded',function(){document.getElementById('loginScreen').style.display='none';});}
window.addEventListener('load',function(){if(sessionStorage.getItem('zpw')!=='ok'){document.getElementById('pwInput').focus();}});
</script>
<header class="header">
  <div class="hleft"><div class="hlogo">Z</div><div><div class="htitle">Portafolio Consolidado Zitron</div><div class="hsub">% avance nivel 1 sin cabeceras ponderado por nivel 2/3</div></div></div>
  <div class="hright"><div>Actualizado: __FECHA__ __HORA__</div><div style="color:#86EFAC" id="hdrTotal">__TOTAL__ proyectos</div></div>
</header>
<div style="max-width:1080px;margin:0 auto;padding:22px 18px;">
  <div class="kpis" id="kpis"></div>
  <div class="ctrl">
    <input type="text" id="srch" placeholder="Buscar proyecto o pais..." oninput="render()">
    <select id="srt" onchange="render()">
      <option value="pd">Mayor avance</option>
      <option value="pa">Menor avance</option>
      <option value="na">Nombre A-Z</option>
      <option value="ct">Pais</option>
      <option value="exwa">EXW mas antigua primero</option>
      <option value="exwz">EXW mas reciente primero</option>
    </select>
  </div>
  <div class="pills">
    <button class="pill act" onclick="setF('all',this)">Todos</button>
    <button class="pill" onclick="setF('mx',this)">Mexico</button>
    <button class="pill" onclick="setF('pe',this)">Peru</button>
    <button class="pill" onclick="setF('cl',this)">Chile</button>
    <button class="pill" onclick="setF('co',this)">Colombia</button>
    <button class="pill" onclick="setF('ot',this)">Otros</button>
    <button class="pill" onclick="setF('des',this)">Despachados</button>
    <button class="pill" onclick="setF('pend',this)">Pend. despacho</button>
  </div>
  <div class="plist" id="plist"></div>
  <div class="sec-title">Tablero Servicios y Mantencion</div>
  <div class="sec-sub">Carpeta Servicios · __SVC_TOTAL__ tareas totales · Actualizado __FECHA__</div>
  <div class="svc-kpis">
    <div class="skpi"><div class="kl">Total tareas</div><div class="kv" style="color:#378ADD">__SVC_TOTAL__</div></div>
    <div class="skpi"><div class="kl">En curso</div><div class="kv" style="color:#EAB308">__SVC_CURSO__</div></div>
    <div class="skpi"><div class="kl">Finalizadas</div><div class="kv" style="color:#1D9E75">__SVC_FIN__</div></div>
    <div class="skpi"><div class="kl">Avance global</div><div class="kv" style="color:#1A3A5C">__SVC_AVANCE__%</div></div>
  </div>
  <div class="svc-grid">
    <div class="svc-card"><h3>Tareas por seccion</h3><div id="svc-sections"></div></div>
    <div class="svc-card"><h3>Distribucion por pais</h3><div id="svc-countries"></div></div>
  </div>
  <div class="svc-tasks"><h3>Tareas activas</h3><div class="stabs" id="svc-tabs"></div><div id="svc-tasklist"></div></div>
  <div class="footer" id="footer"></div>
</div>
<div class="modal-overlay" id="descModal" onclick="if(event.target===this)closeDescModal()">
  <div class="modal-box">
    <div class="modal-head">
      <div>
        <div class="modal-head-title" id="descModalTitle">Detalle del pedido</div>
        <div class="modal-head-sub" id="descModalSub">Tarea Asana: Apertura pedido</div>
      </div>
      <button class="modal-close" onclick="closeDescModal()">&#10005;</button>
    </div>
    <div class="modal-body">
      <div class="modal-label">Descripcion</div>
      <div class="modal-desc" id="descModalBody"></div>
    </div>
    <div class="modal-foot" id="descModalFoot"></div>
  </div>
</div>
<script>
var PM={mx:"Mexico",pe:"Peru",cl:"Chile",co:"Colombia"};
var filt="all";
var TODAY = new Date('__FECHA_ISO__');

var P = __P_JSON__;

function col(p){return p>=76?"#1D9E75":p>=50?"#EAB308":"#E24B4A";}

function isDespachado(x){
  var lp=x.ph.find(function(ph){var n=ph[0].toLowerCase();return n.indexOf("logist")>=0||n.indexOf("despacho")>=0;});
  return !!(lp&&lp[1]>0&&lp[1]===lp[2]);
}

function bi(despachado){if(despachado)return{t:"Completado",c:"#1D9E75"};return{t:"En progreso",c:"#EAB308"};}

// Convierte "DD/MM/YYYY" -> Date para comparar/ordenar
function parseExw(s){
  if(!s)return null;
  var p=s.split('/');
  if(p.length!==3)return null;
  return new Date(parseInt(p[2]),parseInt(p[1])-1,parseInt(p[0]));
}

function getF(){
  var q=(document.getElementById("srch").value||"").toLowerCase();
  var s=document.getElementById("srt").value;
  var d=P.filter(function(x){
    if(q&&x.n.toLowerCase().indexOf(q)<0&&x.p.toLowerCase().indexOf(q)<0)return false;
    if(PM[filt])return x.p===PM[filt];
    if(filt==="ot")return Object.values(PM).indexOf(x.p)<0;
    if(filt==="des")return isDespachado(x);
    if(filt==="pend")return!isDespachado(x);
    return true;
  });
  if(s==="pd")d.sort(function(a,b){return b.pct-a.pct;});
  else if(s==="pa")d.sort(function(a,b){return a.pct-b.pct;});
  else if(s==="ct")d.sort(function(a,b){return a.p.localeCompare(b.p);});
  else if(s==="exwa"){
    d.sort(function(a,b){
      var da=parseExw(a.exw),db=parseExw(b.exw);
      if(!da&&!db)return 0;
      if(!da)return 1;
      if(!db)return -1;
      return da-db;
    });
  } else if(s==="exwz"){
    d.sort(function(a,b){
      var da=parseExw(a.exw),db=parseExw(b.exw);
      if(!da&&!db)return 0;
      if(!da)return 1;
      if(!db)return -1;
      return db-da;
    });
  } else d.sort(function(a,b){return a.n.localeCompare(b.n);});
  return d;
}

function rKPIs(){
  var avg=P.length?Math.round(P.reduce(function(s,x){return s+x.pct;},0)/P.length):0;
  var ts=P.reduce(function(s,x){return s+x.t;},0);
  var ds=P.reduce(function(s,x){return s+x.d;},0);
  document.getElementById("hdrTotal").textContent=P.length+" proyectos";
  var completados=P.filter(function(x){return isDespachado(x);}).length;
  var vencidos=P.filter(function(x){
    var d=parseExw(x.exw);
    return d&&d<TODAY&&!isDespachado(x);
  }).length;
  var ks=[
    {l:"Total proyectos",v:P.length,c:"#378ADD"},
    {l:"Avance promedio",v:avg+"%",c:col(avg)},
    {l:"Subtareas completadas",v:ds+"/"+ts,c:"#1D9E75"},
    {l:"Completados (despachados)",v:completados,c:"#1D9E75"},
    {l:"En progreso",v:P.length-completados,c:"#EAB308"},
    {l:"EXW vencida",v:vencidos,c:"#E24B4A"},
    {l:"Despachados",v:completados,c:"#1D9E75"}
  ];
  document.getElementById("kpis").innerHTML=ks.map(function(k){return '<div class="kpi"><div class="kl">'+k.l+'</div><div class="kv" style="color:'+k.c+'">'+k.v+'</div></div>';}).join("");
}

function toggle(i){
  var card=document.getElementById("c"+i),det=document.getElementById("d"+i);
  if(!det)return;
  var open=det.style.display==="block";
  det.style.display=open?"none":"block";
  if(card)card.classList.toggle("open",!open);
}

var lastRenderData=[];
function render(){
  rKPIs();
  var data=getF();
  lastRenderData=data;
  document.getElementById("plist").innerHTML=data.map(function(x,i){
    var c=col(x.pct);
    var desp=isDespachado(x);
    var b=bi(desp);
    var phs=x.ph.map(function(ph){
      var fn=ph[0],ft=ph[1],fd=ph[2];
      var fp=ft>0?Math.round(fd/ft*100):0;
      return '<div class="phrow"><span class="phname">'+fn+'</span><div class="phbg"><div class="phfill" style="width:'+fp+'%;background:'+col(fp)+';"></div></div><span class="phpct" style="color:'+col(fp)+'">'+fp+'%</span><span class="phct">'+fd+'/'+ft+'</span></div>';
    }).join("");
    var db=desp
      ?'<span style="font-size:11px;font-weight:700;padding:3px 10px;border-radius:12px;background:#D1FAE5;color:#065F46;border:1.5px solid #1D9E75;white-space:nowrap">✓ Despachado</span>'
      :'<span style="font-size:11px;font-weight:700;padding:3px 10px;border-radius:12px;background:#FEE2E2;color:#E24B4A;border:1.5px solid #E24B4A;white-space:nowrap">Pend. despacho</span>';
    // Tag EXW
    var exwTag='';
    if(x.exw){
      var exwDate=parseExw(x.exw);
      var vencida=exwDate&&exwDate<TODAY&&!desp;
      exwTag='<span class="exw-tag'+(vencida?' vencida':'')+'">📦 EXW '+(vencida?'⚠ ':'')+x.exw+'</span>';
    }
    // Tag Fecha de entrega (desde el portafolio de Asana)
    var entregaTag='';
    if(x.entrega){
      var entregaDate=parseExw(x.entrega);
      var entregaVencida=entregaDate&&entregaDate<TODAY&&!desp;
      entregaTag='<span class="entrega-tag'+(entregaVencida?' vencida':'')+'">🗓️ Entrega '+(entregaVencida?'⚠ ':'')+x.entrega+'</span>';
    }
    return '<div class="pcard" id="c'+i+'" onclick="toggle('+i+')">'
      +'<div class="ptop"><span class="pname" title="'+x.n+'">'+x.n+'</span>'
      +'<span class="pcountry">'+x.p+'</span>'
      +'<span class="pbadge" style="background:'+b.c+'22;color:'+b.c+'">'+b.t+'</span>'
      +exwTag
      +entregaTag
      +'<button class="info-btn" onclick="event.stopPropagation();showDesc('+i+')" title="Ver detalle del pedido (Apertura pedido)">&#9432; Detalle</button>'
      +'<span class="ppct" style="color:'+c+'">'+x.pct+'%</span>'
      +db+'<span class="pchev">&#9662;</span></div>'
      +'<div class="pbar-row"><div class="pbar-bg"><div class="pbar-fill" style="width:'+x.pct+'%;background:'+c+';"></div></div>'
      +'<span class="pcnt">'+x.d+'/'+x.t+' subtareas</span></div>'
      +'<div class="pdet" id="d'+i+'">'+phs+'</div></div>';
  }).join("")||'<div style="text-align:center;padding:40px;color:#6B7280">Sin resultados</div>';
  document.getElementById("footer").textContent=data.length+" proyectos · "+P.length+" total · __FECHA__ · Consolidado Asana";
}
function setF(f,btn){filt=f;document.querySelectorAll(".pill").forEach(function(b){b.classList.remove("act");});btn.classList.add("act");render();}

function escHtml(s){return String(s==null?"":s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");}

function showDesc(i){
  var x=lastRenderData[i];
  if(!x)return;
  document.getElementById("descModalTitle").textContent=x.n;
  document.getElementById("descModalSub").textContent="Tarea Asana: Kick Off meeting \u2192 Apertura pedido";
  var body=document.getElementById("descModalBody");
  var txt=(x.desc||"").trim();
  if(txt){
    body.classList.remove("empty");
    body.textContent=txt;
  } else {
    body.classList.add("empty");
    body.textContent="Sin descripcion registrada en Asana para la tarea \"Apertura pedido\" de este proyecto.";
  }
  document.getElementById("descModalFoot").textContent=x.p+" \u00b7 "+x.pct+"% de avance \u00b7 "+x.d+"/"+x.t+" subtareas";
  document.getElementById("descModal").classList.add("show");
}
function closeDescModal(){document.getElementById("descModal").classList.remove("show");}
document.addEventListener("keydown",function(e){if(e.key==="Escape")closeDescModal();});

render();

var SVC_TASKS = __SVC_JSON__;
var SVC_SECTION_COUNTS = __SVC_SECTIONS_JSON__;
var SVC_TODAY=new Date('__FECHA_ISO__');
var SVC_SECTIONS=Object.keys(SVC_SECTION_COUNTS);
var SVC_SECTION = SVC_SECTIONS.length?SVC_SECTIONS[0]:"";
function sbarHtml(lbl,v,mx,clr){var p=mx>0?Math.round(v/mx*100):0;return '<div class="sbar-row"><span class="sbar-label">'+lbl+'</span><div class="sbar-track"><div class="sbar-fill" style="width:'+p+'%;background:'+clr+';"></div></div><span class="sbar-count">'+v+'</span></div>';}
var SECTION_COLOR_MAP=__SECTION_COLORS_JSON__;
function colorForSection(name){var k=name.toLowerCase();for(var key in SECTION_COLOR_MAP){if(k.indexOf(key)>=0)return SECTION_COLOR_MAP[key];}return "__SECTION_DEFAULT__";}
function renderSvcSections(){var entries=Object.entries(SVC_SECTION_COUNTS).sort(function(a,b){return b[1]-a[1];});var mx=entries.length?entries[0][1]:1;document.getElementById("svc-sections").innerHTML=entries.map(function(r){return sbarHtml(r[0],r[1],mx,colorForSection(r[0]));}).join("");}
function renderSvcCountries(){var cc={};SVC_TASKS.forEach(function(t){var p=t.pais||"Sin pais";cc[p]=(cc[p]||0)+1;});var sorted=Object.entries(cc).sort(function(a,b){return b[1]-a[1];}).slice(0,7);if(!sorted.length){document.getElementById("svc-countries").innerHTML='<i style="font-size:12px;color:#6B7280">Sin datos</i>';return;}var mx=sorted[0][1];document.getElementById("svc-countries").innerHTML=sorted.map(function(r){return sbarHtml(r[0],r[1],mx,"#378ADD");}).join("");}
function renderSvcTabs(){document.getElementById("svc-tabs").innerHTML=SVC_SECTIONS.map(function(s){var cnt=SVC_TASKS.filter(function(t){return t.section===s;}).length;return '<button class="stab'+(s===SVC_SECTION?" act":"")+'" onclick="setSvcSec(\''+s+'\',this)">'+s+' ('+cnt+')</button>';}).join("");}
function setSvcSec(s,btn){SVC_SECTION=s;document.querySelectorAll(".stab").forEach(function(b){b.classList.remove("act");});btn.classList.add("act");renderSvcList();}
function renderSvcList(){var po={Alta:0,Media:1,Baja:2,"":3};var tasks=SVC_TASKS.filter(function(t){return t.section===SVC_SECTION;}).sort(function(a,b){return(po[a.prioridad]||3)-(po[b.prioridad]||3);});if(!tasks.length){document.getElementById("svc-tasklist").innerHTML='<p style="font-size:12px;color:#6B7280;padding:8px 0">Sin tareas.</p>';return;}document.getElementById("svc-tasklist").innerHTML=tasks.map(function(t){var over=t.due&&new Date(t.due+"T00:00:00")<SVC_TODAY;var dt=t.due?new Date(t.due+"T00:00:00").toLocaleDateString("es-CL",{day:"2-digit",month:"short"}):"";var pill=t.prioridad?'<span class="stpill p-'+t.prioridad.toLowerCase()+'">'+t.prioridad+"</span>":'';var nameHtml=t.url?('<a href="'+t.url+'" target="_blank">'+t.name+'</a>'):t.name;return '<div class="strow"><span class="stname">'+nameHtml+'</span>'+pill+'<span class="stassign">'+t.assignee+'</span><span class="stdate'+(over?" ov":"")+'">'+dt+(over?" ⚠":"")+"</span></div>";}).join("");}
renderSvcSections();renderSvcCountries();renderSvcTabs();renderSvcList();
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():
    dir_proyectos = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads"
    dir_servicios = sys.argv[2] if len(sys.argv) > 2 else "/mnt/user-data/uploads/servicios"
    salida        = sys.argv[3] if len(sys.argv) > 3 else "/mnt/user-data/outputs/index.html"
    dir_exw       = sys.argv[4] if len(sys.argv) > 4 else dir_proyectos

    # --- Leer fechas EXW ---
    fechas_exw = {}
    exw_files = sorted([f for f in glob.glob(os.path.join(dir_exw, "*.xlsx")) if any(k in os.path.basename(f).lower() for k in ["exw","fabricacion","entrega","reporte"]) and "portafolio" not in os.path.basename(f).lower()])
    exw_files = [f for f in exw_files if not os.path.basename(f).startswith("~$")]
    if exw_files:
        fechas_exw = leer_fechas_exw(exw_files[0])
    else:
        print(f"  [AVISO] No se encontro Excel EXW en {dir_exw}")

    # --- Leer fechas de entrega del portafolio (columna "Fecha de entrega"
    # de la vista de portafolio de Asana, exportada por asana_exporter.py) ---
    entrega_por_pedido = {}
    entrega_files = sorted(
        [f for f in glob.glob(os.path.join(dir_proyectos, "*.xlsx")) + glob.glob(os.path.join(dir_proyectos, "*.csv"))
         if "portafolio" in os.path.basename(f).lower()]
    )
    entrega_files = [f for f in entrega_files if not os.path.basename(f).startswith("~$")]
    if entrega_files:
        entrega_por_pedido = leer_fechas_entrega_portafolio(entrega_files[0])
    else:
        print(f"  [AVISO] No se encontro Excel de fechas de entrega del portafolio en {dir_proyectos}")

    # --- Leer proyectos ---
    files = sorted(glob.glob(os.path.join(dir_proyectos, "*.xlsx")))
    files = [f for f in files if not os.path.basename(f).startswith("~$") and not any(k in os.path.basename(f).lower() for k in ["exw","fabricacion","entrega","reporte","portafolio"])]
    if not files:
        print(f"No se encontraron .xlsx en {dir_proyectos}")
        import sys as _sys; _sys.exit(1)

    def _clave(path):
        base = os.path.splitext(os.path.basename(path))[0]
        base = _re.sub(r'\s*\(\d+\)\s*$', '', base)
        base = _re.sub(r'[\s_-]+\d{2}-\d{2}-\d{4}.*$', '', base)
        base = _re.sub(r'[\s_-]+\d{5,}.*$', '', base)
        return base.strip().lower()

    mejor = {}
    for f in files:
        clave = _clave(f)
        mtime = os.path.getmtime(f)
        if clave not in mejor or mtime > mejor[clave][0]:
            mejor[clave] = (mtime, f)
    files = sorted([v[1] for v in mejor.values()])
    print(f"Archivos unicos a procesar: {len(files)}")

    P = []
    errores = []
    for f in files:
        try:
            r = process_project(f)
            if r is None:
                errores.append((f, "encabezados/columnas no reconocidas"))
                continue
            # Cruzar con fecha EXW: buscar numero de 8 digitos en el nombre
            m = _re.search(r'(\d{8})', r["n"])
            if m:
                r["exw"] = fechas_exw.get(m.group(1), "")
            r["entrega"] = emparejar_fecha_entrega(r["n"], entrega_por_pedido)
            P.append(r)
        except Exception as e:
            errores.append((f, str(e)))

    if not P:
        print("No se pudo procesar ningun proyecto.")
        for f, e in errores:
            print(f"  - {f}: {e}")
        import sys as _sys; _sys.exit(1)

    # --- Servicios ---
    svc = {"total":0,"en_curso":0,"finalizadas":0,"avance":0,"secciones":{},"tasks":[]}
    svc_files = sorted(glob.glob(os.path.join(dir_servicios, "*.xlsx"))) if os.path.isdir(dir_servicios) else []
    svc_files = [f for f in svc_files if not os.path.basename(f).startswith("~$")]
    if svc_files:
        r = process_servicios(svc_files[0])
        if r: svc = r
    else:
        print(f"  [AVISO] no se encontro xlsx de servicios en {dir_servicios}")

    now_chile = datetime.datetime.now(ZoneInfo("America/Santiago"))
    fecha_str = now_chile.strftime("%d/%m/%Y")
    hora_str  = now_chile.strftime("%H:%M hrs")
    fecha_iso = now_chile.strftime("%Y-%m-%d")

    html_out = TEMPLATE
    html_out = html_out.replace("__PASSWORD__",          PASSWORD)
    html_out = html_out.replace("__FECHA_ISO__",         fecha_iso)
    html_out = html_out.replace("__FECHA__",             fecha_str)
    html_out = html_out.replace("__HORA__",              hora_str)
    html_out = html_out.replace("__TOTAL__",             str(len(P)))
    html_out = html_out.replace("__P_JSON__",            json.dumps(P, ensure_ascii=False))
    html_out = html_out.replace("__SVC_JSON__",          json.dumps(svc["tasks"], ensure_ascii=False))
    html_out = html_out.replace("__SVC_SECTIONS_JSON__", json.dumps(svc["secciones"], ensure_ascii=False))
    html_out = html_out.replace("__SVC_TOTAL__",         str(svc["total"]))
    html_out = html_out.replace("__SVC_CURSO__",         str(svc["en_curso"]))
    html_out = html_out.replace("__SVC_FIN__",           str(svc["finalizadas"]))
    html_out = html_out.replace("__SVC_AVANCE__",        str(svc["avance"]))
    html_out = html_out.replace("__SECTION_COLORS_JSON__", json.dumps(SECTION_COLORS, ensure_ascii=False))
    html_out = html_out.replace("__SECTION_DEFAULT__",   DEFAULT_SECTION_COLOR)

    out_dir = os.path.dirname(salida)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(salida, "w", encoding="utf-8") as fh:
        fh.write(html_out)

    print(f"Proyectos procesados: {len(P)} / {len(files)}")
    for f, e in errores:
        print(f"  [ERROR] {os.path.basename(f)}: {e}")
    print(f"Servicios: {svc['total']} tareas ({len(svc['tasks'])} activas)")
    print(f"HTML generado en: {salida}")

    # -------------------------------------------------------------
    # DIAGNOSTICO: descripcion de "Apertura pedido" por proyecto.
    # Se imprime siempre en el log (aparece en el log de GitHub
    # Actions del paso "Generar index.html") para poder ver, sin
    # entrar al sitio, que proyectos trajeron o no su descripcion.
    # -------------------------------------------------------------
    con_desc = [x["n"] for x in P if x["desc"].strip()]
    sin_desc = [x["n"] for x in P if not x["desc"].strip()]

    print("\n" + "=" * 60)
    print(f"  DIAGNOSTICO 'Apertura pedido': {len(con_desc)}/{len(P)} proyectos con descripcion")
    print("=" * 60)
    if sin_desc:
        print(f"  Sin descripcion ({len(sin_desc)}):")
        for n in sin_desc:
            print(f"    - {n}")

    # Deteccion de descripciones duplicadas entre proyectos DISTINTOS
    # (senal de que en Asana se copio el texto de un pedido a otro sin editarlo).
    from collections import defaultdict as _dd
    por_desc = _dd(list)
    for x in P:
        if x["desc"].strip():
            por_desc[x["desc"]].append(x["n"])
    duplicados = {d: ns for d, ns in por_desc.items() if len(ns) > 1}
    if duplicados:
        print(f"\n  [ALERTA] Descripciones IDENTICAS en mas de un proyecto (revisar en Asana):")
        for _, nombres in duplicados.items():
            print(f"    - Compartida por {len(nombres)} proyectos: {nombres}")
    print("=" * 60)


if __name__ == "__main__":
    main()
