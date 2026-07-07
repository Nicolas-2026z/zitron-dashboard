#!/usr/bin/env python3
"""
Genera KPI.html - "Dashboard Desempeño de Equipos"
basado en exports de Asana (.xlsx), replicando la estructura de las
capturas: tarjetas globales, Resumen por área, Detalle por persona,
Detalle de tareas, con semáforo en días hábiles de Chile.

REGLAS
------
- "Tarea nivel 2"  = filas cuyo campo 'Parent task' NO está vacío
                     (subtareas de una tarea padre).
- Semáforo:
    VERDE = completada en la fecha de entrega o antes (días hábiles Chile)
    ROJO  = todo lo demás:
              - completada 1+ día hábil después de la fecha de entrega
              - o no completada y la fecha de entrega ya pasó (1+ día hábil)
    (tareas en curso, sin vencer aún -> estado "En curso", no entran al %)

- Área del equipo: se determina por assignee (diccionario editable
  ASSIGNEE_AREA) y, si no está mapeado, por palabras clave en
  'Section/Column' (ver SECTION_AREA_KEYWORDS).

USO
---
  python3 generar_kpi.py <carpeta_con_excels> <archivo_salida_html>

Por defecto:
  carpeta_con_excels  = /mnt/user-data/uploads
  archivo_salida_html = /mnt/user-data/outputs/KPI.html
"""

import sys
from pathlib import Path
from zoneinfo import ZoneInfo
import os
import glob
import json
import datetime
import warnings
import openpyxl

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------
# CONFIGURACIÓN EDITABLE
# ---------------------------------------------------------------------

FERIADOS_CHILE = {
    datetime.date(2025, 1, 1), datetime.date(2025, 4, 18), datetime.date(2025, 4, 19),
    datetime.date(2025, 5, 1), datetime.date(2025, 5, 21), datetime.date(2025, 6, 20),
    datetime.date(2025, 6, 29), datetime.date(2025, 7, 16), datetime.date(2025, 8, 15),
    datetime.date(2025, 9, 18), datetime.date(2025, 9, 19), datetime.date(2025, 10, 12),
    datetime.date(2025, 10, 31), datetime.date(2025, 11, 1), datetime.date(2025, 12, 8),
    datetime.date(2025, 12, 25),
    datetime.date(2026, 1, 1), datetime.date(2026, 4, 3), datetime.date(2026, 4, 4),
    datetime.date(2026, 5, 1), datetime.date(2026, 5, 21), datetime.date(2026, 6, 20),
    datetime.date(2026, 6, 29), datetime.date(2026, 7, 16), datetime.date(2026, 8, 15),
    datetime.date(2026, 9, 18), datetime.date(2026, 9, 19), datetime.date(2026, 10, 12),
    datetime.date(2026, 10, 31), datetime.date(2026, 11, 1), datetime.date(2026, 12, 8),
    datetime.date(2026, 12, 25),
}

ASSIGNEE_AREA = {
    # Servicios
    "Ignacio García": "Servicios",
    "Ignacio Garcia": "Servicios",
    "IGNACIO GARCIA MARTINEZ": "Servicios",
    "Ignacio Garcia Martinez": "Servicios",
    "Sergio Saavedra": "Servicios",
    "sergio saavedra fernandez": "Servicios",
    "Sergio Saavedra Fernandez": "Servicios",
    # Equipo Proyecto
    "Carlos Pérez": "Equipo Proyecto",
    "Carlos Perez": "Equipo Proyecto",
    "Francisca Ramos": "Equipo Proyecto",
    "Francisca Alejandra Ramos Aravales": "Equipo Proyecto",
    "Sergio de la Fuente": "Equipo Proyecto",
    "Sergio de la Fuente Fernandez": "Equipo Proyecto",
    "David Blazquez": "Equipo Proyecto",
    "DAVID BLAZQUEZ": "Equipo Proyecto",
    # Producción
    "Benjamín Bustos": "Producción",
    "benjamin.bustos@zitron.com": "Producción",
    "Benjamin Bustos": "Producción",
    "Benjamin Umaña": "Producción",
    "Hugo Isla": "Producción",
    "hugo isla martinez": "Producción",
    "Nicolás Mol": "Producción",
    "Nicolas Mol": "Producción",
    "Nicolás López": "Producción",
    "Nicolas Lopez": "Producción",
    # Compras
    "Eliana": "Compras",
    "Yerlia": "Compras",
    "Yerlia Ayleen Castillo Diaz": "Compras",
    "Rose": "Compras",
    "Rosemary Singh": "Compras",
    # Ingeniería
    "Hernán Gutierrez": "Ingeniería",
    "Hernan Roberto Gutierrez Barrientos": "Ingeniería",
    "Hernan Gutierrez": "Ingeniería",
    "Francisca González": "Ingeniería",
    "Francisca González Cornejo": "Ingeniería",
    "Francisca Gonzalez": "Ingeniería",
    "Gonzalo Davila": "Ingeniería",
    "Gonzalo Dávila": "Ingeniería",
    "Gabriel Venegas": "Ingeniería",
    "Gabriel Venega": "Ingeniería",
    # Bodega
    "Víctor Muñoz": "Bodega",
    "Victor Muñoz": "Bodega",
    "Victor Munoz": "Bodega",
    # Logística
    "Karin Pinto": "Logística",
}

SECTION_AREA_KEYWORDS = [
    ("produccion", "Producción"), ("producción", "Producción"),
    ("ingenier", "Ingeniería"),
    ("compra", "Compras"),
    ("bodega", "Bodega"),
    ("logist", "Logística"), ("logíst", "Logística"),
]

# ---------------------------------------------------------------------
# UTILIDADES DE FECHAS (días hábiles Chile)
# ---------------------------------------------------------------------

def es_habil(d: datetime.date) -> bool:
    return d.weekday() < 5 and d not in FERIADOS_CHILE


def dias_habiles_entre(d1: datetime.date, d2: datetime.date) -> int:
    if d1 == d2:
        return 0
    a, b = (d1, d2) if d2 > d1 else (d2, d1)
    count = 0
    cur = a + datetime.timedelta(days=1)
    while cur <= b:
        if es_habil(cur):
            count += 1
        cur += datetime.timedelta(days=1)
    return count if d2 > d1 else -count


def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


# ---------------------------------------------------------------------
# LECTURA DE EXCELS
# ---------------------------------------------------------------------

def find_header_row(ws, max_scan=10):
    for r in range(1, max_scan + 1):
        values = [c.value for c in ws[r]]
        if "Task ID" in values and "Name" in values:
            return r, values
    return None, None


def _norm(s):
    s = str(s or "").lower().strip()
    repl = {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n"}
    for a, b in repl.items():
        s = s.replace(a, b)
    return s


def area_for(assignee, section):
    sec = (section or "").lower()
    for kw, area in SECTION_AREA_KEYWORDS:
        if kw in sec:
            return area
    if assignee:
        a_norm = _norm(assignee)
        a_words = set(a_norm.split())
        for k, v in ASSIGNEE_AREA.items():
            k_norm = _norm(k)
            if k_norm == a_norm or k_norm in a_norm:
                return v
            k_words = set(k_norm.split())
            if k_words and k_words.issubset(a_words):
                return v
    return "Equipo Proyecto"


def process_file(path, today):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    project_name = Path(path).stem.strip()

    header_row, headers = find_header_row(ws)
    if header_row is None:
        return None

    col = {name: idx for idx, name in enumerate(headers) if name is not None}
    required = ["Name", "Due Date", "Completed At", "Parent task"]
    if not all(r in col for r in required):
        return None

    tasks = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[col["Name"]] in (None, ""):
            continue

        parent = row[col["Parent task"]]
        if parent in (None, ""):
            continue

        name = str(row[col["Name"]]).strip()
        section = row[col.get("Section/Column")] if "Section/Column" in col else ""
        assignee = row[col.get("Assignee")] if "Assignee" in col else ""

        if assignee and str(assignee).strip().lower() in ("nicolás", "nicolas"):
            continue

        # Excluir tareas de la sección "Cierre de proyecto"
        if section and "cierre de proyecto" in str(section).strip().lower():
            continue

        # Excluir tareas llamadas "Costos"
        if name and str(name).strip().lower() == "costos":
            continue

        start = to_date(row[col["Start Date"]]) if "Start Date" in col else None
        due = to_date(row[col["Due Date"]])
        completed = to_date(row[col["Completed At"]])

        area = area_for(assignee, section)

        if completed is not None:
            if due is not None:
                diff = dias_habiles_entre(due, completed)
                if diff <= 0:
                    estado = "verde"
                    estado_plazo = "a tiempo" if diff == 0 else f"{abs(diff)}d antes"
                else:
                    estado = "rojo"
                    estado_plazo = f"{diff}d tarde"
            else:
                estado = "verde"
                estado_plazo = "sin fecha"
            estado_general = "Completada"
        else:
            if due is not None:
                diff = dias_habiles_entre(due, today)
                if diff > 0:
                    estado = "rojo"
                    estado_plazo = f"{diff}d vencida"
                    estado_general = "Vencida"
                else:
                    estado = "encurso"
                    estado_plazo = f"{abs(diff)}d rest." if diff < 0 else "vence hoy"
                    estado_general = "En curso"
            else:
                estado = "encurso"
                estado_plazo = "sin fecha"
                estado_general = "En curso"

        blocked_by = row[col["Blocked By (Dependencies)"]] if "Blocked By (Dependencies)" in col else ""
        blocking = row[col["Blocking (Dependencies)"]] if "Blocking (Dependencies)" in col else ""
        task_id = row[col["Task ID"]] if "Task ID" in col else None

        tasks.append({
            "name": name,
            "section": section or "",
            "assignee": assignee or "(sin asignar)",
            "area": area,
            "project": project_name,
            "start_iso": start.isoformat() if start else (due.isoformat() if due else None),
            "due_iso": due.isoformat() if due else None,
            "completed": completed.isoformat() if completed else None,
            "start_fmt": start.strftime("%d/%m/%y") if start else (due.strftime("%d/%m/%y") if due else "--"),
            "due_fmt": due.strftime("%d/%m/%y") if due else "--",
            "completed_fmt": completed.strftime("%d/%m/%y") if completed else "--",
            "duracion_prevista": dias_habiles_entre(start, due) if start and due else (1 if due else None),
            "blocked_by": [x.strip() for x in str(blocked_by).split(",") if x.strip()] if blocked_by else [],
            "blocking": [x.strip() for x in str(blocking).split(",") if x.strip()] if blocking else [],
            "bloqueada": bool(blocked_by and str(blocked_by).strip()),
            "asana_url": f"https://app.asana.com/0/0/{int(task_id)}/f" if task_id else None,
            "atraso_dias": abs(dias_habiles_entre(due, completed)) if completed and due and completed > due else (abs(dias_habiles_entre(due, today)) if due and today > due and not completed else 0),
            "dias_plazo": dias_habiles_entre(due, completed) if completed and due else None,
            "estado": estado,
            "estado_plazo": estado_plazo,
            "estado_general": estado_general,
        })

    return {"project": project_name, "file": os.path.basename(path), "tasks": tasks}


TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<title>Dashboard Desempeño de Equipos</title>
<style>
  :root {
    --verde: #1e8e3e; --verde-bg: #e6f4ea;
    --rojo: #b3261e; --rojo-bg: #fce8e6;
    --amarillo: #eab308; --amarillo-bg: #fef9c3;
    --azul: #1a73e8;
    --gris: #80868b; --gris-bg: #f1f3f4;
    --bg: #f1efe9; --card: #ffffff; --texto: #202124; --borde: #e6e3dc;
  }
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, "Segoe UI", Roboto, Arial, sans-serif;
    background: var(--bg); color: var(--texto); margin: 0; padding: 28px;
    max-width: 1280px; margin-left: auto; margin-right: auto;
  }
  h1 { font-size: 26px; margin: 0 0 4px 0; }
  .subt { color: #5f6368; font-size: 14px; margin: 2px 0; }
  .subt2 { color: #9aa0a6; font-size: 12px; margin: 2px 0 18px 0; }

  .cards-globales { display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 14px; margin-bottom: 18px; }
  .card-g {
    background: var(--card); border: 1px solid var(--borde); border-radius: 10px;
    padding: 12px 16px;
  }
  .card-g .lbl { font-size: 12px; color: #80868b; text-transform: uppercase; letter-spacing: .04em; }
  .card-g .val { font-size: 26px; font-weight: 700; margin: 4px 0; }
  .card-g .sub { font-size: 11.5px; color: #9aa0a6; line-height: 1.4; }
  .card-g .val.verde { color: var(--verde); }
  .card-g .val.rojo { color: var(--rojo); }

  .tabs { display: flex; gap: 28px; border-bottom: 1px solid var(--borde); margin-bottom: 18px; }
  .tab { padding: 10px 2px; cursor: pointer; font-size: 14px; color: #5f6368; border-bottom: 2px solid transparent; }
  .tab.active { color: var(--texto); font-weight: 600; border-bottom-color: var(--texto); }
  .tabcontent { display: none; }
  .tabcontent.active { display: block; }

  .area-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(190px, 1fr)); gap: 14px; }
  .area-card {
    background: var(--card); border: 1px solid var(--borde); border-radius: 10px;
    padding: 14px; width: auto;
  }
  .area-card .titulo { font-size: 14px; font-weight: 600; margin-bottom: 6px; }
  .area-card .pct { font-size: 28px; font-weight: 700; margin: 4px 0; }
  .area-card .pct.verde { color: var(--verde); }
  .area-card .pct.rojo { color: var(--rojo); }
  .area-card .pct.amarillo { color: var(--amarillo); }
  .area-card .meta { font-size: 12px; color: #5f6368; margin-top: 6px; line-height: 1.6; }
  .area-card .prom-box {
    margin-top: 10px;
    border-top: 1px solid var(--borde);
    padding-top: 8px;
    display: flex;
    align-items: center;
    gap: 8px;
  }
  .area-card .prom-val {
    font-size: 18px;
    font-weight: 700;
    color: var(--texto);
  }
  .area-card .prom-lbl {
    font-size: 11px;
    color: #9aa0a6;
    line-height: 1.3;
  }
  .dot { width: 9px; height: 9px; border-radius: 50%; display: inline-block; margin-right: 3px; }
  .dot.verde { background: var(--verde); } .dot.rojo { background: var(--rojo); } .dot.gris { background: var(--gris); }

  table { width: 100%; border-collapse: collapse; background: var(--card); border-radius: 10px; overflow: hidden; }
  thead th {
    text-align: left; background: #fafafa; border-bottom: 2px solid var(--borde);
    padding: 9px 10px; font-size: 12px; color: #5f6368; text-transform: uppercase; letter-spacing: .03em;
  }
  tbody td { padding: 8px 10px; border-bottom: 1px solid var(--borde); font-size: 13px; }
  tbody tr:hover { background: #fafafa; }
  .pill { padding: 2px 8px; border-radius: 10px; font-size: 12px; font-weight: 600; white-space: nowrap; }
  .pill.verde { background: var(--verde-bg); color: var(--verde); }
  .pill.rojo { background: var(--rojo-bg); color: var(--rojo); }
  .pill.amarillo { background: var(--amarillo-bg); color: var(--amarillo); }
  .estado-txt.rojo { color: var(--rojo); font-weight: 600; }
  .estado-txt.verde { color: var(--verde); font-weight: 600; }
  .estado-txt.encurso { color: var(--azul); font-weight: 600; }

  .filtros { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 14px; }
  .filtros select {
    padding: 6px 10px; border: 1px solid var(--borde); border-radius: 6px; background: var(--card); font-size: 13px;
  }
  .resumen-linea { font-size: 13px; color: #5f6368; margin-bottom: 10px; }
  .resumen-linea b.rojo { color: var(--rojo); } .resumen-linea b.azul { color: var(--azul); } .resumen-linea b.verde { color: var(--verde); }

  .leyenda { margin-top: 18px; font-size: 12px; color: #80868b; display: flex; gap: 18px; flex-wrap: wrap; align-items: center; }
  .area-pills { display:flex; gap:8px; flex-wrap:wrap; margin-bottom:14px; }
  .area-pill {
    padding: 6px 14px; border-radius: 16px; background: var(--card); border: 1px solid var(--borde);
    font-size: 13px; cursor: pointer;
  }
  .area-pill.active { background: var(--texto); color: #fff; border-color: var(--texto); }

  .cascada-chain { background: var(--card); border: 1px solid var(--borde); border-radius: 10px; margin-bottom: 16px; overflow: hidden; }
  .cascada-chain .chain-header { padding: 10px 16px; background: #fafafa; border-bottom: 1px solid var(--borde); font-size: 13px; font-weight: 600; color: #5f6368; }
  .cascada-node { display: flex; align-items: flex-start; padding: 10px 16px; border-bottom: 1px solid var(--borde); gap: 12px; }
  .cascada-node:last-child { border-bottom: none; }
  .cascada-arrow { color: #9aa0a6; font-size: 18px; padding-top: 2px; min-width: 20px; }
  .cascada-task { flex: 1; }
  .cascada-task .task-name { font-weight: 600; font-size: 13px; margin-bottom: 3px; }
  .cascada-task .task-meta { font-size: 12px; color: #5f6368; display: flex; gap: 16px; flex-wrap: wrap; }
  .cascada-task .task-meta span { display: flex; align-items: center; gap: 4px; }
  .badge-atraso { background: var(--rojo-bg); color: var(--rojo); padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 700; }
  .badge-ok { background: var(--verde-bg); color: var(--verde); padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 700; }
  .badge-encurso { background: #e8f0fe; color: var(--azul); padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 700; }

  #gate {
    position: fixed; inset: 0; background: var(--bg); z-index: 9999;
    display: flex; align-items: center; justify-content: center;
  }
  #gate.hidden { display: none; }
  #gate .box {
    background: var(--card); border: 1px solid var(--borde); border-radius: 14px;
    padding: 40px 44px; text-align: center; box-shadow: 0 4px 20px rgba(0,0,0,0.06);
    width: 320px;
  }
  #gate .box .icono { font-size: 32px; margin-bottom: 8px; }
  #gate .box h2 { margin: 0 0 4px 0; font-size: 18px; font-weight: 700; }
  #gate .box .sub { color: #9aa0a6; font-size: 13px; margin-bottom: 18px; }
  #gate .box input {
    padding: 11px 14px; border: 1px solid var(--borde); border-radius: 8px;
    font-size: 14px; width: 100%; margin-bottom: 12px; box-sizing: border-box;
  }
  #gate .box button {
    padding: 11px 18px; border: none; border-radius: 8px; background: var(--texto);
    color: #fff; font-size: 14px; font-weight: 600; cursor: pointer; width: 100%;
  }
  #gate .box .error { color: var(--rojo); font-size: 13px; margin-top: 8px; height: 16px; }
  #app { display: none; }
  #app.show { display: block; }

  .selector-multi { position: relative; display: inline-block; margin-bottom: 18px; }
  .selector-multi .selector-btn {
    display: inline-flex; align-items: center; gap: 8px;
    background: var(--card); border: 1px solid var(--borde); border-radius: 8px;
    padding: 8px 14px; font-size: 14px; font-weight: 600; cursor: pointer;
  }
  .selector-multi .panel {
    display: none; position: absolute; top: 110%; left: 0; z-index: 100;
    background: var(--card); border: 1px solid var(--borde); border-radius: 8px;
    box-shadow: 0 4px 16px rgba(0,0,0,0.08); padding: 8px 0;
    max-height: 320px; overflow-y: auto; min-width: 320px;
  }
  .selector-multi.open .panel { display: block; }
  .selector-multi .panel label {
    display: flex; align-items: center; gap: 8px; padding: 6px 14px;
    font-size: 13px; font-weight: 400; cursor: pointer; white-space: nowrap;
  }
  .selector-multi .panel label:hover { background: #fafafa; }
  .selector-multi .panel .acciones {
    display: flex; gap: 10px; padding: 6px 14px 10px 14px; border-bottom: 1px solid var(--borde);
    margin-bottom: 4px; font-size: 12px;
  }
  .selector-multi .panel .acciones a { color: var(--azul); cursor: pointer; text-decoration: none; }

  /* Columna prom cierre en tabla personas */
  .prom-cierre-cell {
    font-weight: 700;
    color: var(--texto);
    white-space: nowrap;
  }
  .prom-cierre-cell span {
    font-size: 11px;
    font-weight: 400;
    color: #9aa0a6;
    display: block;
  }
</style>
</head>
<body>

<div id="gate">
  <div class="box">
    <div class="icono">📊</div>
    <h2>Dashboard Desempeño</h2>
    <div class="sub">Ingresa la clave para continuar</div>
    <input type="password" id="gatePass" placeholder="Contraseña" onkeydown="if(event.key==='Enter')checkPass()">
    <button onclick="checkPass()">Ingresar</button>
    <div class="error" id="gateError"></div>
  </div>
</div>

<div id="app">

<h1>Dashboard Desempeño de Equipos</h1>
<div class="subt">Indicador ejecutivo basado en tareas Asana — nivel 2 únicamente (subtareas, sin cabeceras de sección)</div>
<div class="subt2">Días calculados en días hábiles Chile · Última actualización: __FECHA__ · <span id="horaActual"></span></div>

<div class="selector-multi" id="selectorMulti">
  <div class="selector-btn" onclick="toggleSelector()">📁 Proyectos: <span id="selectorLabel"></span> ▾</div>
  <div class="panel" id="selectorPanel">
    <div class="acciones">
      <a onclick="seleccionarTodos(true)">Marcar todos</a>
      <a onclick="seleccionarTodos(false)">Desmarcar todos</a>
    </div>
    <div id="selectorOpciones"></div>
  </div>
</div>

<div id="globales"></div>

<div class="tabs">
  <div class="tab active" onclick="cambiarTab('area')">Resumen por área</div>
  <div class="tab" onclick="cambiarTab('persona')">Detalle por persona</div>
  <div class="tab" onclick="cambiarTab('tareas')">📄 Detalle tareas</div>
  <div class="tab" onclick="cambiarTab('cascada')">🔗 Cascada de atrasos</div>
</div>

<div class="tabcontent active" id="tab-area">
  <div class="area-grid" id="areaGrid"></div>
</div>

<div class="tabcontent" id="tab-persona">
  <div class="area-pills" id="areaPills"></div>
  <table>
    <thead><tr>
      <th></th>
      <th>Persona</th>
      <th>Área</th>
      <th>Tareas</th>
      <th>Compl.</th>
      <th>% Compl.</th>
      <th>⏱ Prom. plazo</th>
      <th>Cerradas en tiempo</th>
      <th>Cerradas fuera de tiempo</th>
      <th>Abiertas en tiempo</th>
      <th>Abiertas fuera de tiempo</th>
    </tr></thead>
    <tbody id="personaBody"></tbody>
  </table>
  <div style="margin-top:10px;font-size:11px;color:#9aa0a6;">
    ⏱ Prom. plazo = promedio de días hábiles entre fecha de vencimiento y fecha de completado. Verde = anticipado · Rojo = con atraso
  </div>
</div>

<div class="tabcontent" id="tab-tareas">
  <div class="filtros">
    <select id="fUsuario" onchange="renderTareas()"><option value="">Usuario: Todos</option></select>
    <select id="fArea" onchange="renderTareas()"><option value="">Área: Todas</option></select>
    <select id="fSeccion" onchange="renderTareas()"><option value="">Sección: Todas</option></select>
    <select id="fProyecto" onchange="renderTareas()"><option value="">Proyecto: Todos</option></select>
    <select id="fEstado" onchange="renderTareas()">
      <option value="">Estado: Todos</option>
      <option value="Completada">Completada</option>
      <option value="Vencida">Vencida</option>
      <option value="En curso">En curso</option>
    </select>
    <select id="fBloqueada" onchange="renderTareas()">
      <option value="">Bloqueo: Todas</option>
      <option value="si">Solo bloqueadas</option>
      <option value="no">Solo liberadas</option>
    </select>
  </div>
  <div class="resumen-linea" id="resumenTareas"></div>
  <table>
    <thead><tr><th>Tarea</th><th>Proyecto</th><th>Sección</th><th>Usuario</th><th>Área</th><th>Inicio</th><th>Vence</th><th>Completó</th><th>Duración prevista</th><th style="cursor:pointer;white-space:nowrap;" onclick="toggleSortPlazo()">Estado plazo ↕<span id="sortPlazoIcon"></span></th><th>Estado</th><th>Bloqueo</th><th>Bloqueada por</th><th>Link</th></tr></thead>
    <tbody id="tareasBody"></tbody>
  </table>
</div>

<div class="tabcontent" id="tab-cascada">
  <div class="filtros" style="align-items:flex-start;flex-wrap:wrap;gap:10px;">
    <select id="fCascadaEstado" onchange="renderCascada()">
      <option value="">Todas las cadenas</option>
      <option value="con_atraso">Solo con atraso</option>
    </select>
    <div class="selector-multi" id="selectorSeccionesCascada">
      <div class="selector-btn" onclick="toggleSelectorSecciones()">🏷️ Etapas: <span id="labelSeccionesCascada">Todas</span> ▾</div>
      <div class="panel" id="panelSeccionesCascada">
        <div class="acciones">
          <a onclick="seleccionarTodasSecciones(true)">Marcar todas</a>
          <a onclick="seleccionarTodasSecciones(false)">Desmarcar todas</a>
        </div>
        <label><input type="checkbox" value="kick_off" onchange="renderCascada()"> Kick off</label>
        <label><input type="checkbox" value="ingenieria" onchange="renderCascada()"> Ingeniería</label>
        <label><input type="checkbox" value="propuesta" onchange="renderCascada()"> Propuesta compras</label>
        <label><input type="checkbox" value="compras" onchange="renderCascada()"> Compras</label>
        <label><input type="checkbox" value="ingenieria_diseno" onchange="renderCascada()"> Ingeniería y diseño</label>
        <label><input type="checkbox" value="bodega" onchange="renderCascada()"> Bodega</label>
        <label><input type="checkbox" value="montaje" onchange="renderCascada()"> Montaje</label>
        <label><input type="checkbox" value="logistica" onchange="renderCascada()"> Logística</label>
        <label><input type="checkbox" value="cierre" onchange="renderCascada()"> Cierre proyecto</label>
      </div>
    </div>
  </div>
  <div id="cascadaContainer"></div>
</div>

<div class="leyenda">
  <span><span class="dot verde"></span> Verde: cerrada en fecha o antes</span>
  <span><span class="dot rojo"></span> Rojo: 1+ día hábil de atraso</span>
  <span style="margin-left:auto;">Solo tareas nivel 2 (con tarea padre) · días hábiles Chile</span>
</div>

<script>
// Mostrar hora actual de Chile
(function() {
  function actualizarHora() {
    const ahora = new Date();
    const opciones = { timeZone: 'America/Santiago', hour: '2-digit', minute: '2-digit', hour12: false };
    const hora = ahora.toLocaleTimeString('es-CL', opciones);
    const fecha = ahora.toLocaleDateString('es-CL', { timeZone: 'America/Santiago', weekday: 'long', day: '2-digit', month: '2-digit', year: 'numeric' });
    const el = document.getElementById('horaActual');
    if (el) el.textContent = '· Hora actual: ' + fecha + ' ' + hora + ' hrs';
  }
  document.addEventListener('DOMContentLoaded', function() {
    actualizarHora();
    setInterval(actualizarHora, 60000);
  });
})();

const CLAVE = "zitron2026!";
function checkPass() {
  const val = document.getElementById('gatePass').value;
  if (val === CLAVE) {
    sessionStorage.setItem('zitron_ok', '1');
    document.getElementById('gate').classList.add('hidden');
    document.getElementById('app').classList.add('show');
  } else {
    document.getElementById('gateError').textContent = 'Contraseña incorrecta';
  }
}
if (sessionStorage.getItem('zitron_ok') === '1') {
  document.getElementById('gate').classList.add('hidden');
  document.getElementById('app').classList.add('show');
}

const DATA = __DATA__;
const PROYECTOS = Object.keys(DATA);
let seleccionados = new Set(PROYECTOS);
let areaFiltroPersona = "Todos";

function cambiarTab(tab) {
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.tabcontent').forEach(t => t.classList.remove('active'));
  const idx = tab === 'area' ? 1 : (tab === 'persona' ? 2 : (tab === 'tareas' ? 3 : 4));
  document.querySelector(`.tab:nth-child(${idx})`).classList.add('active');
  document.getElementById('tab-' + tab).classList.add('active');
  if (tab === 'cascada') renderCascada();
}

function toggleSelectorSecciones() {
  document.getElementById('selectorSeccionesCascada').classList.toggle('open');
}
document.addEventListener('click', function(e) {
  const sel = document.getElementById('selectorSeccionesCascada');
  if (sel && !sel.contains(e.target)) sel.classList.remove('open');
});

function seleccionarTodasSecciones(valor) {
  document.querySelectorAll('#panelSeccionesCascada input[type=checkbox]').forEach(cb => cb.checked = valor);
  actualizarLabelSecciones();
  renderCascada();
}

function actualizarLabelSecciones() {
  const checks = document.querySelectorAll('#panelSeccionesCascada input[type=checkbox]:checked');
  const total = document.querySelectorAll('#panelSeccionesCascada input[type=checkbox]').length;
  const label = document.getElementById('labelSeccionesCascada');
  if (!label) return;
  if (checks.length === 0 || checks.length === total) {
    label.textContent = 'Todas';
  } else {
    label.textContent = checks.length + ' seleccionadas';
  }
}

function getSeccionesSeleccionadas() {
  const checks = document.querySelectorAll('#panelSeccionesCascada input[type=checkbox]:checked');
  return Array.from(checks).map(cb => cb.value);
}

function toggleSelector() {
  document.getElementById('selectorMulti').classList.toggle('open');
}
document.addEventListener('click', function(e) {
  const sel = document.getElementById('selectorMulti');
  if (sel && !sel.contains(e.target)) sel.classList.remove('open');
});

function seleccionarTodos(valor) {
  if (valor) PROYECTOS.forEach(p => seleccionados.add(p));
  else seleccionados.clear();
  refrescarSelector();
  areaFiltroPersona = "Todos";
  render();
  if (document.getElementById('tab-cascada').classList.contains('active')) renderCascada();
}

function toggleProyecto(p, checked) {
  if (checked) seleccionados.add(p); else seleccionados.delete(p);
  refrescarSelector();
  areaFiltroPersona = "Todos";
  render();
  if (document.getElementById('tab-cascada').classList.contains('active')) renderCascada();
}

function refrescarSelector() {
  document.getElementById('selectorLabel').textContent = `${seleccionados.size} de ${PROYECTOS.length}`;
  const cont = document.getElementById('selectorOpciones');
  cont.innerHTML = PROYECTOS.map(p =>
    `<label><input type="checkbox" ${seleccionados.has(p) ? 'checked' : ''} onchange="toggleProyecto('${p.replace(/'/g, "\\'")}', this.checked)"> ${p}</label>`
  ).join('');
}

function tareasSeleccionadas() {
  let out = [];
  PROYECTOS.forEach(p => { if (seleccionados.has(p)) out = out.concat(DATA[p]); });
  return out;
}

function pct(verde, total) { return total ? (100 * verde / total) : 0; }

function diasHabilesJS(d1str, d2str) {
  if (!d1str || !d2str) return null;
  const feriados = new Set([
    '2025-04-18','2025-04-19','2025-05-01','2025-05-21','2025-06-20','2025-06-29',
    '2025-07-16','2025-08-15','2025-09-18','2025-09-19','2025-10-12','2025-10-31',
    '2025-11-01','2025-12-08','2025-12-25','2026-01-01','2026-04-03','2026-04-04',
    '2026-05-01','2026-05-21','2026-06-20','2026-06-29','2026-07-16','2026-08-15',
    '2026-09-18','2026-09-19','2026-10-12','2026-10-31','2026-11-01','2026-12-08','2026-12-25'
  ]);
  let d1 = new Date(d1str + 'T12:00:00'), d2 = new Date(d2str + 'T12:00:00');
  const sign = d2 >= d1 ? 1 : -1;
  if (sign < 0) { let tmp=d1; d1=d2; d2=tmp; }
  let count = 0, cur = new Date(d1);
  cur.setDate(cur.getDate()+1);
  while (cur <= d2) {
    const iso = cur.toISOString().slice(0,10);
    if (cur.getDay()!==0 && cur.getDay()!==6 && !feriados.has(iso)) count++;
    cur.setDate(cur.getDate()+1);
  }
  return count * sign;
}

function render() {
  const tasks = tareasSeleccionadas();
  calcularBloqueo(tasks);
  const total = tasks.length;

  const complAtiempo = tasks.filter(t => t.estado_general === "Completada" && t.estado === "verde").length;
  const complAtraso  = tasks.filter(t => t.estado_general === "Completada" && t.estado === "rojo").length;
  const cursoAtiempo = tasks.filter(t => t.estado_general === "En curso").length;
  const cursoAtraso  = tasks.filter(t => t.estado_general === "Vencida").length;

  const subLabel = seleccionados.size === 1 ? Array.from(seleccionados)[0] : `${seleccionados.size} proyectos`;

  document.getElementById('globales').innerHTML = `
    <div class="cards-globales">
      <div class="card-g"><div class="lbl">Tareas nivel 2</div><div class="val">${total}</div><div class="sub">${subLabel}</div></div>
      <div class="card-g"><div class="lbl">Cerradas a tiempo</div><div class="val verde">${complAtiempo}</div><div class="sub">${pct(complAtiempo,total).toFixed(0)}% &middot; completadas en fecha o antes</div></div>
      <div class="card-g"><div class="lbl">Cerradas con atraso</div><div class="val rojo">${complAtraso}</div><div class="sub">${pct(complAtraso,total).toFixed(0)}% &middot; completadas 1+ día tarde</div></div>
      <div class="card-g"><div class="lbl">En curso a tiempo</div><div class="val verde">${cursoAtiempo}</div><div class="sub">${pct(cursoAtiempo,total).toFixed(0)}% &middot; aún dentro de plazo</div></div>
      <div class="card-g"><div class="lbl">En curso atrasadas</div><div class="val rojo">${cursoAtraso}</div><div class="sub">${pct(cursoAtraso,total).toFixed(0)}% &middot; vencidas sin completar</div></div>
    </div>`;

  renderAreas(tasks);
  renderPersonas(tasks);
  fillFiltrosTareas(tasks);
  renderTareas();
}

function renderAreas(tasks) {
  const ORDEN = ["Equipo Proyecto", "Servicios", "Compras", "Ingeniería", "Producción", "Bodega", "Logística"];
  const areas = {};

  // Primero acumular contadores generales desde las tareas (total, cerradas, verde, etc.)
  tasks.forEach(t => {
    areas[t.area] = areas[t.area] || {total:0, cerradas:0, verde:0, cerrada_atraso:0, abierta_vencida:0, dias_totales:0, dias_count:0};
    areas[t.area].total++;
    if (t.estado_general === 'Completada') areas[t.area].cerradas++;
    if (t.estado === 'verde') areas[t.area].verde++;
    if (t.estado_general === 'Completada' && t.estado === 'rojo') areas[t.area].cerrada_atraso++;
    if (t.estado_general === 'Vencida') areas[t.area].abierta_vencida++;
  });

  // Promedio por persona usando dias_plazo (días entre vencimiento y completado, con signo).
  // Negativo = antes del plazo, positivo = tarde. Promediamos esos valores por persona,
  // luego promediamos los promedios de personas por área.
  const personas = {};
  tasks.forEach(t => {
    if (!t.assignee || t.assignee === '(sin asignar)') return;
    if (t.estado_general !== 'Completada' || t.dias_plazo === null || t.dias_plazo === undefined) return;
    personas[t.assignee] = personas[t.assignee] || {area: t.area, dias_totales: 0, dias_count: 0};
    personas[t.assignee].dias_totales += t.dias_plazo;
    personas[t.assignee].dias_count++;
  });

  // Sumar los promedios individuales por área y dividir por cantidad de personas
  const areaProms = {};
  Object.values(personas).forEach(p => {
    if (p.dias_count === 0) return;
    const promPersona = p.dias_totales / p.dias_count;
    areaProms[p.area] = areaProms[p.area] || {suma: 0, cant: 0};
    areaProms[p.area].suma += promPersona;
    areaProms[p.area].cant++;
  });
  Object.keys(areaProms).forEach(area => {
    if (areas[area]) {
      const ap = areaProms[area];
      areas[area].prom_area = ap.cant > 0 ? ap.suma / ap.cant : null;
      areas[area].prom_personas = ap.cant;
    }
  });

  const keys = Object.keys(areas).sort((a,b) => {
    const ia = ORDEN.indexOf(a), ib = ORDEN.indexOf(b);
    return (ia === -1 ? 999 : ia) - (ib === -1 ? 999 : ib);
  });

  let html = '';
  keys.forEach(area => {
    const a = areas[area];
    const p = pct(a.cerradas, a.total);
    const cls = a.total === 0 ? '' : (p >= 80 ? 'verde' : (p >= 50 ? 'amarillo' : 'rojo'));
    const promDias = a.prom_area != null ? a.prom_area.toFixed(1) : null;
    const promPersonas = a.prom_personas || 0;

    html += `<div class="area-card">
      <div class="titulo">${area}</div>
      <div class="pct ${cls}">${p.toFixed(0)}%</div>
      <div class="meta">
        ${a.total} tareas · ${a.cerradas} cerradas<br>
        <span class="dot verde"></span>${a.verde} a tiempo
        &nbsp;<span class="dot rojo"></span>${a.cerrada_atraso} cerradas tarde
      </div>
      <div class="prom-box">
        <div>
          <div class="prom-val" style="${promDias !== null ? (parseFloat(promDias) > 0 ? 'color:var(--rojo)' : parseFloat(promDias) < 0 ? 'color:var(--verde)' : '') : ''}">${promDias !== null ? (parseFloat(promDias) > 0 ? '+' : '') + promDias + 'd' : '—'}</div>
          <div class="prom-lbl">${promDias !== null && parseFloat(promDias) > 0 ? 'prom. atraso' : promDias !== null && parseFloat(promDias) < 0 ? 'prom. anticipado' : 'prom. exacto'}<br>(${promPersonas} ${promPersonas === 1 ? 'persona' : 'personas'})</div>
        </div>
      </div>
    </div>`;
  });

  document.getElementById('areaGrid').innerHTML = html || '<i>Sin tareas nivel 2 en este proyecto</i>';
}

function renderPersonas(tasks) {
  const areasSet = new Set(tasks.map(t => t.area));
  const pillsHtml = ['Todos', ...Array.from(areasSet)].map(a =>
    `<div class="area-pill ${a===areaFiltroPersona?'active':''}" onclick="setAreaPersona('${a}')">${a}</div>`).join('');
  document.getElementById('areaPills').innerHTML = pillsHtml;

  const filtradas = areaFiltroPersona === 'Todos' ? tasks : tasks.filter(t => t.area === areaFiltroPersona);

  const personas = {};
  filtradas.forEach(t => {
    const key = t.assignee;
    personas[key] = personas[key] || {area: t.area, total:0, compl:0, verde:0, rojo:0, abierta_tiempo:0, abierta_atraso:0, cerrada_tiempo:0, cerrada_atraso:0, dias_totales:0, dias_count:0};
    personas[key].total++;
    if (t.estado_general === 'Completada') {
      personas[key].compl++;
      // Usar dias_plazo (diferencia entre vencimiento y completado, con signo)
      if (t.dias_plazo !== null && t.dias_plazo !== undefined) {
        personas[key].dias_totales += t.dias_plazo;
        personas[key].dias_count++;
      }
    }
    if (t.estado_general === 'En curso') personas[key].abierta_tiempo++;
    if (t.estado_general === 'Vencida') personas[key].abierta_atraso++;
    if (t.estado_general === 'Completada' && t.estado === 'verde') personas[key].cerrada_tiempo++;
    if (t.estado_general === 'Completada' && t.estado === 'rojo') personas[key].cerrada_atraso++;
  });

  const ordenadas = Object.entries(personas).sort((a,b) => b[1].total - a[1].total);
  let html = '';
  ordenadas.forEach(([nombre, d]) => {
    const p = pct(d.compl, d.total);
    const promDias = d.dias_count > 0 ? (d.dias_totales / d.dias_count).toFixed(1) : null;
    html += `<tr>
      <td></td>
      <td><b>${nombre}</b></td>
      <td>${d.area}</td>
      <td>${d.total}</td>
      <td>${d.compl}</td>
      <td><span class="pill ${p>=80?'verde':(p>=50?'amarillo':'rojo')}">${p.toFixed(0)}%</span></td>
      <td class="prom-cierre-cell">
        ${promDias !== null
          ? `<span style="color:${parseFloat(promDias) > 0 ? 'var(--rojo)' : parseFloat(promDias) < 0 ? 'var(--verde)' : 'var(--texto)'}">${parseFloat(promDias) > 0 ? '+' : ''}${promDias}d</span> <span>${d.dias_count} tareas</span>`
          : `<span style="color:#ccc;">—</span>`}
      </td>
      <td><span class="dot verde"></span>${d.cerrada_tiempo}</td>
      <td><span class="dot rojo"></span>${d.cerrada_atraso}</td>
      <td><span class="dot verde"></span>${d.abierta_tiempo}</td>
      <td><span class="dot rojo"></span>${d.abierta_atraso}</td>
    </tr>`;
  });
  document.getElementById('personaBody').innerHTML = html || '<tr><td colspan="11"><i>Sin datos</i></td></tr>';
}

function setAreaPersona(a) { areaFiltroPersona = a; renderPersonas(tareasSeleccionadas()); }

function fillFiltrosTareas(tasks) {
  const usuarios = [...new Set(tasks.map(t => t.assignee))].sort();
  const areas = [...new Set(tasks.map(t => t.area))].sort();
  const secciones = [...new Set(tasks.map(t => t.section).filter(s => s))].sort();
  const proyectos = [...new Set(tasks.map(t => t.project).filter(p => p))].sort();

  const fill = (id, items, label) => {
    const sel = document.getElementById(id);
    const current = sel.value;
    sel.innerHTML = `<option value="">${label}: Todos</option>` +
      items.map(i => `<option value="${i}">${i}</option>`).join('');
    sel.value = current;
  };
  fill('fUsuario', usuarios, 'Usuario');
  fill('fArea', areas, 'Área');
  fill('fSeccion', secciones, 'Sección');
  fill('fProyecto', proyectos, 'Proyecto');
}

let sortPlazoDir = null;

function toggleSortPlazo() {
  sortPlazoDir = sortPlazoDir === 'desc' ? 'asc' : 'desc';
  document.getElementById('sortPlazoIcon').textContent = sortPlazoDir === 'desc' ? ' ▼' : ' ▲';
  renderTareas();
}

function diasAtraso(estadoPlazo) {
  const m = estadoPlazo.match(/^(\d+)d/);
  return m ? parseInt(m[1]) : 0;
}

function calcularBloqueo(tasks) {
  const gruposPorNombre = {};
  tasks.forEach(t => {
    const n = t.name.trim().replace(/:$/, '').trim();
    const key = (t.project||'') + '||' + n;
    gruposPorNombre[key] = gruposPorNombre[key] || [];
    gruposPorNombre[key].push(t);
  });

  function buscarGrupo(nombre, proyecto) {
    const n = nombre.trim().replace(/:$/, '').trim();
    const key = (proyecto||'') + '||' + n;
    if (gruposPorNombre[key]) return gruposPorNombre[key];
    const nl = n.toLowerCase();
    const candidatos = tasks.filter(t => t.project === proyecto &&
      (t.name.toLowerCase().startsWith(nl) || nl.startsWith(t.name.toLowerCase())));
    return candidatos.length ? candidatos : null;
  }

  tasks.forEach(t => {
    if (!t.blocked_by || t.blocked_by.length === 0) {
      t.bloqueada_real = false;
      t.bloqueada_por = null;
      return;
    }
    let pendienteEncontrada = null;
    for (const nombre of t.blocked_by) {
      const grupo = buscarGrupo(nombre, t.project);
      if (!grupo || grupo.length === 0) continue;
      let candidato;
      if (grupo.length === 1) {
        candidato = grupo[0];
      } else {
        const refDate = t.start_iso || t.due_iso;
        if (refDate) {
          let mejor = null, mejorDiff = Infinity;
          grupo.forEach(g => {
            if (!g.due_iso) return;
            const diff = Math.abs(new Date(g.due_iso) - new Date(refDate));
            if (diff < mejorDiff) { mejorDiff = diff; mejor = g; }
          });
          if (mejor) {
            const subgrupo = grupo.filter(g => g.due_iso === mejor.due_iso);
            candidato = subgrupo.every(g => g.estado_general !== 'Completada')
              ? (subgrupo.find(g => g.estado_general !== 'Completada') || subgrupo[0])
              : null;
          }
        }
        if (candidato === undefined) {
          const todasPendientes = grupo.every(g => g.estado_general !== 'Completada');
          candidato = todasPendientes ? (grupo.find(g => g.estado_general !== 'Completada') || grupo[0]) : null;
        }
      }
      if (candidato && candidato.estado_general !== 'Completada') {
        pendienteEncontrada = candidato;
        break;
      }
    }
    t.bloqueada_real = !!pendienteEncontrada;
    t.bloqueada_por = pendienteEncontrada || null;
  });
}

function renderTareas() {
  const tasks = tareasSeleccionadas();
  const fu = document.getElementById('fUsuario').value;
  const fa = document.getElementById('fArea').value;
  const fs = document.getElementById('fSeccion').value;
  const fp = document.getElementById('fProyecto').value;
  const fe = document.getElementById('fEstado').value;
  const fb = document.getElementById('fBloqueada').value;

  const filtradas = tasks.filter(t =>
    (!fu || t.assignee === fu) &&
    (!fa || t.area === fa) &&
    (!fs || t.section === fs) &&
    (!fp || t.project === fp) &&
    (!fe || t.estado_general === fe) &&
    (!fb || (fb === 'si' ? t.bloqueada_real === true : t.bloqueada_real === false))
  );

  let sorted = [...filtradas];
  if (sortPlazoDir) {
    sorted.sort((a, b) => {
      const da = diasAtraso(a.estado_plazo), db = diasAtraso(b.estado_plazo);
      return sortPlazoDir === 'desc' ? db - da : da - db;
    });
  }

  const completadas = filtradas.filter(t => t.estado_general === 'Completada').length;
  const enCurso = filtradas.filter(t => t.estado_general === 'En curso').length;
  const vencidas = filtradas.filter(t => t.estado_general === 'Vencida').length;

  document.getElementById('resumenTareas').innerHTML =
    `${filtradas.length} tareas &nbsp; <b class="verde">${completadas} completadas (${pct(completadas,filtradas.length).toFixed(0)}%)</b> &nbsp; ` +
    `<b class="azul">${enCurso} en curso</b> &nbsp; <b class="rojo">${vencidas} vencidas</b>`;

  let html = '';
  sorted.forEach(t => {
    const cls = t.estado === 'verde' ? 'verde' : (t.estado === 'rojo' ? 'rojo' : 'encurso');
    const estadoLabel = t.estado_general === 'Vencida' ? '⚠ Vencida' :
                         t.estado_general === 'En curso' ? '● En curso' : '✓ Completada';
    html += `<tr>
      <td>${t.name}</td>
      <td style="font-size:12px;color:#5f6368;">${t.project}</td>
      <td>${t.section}</td>
      <td><b>${t.assignee}</b></td>
      <td>${t.area}</td>
      <td>${t.start_fmt || '--'}</td>
      <td>${t.due_fmt}</td>
      <td>${t.completed_fmt}</td>
      <td>${t.duracion_prevista !== null && t.duracion_prevista !== undefined ? t.duracion_prevista + 'd hábiles' : '--'}</td>
      <td class="estado-txt ${cls}">${t.estado_plazo}</td>
      <td><span class="estado-txt ${cls}">${estadoLabel}</span></td>
      <td>${t.bloqueada_real
        ? `<span style="color:var(--rojo);font-weight:700;" title="Bloqueada por: ${(t.blocked_by||[]).join(', ')}">🔴 Bloqueada</span>`
        : `<span style="color:var(--verde);font-weight:700;">🟢 Liberada</span>`}</td>
      <td style="font-size:12px;">${t.bloqueada_por ? `<b>${t.bloqueada_por.name}</b><br>👤 ${t.bloqueada_por.assignee}` : '—'}</td>
      <td>${t.asana_url ? `<a href="${t.asana_url}" target="_blank" style="color:var(--azul);text-decoration:none;font-weight:600;">🔗 Abrir</a>` : '—'}</td>
    </tr>`;
  });
  document.getElementById('tareasBody').innerHTML = html || '<tr><td colspan="14"><i>Sin tareas</i></td></tr>';
}

// Orden canónico global — nombres exactos del Excel (sin cabeceras)
const ORDEN_GLOBAL = [
  "Kick off meeting",
  "Definicion Jefe de proyecto",
  "Apertura pedido",
  "Alcance del proyecto",
  "Definicion Tecnica de componentes principales",
  "Plazo Contractual",
  "Ingenieria electrica",
  "Revision tableros pruebas",
  "Ingenieria Basica Motor",
  "Ingenieria basica Rodete",
  "Ingenieria Detalle",
  "Analisis de costeo",
  "propuesta motor",
  "propuesta alabes",
  "propuesta tableros",
  "propuesta nucleo rodete",
  "Propuesta sensores y accesorios",
  "propuesta compras a codigo",
  "Generación OC Alabe",
  "Fabricación Alabe",
  "Generación OC Tableros",
  "Fabricación Tablero",
  "Generación OC Rodete",
  "Fabricación Rodete",
  "Generación OC motor OT",
  "Fabricación motor OT",
  "Planos motor proveedor OT",
  "Materiales a codigo",
  "Generación OC materiales",
  "Fabricacion a codigo y compras locales",
  "Generación OC Sensores y accesorios",
  "Fabricacion Sensores y accesorios",
  "Planos Preliminar",
  "Planos Definitivos",
  "Check Planos",
  "Check pruebas FAT",
  "Recepcion materiales a codigos",
  "Control de calidad compras materiales",
  "Generación OC FABRICACIÓN EXTERNA",
  "Armado FABRICACIÓN EXTERNA",
  "Control de calidad Armado",
  "Recepcion Alabe",
  "Recepcion Rodete",
  "Recepcion Motor",
  "Recepcion Tableros",
  "Recepcion sensores y accesorios",
  "Revestimiento",
  "Montaje motor",
  "Montaje Alabe",
  "Montaje Rodete",
  "Montaje sensores y accesorios",
  "Pruebas FAT",
  "RE-PINTADO",
  "Coordinacion Entrega con Cliente",
  "Embalaje",
  "PACKING LIST",
  "Despacho",
  "Gestion",
  "Puesta en Marcha Servicios",
  "instalación Componentes",
  "Pruebas de instalación Componentes"
];

// Nombres exactos del Excel que corresponden a cabeceras de sección (se excluyen de la cascada)
const CABECERAS_SECCION = [
  "ingenieria de servicio:",
  "ingenieria jefe de proyecto:",
  "propuesta compras:",
  "compras:",
  "ingenieria y diseño:",
  "bodega:",
  "montaje:",
  "logistica:",
  "cierre proyecto:"
];

function esCabecera(nombre) {
  const n = normStr(nombre);
  return CABECERAS_SECCION.some(c => n === normStr(c) || n.endsWith(':'));
}

// Orden canónico global — nombres exactos del Excel (sin cabeceras)
const ORDEN_GLOBAL_UNUSED = [];

const SECCIONES_ORDEN = [
  { key: "kick_off",          label: "Kick off",             keywords: ["kick off"] },
  { key: "ingenieria",        label: "Ingeniería",           keywords: ["ingenieria de servicio","ingeniería de servicio","ingenieria jefe","ingeniería jefe"] },
  { key: "propuesta",         label: "Propuesta compras",    keywords: ["propuesta compra","propuesta de compra"] },
  { key: "compras",           label: "Compras",              keywords: ["compras"] },
  { key: "ingenieria_diseno", label: "Ingeniería y diseño",  keywords: ["ingenieria y diseño","ingeniería y diseño","ingenieria y diseno"] },
  { key: "bodega",            label: "Bodega",               keywords: ["bodega"] },
  { key: "montaje",           label: "Montaje",              keywords: ["montaje"] },
  { key: "logistica",         label: "Logística",            keywords: ["logistica","logística"] },
  { key: "cierre",            label: "Cierre proyecto",      keywords: ["cierre proyecto","cierre de proyecto"] },
];

function normStr(s) {
  return (s||'').toLowerCase().normalize("NFD").replace(/[̀-ͯ]/g,"").trim();
}

function ordenGlobal(nombre) {
  const n = normStr(nombre);
  // Primero busca coincidencia exacta
  for (let i = 0; i < ORDEN_GLOBAL.length; i++) {
    if (n === normStr(ORDEN_GLOBAL[i])) return i;
  }
  // Luego busca si empieza igual (para variantes con OT number)
  for (let i = 0; i < ORDEN_GLOBAL.length; i++) {
    const og = normStr(ORDEN_GLOBAL[i]);
    if (n.startsWith(og) || og.startsWith(n)) return i;
  }
  return 999;
}

function getSeccionLabel(section) {
  const sec = normStr(section||'');
  for (const s of SECCIONES_ORDEN) {
    if (s.keywords.some(k => sec.includes(normStr(k)))) return s.label;
  }
  return section || '';
}

function diasHabilesJSCascada(d1str, d2str) {
  if (!d1str || !d2str) return null;
  const feriados = new Set([
    '2025-04-18','2025-04-19','2025-05-01','2025-05-21','2025-06-20','2025-06-29',
    '2025-07-16','2025-08-15','2025-09-18','2025-09-19','2025-10-12','2025-10-31',
    '2025-11-01','2025-12-08','2025-12-25','2026-01-01','2026-04-03','2026-04-04',
    '2026-05-01','2026-05-21','2026-06-20','2026-06-29','2026-07-16','2026-08-15',
    '2026-09-18','2026-09-19','2026-10-12','2026-10-31','2026-11-01','2026-12-08','2026-12-25'
  ]);
  let d1 = new Date(d1str + 'T12:00:00'), d2 = new Date(d2str + 'T12:00:00');
  const sign = d2 >= d1 ? 1 : -1;
  if (sign < 0) { let tmp=d1; d1=d2; d2=tmp; }
  let count = 0, cur = new Date(d1);
  cur.setDate(cur.getDate()+1);
  while (cur <= d2) {
    const iso = cur.toISOString().slice(0,10);
    if (cur.getDay()!==0 && cur.getDay()!==6 && !feriados.has(iso)) count++;
    cur.setDate(cur.getDate()+1);
  }
  return count * sign;
}

function toggleSelectorSecciones() {
  document.getElementById('selectorSeccionesCascada').classList.toggle('open');
}
document.addEventListener('click', function(e) {
  const sel = document.getElementById('selectorSeccionesCascada');
  if (sel && !sel.contains(e.target)) sel.classList.remove('open');
});

function seleccionarTodasSecciones(valor) {
  document.querySelectorAll('#panelSeccionesCascada input[type=checkbox]').forEach(cb => cb.checked = valor);
  actualizarLabelSecciones();
  renderCascada();
}

function actualizarLabelSecciones() {
  const checks = document.querySelectorAll('#panelSeccionesCascada input[type=checkbox]:checked');
  const total = document.querySelectorAll('#panelSeccionesCascada input[type=checkbox]').length;
  const label = document.getElementById('labelSeccionesCascada');
  if (!label) return;
  label.textContent = (checks.length === 0 || checks.length === total) ? 'Todas' : checks.length + ' seleccionadas';
}

function getSeccionesSeleccionadas() {
  return Array.from(document.querySelectorAll('#panelSeccionesCascada input[type=checkbox]:checked')).map(cb => cb.value);
}

function renderCascada() {
  const tasks = tareasSeleccionadas();
  const soloAtraso = document.getElementById('fCascadaEstado').value === 'con_atraso';
  const seccionesSeleccionadas = getSeccionesSeleccionadas();
  actualizarLabelSecciones();

  // Agrupar por proyecto y ordenar
  const porProyecto = {};
  tasks.forEach(t => {
    if (!t.project) return;
    porProyecto[t.project] = porProyecto[t.project] || [];
    porProyecto[t.project].push(t);
  });

  let cadenas = Object.entries(porProyecto).map(([proyecto, tareas]) => ({
    proyecto,
    tareas: tareas
      .filter(t => t.name && t.name.trim() && !esCabecera(t.name))
      .sort((a, b) => ordenGlobal(a.name) - ordenGlobal(b.name))
  })).filter(c => c.tareas.length > 0);

  if (soloAtraso) cadenas = cadenas.filter(c => c.tareas.some(t => t.estado_general === 'Vencida'));

  if (seccionesSeleccionadas.length > 0) {
    cadenas = cadenas.map(c => ({
      proyecto: c.proyecto,
      tareas: c.tareas.filter(t => {
        const sec = normStr(t.section||'');
        return seccionesSeleccionadas.some(sel => {
          const found = SECCIONES_ORDEN.find(s => s.key === sel);
          return found && found.keywords.some(k => sec.includes(normStr(k)));
        });
      })
    })).filter(c => c.tareas.length > 0);
  }

  if (cadenas.length === 0) {
    document.getElementById('cascadaContainer').innerHTML =
      '<p style="color:#9aa0a6;padding:16px;">No se encontraron tareas que coincidan con los filtros.</p>';
    return;
  }

  function getColor(t) {
    if (t.estado_general === 'Completada') return { bg: '#1e8e3e', border: '#14532d', text: '#fff', label: '✓ Completada' };
    if (t.estado_general === 'Vencida')    return { bg: '#b3261e', border: '#7f1d1d', text: '#fff', label: '⚠ Vencida' };
    return { bg: '#1a73e8', border: '#1558b0', text: '#fff', label: '● En curso' };
  }

  let html = '';
  cadenas.forEach(({ proyecto, tareas }) => {
    html += `<div style="margin-bottom:32px;">
      <div style="font-size:13px;font-weight:700;color:#5f6368;margin-bottom:12px;">
        📁 ${proyecto} — ${tareas.length} tareas
      </div>
      <div style="display:flex;flex-direction:column;gap:4px;">`;

    tareas.forEach((t, i) => {
      const col = getColor(t);
      const diasReales = (t.completed && t.start_iso) ? diasHabilesJSCascada(t.start_iso, t.completed) : null;
      const antecesoras = (t.blocked_by && t.blocked_by.length > 0) ? t.blocked_by.join(' · ') : null;
      const secLabel = getSeccionLabel(t.section);

      if (i > 0) {
        html += `<div style="padding:1px 0 1px 16px;font-size:18px;color:#9aa0a6;">↓</div>`;
      }

      html += `
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;align-items:stretch;">
        <!-- Columna izquierda: tarea -->
        <div style="border-left:5px solid ${col.border};background:${col.bg};
                    border-radius:10px;padding:12px 16px;color:${col.text};box-sizing:border-box;">
          <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:6px;flex-wrap:wrap;">
            <div style="font-weight:700;font-size:14px;flex:1;">${t.name}</div>
            <div style="font-size:12px;font-weight:700;background:rgba(0,0,0,0.2);padding:2px 10px;border-radius:10px;white-space:nowrap;">${col.label}</div>
          </div>
          ${secLabel ? `<div style="font-size:11px;opacity:0.75;margin-top:3px;">📂 ${secLabel}</div>` : ''}
          <div style="font-size:12px;margin-top:6px;opacity:0.9;display:flex;gap:12px;flex-wrap:wrap;align-items:center;">
            <span>👤 ${t.assignee}</span>
            <span>🏢 ${t.area}</span>
          </div>
          <div style="font-size:12px;margin-top:4px;opacity:0.9;display:flex;gap:12px;flex-wrap:wrap;align-items:center;">
            <span>📅 ${t.start_fmt} → ${t.due_fmt}</span>
            ${t.duracion_prevista !== null && t.duracion_prevista !== undefined ? `<span>⏳ ${t.duracion_prevista}d prev.</span>` : ''}
          </div>
          <div style="font-size:12px;margin-top:4px;opacity:0.9;display:flex;gap:12px;flex-wrap:wrap;align-items:center;">
            ${t.completed_fmt !== '--' ? `<span>✅ ${t.completed_fmt}</span>` : ''}
            ${diasReales !== null ? `<span style="font-weight:700;">⏱ ${diasReales}d reales</span>` : ''}
            ${t.atraso_dias > 0 ? `<span style="font-weight:700;background:rgba(0,0,0,0.2);padding:1px 8px;border-radius:8px;">+${t.atraso_dias}d atraso</span>` : ''}
          </div>
        </div>
        <!-- Columna derecha: antecesoras -->
        <div style="border-left:3px solid #e6e3dc;background:#f8f8f6;
                    border-radius:10px;padding:12px 16px;box-sizing:border-box;">
          <div style="font-size:11px;font-weight:700;color:#5f6368;text-transform:uppercase;letter-spacing:.04em;margin-bottom:6px;">⬆ Bloqueada por</div>
          ${antecesoras
            ? antecesoras.split(' · ').map(a => `<div style="font-size:12px;color:#202124;padding:3px 0;border-bottom:1px solid #e6e3dc;">${a.trim()}</div>`).join('')
            : '<div style="font-size:12px;color:#9aa0a6;font-style:italic;">Sin antecesoras</div>'
          }
        </div>
      </div>`;
    });

    html += `</div></div>`;
  });

  document.getElementById('cascadaContainer').innerHTML = html;
}
refrescarSelector();
render();
</script>
</body>
</html>
"""


def main():
    input_dir = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "/mnt/user-data/outputs/KPI.html"

    input_dir_abs = os.path.abspath(input_dir)
    print(f"Buscando archivos .xlsx en: {input_dir_abs}")

    if not os.path.isdir(input_dir_abs):
        print(f"[ERROR] Esa carpeta NO existe. Verifica la ruta.")
        sys.exit(1)

    files = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]

    print(f"Archivos .xlsx encontrados: {len(files)}")
    if not files:
        # Mostrar qué tipo de archivos SI hay en la carpeta, para ayudar a diagnosticar
        todos = os.listdir(input_dir_abs)
        ejemplo = todos[:10]
        print(f"No se encontraron archivos .xlsx en {input_dir_abs}")
        if todos:
            print(f"La carpeta tiene {len(todos)} archivos/carpetas, por ejemplo: {ejemplo}")
        else:
            print("La carpeta está vacía.")
        sys.exit(1)

    today = datetime.date.today()
    now_chile = datetime.datetime.now(ZoneInfo("America/Santiago"))
    DIAS = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
    fecha_str = f"{DIAS[now_chile.weekday()]} {now_chile.strftime('%d/%m/%Y')} {now_chile.strftime('%H:%M')} hrs (Chile)"
    data = {}
    errores = []
    for f in files:
        try:
            r = process_file(f, today)
            if r is None:
                errores.append((f, "encabezados/columnas no reconocidas"))
                continue
            data[r["project"]] = r["tasks"]
        except Exception as e:
            errores.append((f, str(e)))

    if not data:
        print("No se pudo procesar ningun archivo.")
        for f, e in errores:
            print(f"  - {f}: {e}")
        sys.exit(1)

    html_out = TEMPLATE.replace("__DATA__", json.dumps(data, ensure_ascii=False))
    html_out = html_out.replace("__FECHA__", fecha_str)


    out_dir = os.path.dirname(output_file)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(output_file, "w", encoding="utf-8") as fh:
        fh.write(html_out)

    print(f"Proyectos procesados: {len(data)} / {len(files)}")
    for f, e in errores:
        print(f"  [ERROR] {os.path.basename(f)}: {e}")
    print(f"KPI generado en: {output_file}")


if __name__ == "__main__":
    main()
