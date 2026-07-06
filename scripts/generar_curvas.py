#!/usr/bin/env python3
"""
generar_curvas.py — Portafolio Zitron · Curva S
Genera const DATA = [...] en CURVAS.HTML
Uso: python3 scripts/generar_curvas.py data CURVAS.HTML CURVAS.HTML
"""
import os, sys, re, json
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

HORAS_DIA = 8.3

CATALOGO_MAP = {
    '50001566':'1215922055649550','50001563':'1215727332551578',
    '50001559':'1215504785832997','50001558':'1215137857526702',
    '50001557':'1215139673478155','50001557b':'1215137857526520',
    '50001555':'1214922603742599','50001554':'1215118022011721',
    '50001553-OT4326':'1214969047721199','50001553-OT4327':'1214969047720950',
    '50001553-OT4324':'1214969284200793','50001553':'1214969047721199',
    '50001547':'1214787972061369','50001546':'1214739697907723',
    '50001545':'1214563704105168','50001544':'1214892476594825',
    '50001543':'1214508463084754','50001541':'1214137717389412',
    '50001534':'1213881596172396','50001533':'1213955236952392',
    '50001532':'1213963037596622','50001525':'1213832650589314',
    '50001524':'1213458788103499','50001520':'1213396195711984',
    '50001518':'1213458785834993','50001516':'1213391174146526',
    '50001515':'1213234368821945','50001514':'1213185599076077',
    '50001508':'1213362937195444','50001506':'1213244147627519',
    '50001504-B':'1213377149548924','50001504':'1213400421980747',
    '50001501':'1213377149548860','50001500':'1213234368822465',
    '50001499':'1213244147627934','50001498':'1213377149548798',
    '50001497':'1213193352022841','50001490':'1213377149548731',
    '50001485':'1213377149548665','50001477':'1213377149548590',
    '50001473':'1213377149548525','50001466':'1213997266064436',
    '50001451':'1213391072478496','50001446':'1213377149548988',
    '50001415':'1213391072478428','50001309':'1213377149548388',
}

KICKOFF_KW = {'apertura pedido','alcance del proyecto','definición técnica',
              'definicion tecnica','plazo contractual','definición jefe','definicion jefe'}
DUE_KW = {'despacho','exw','coordinacion entrega','coordinación entrega'}
HOY = date.today()

# Calendario semanas dom-sáb (fijo 2026)
SEMANAS_CAL = []
_base_sem = date(2025, 12, 28)  # domingo 28 dic 2025
_PV_SEMS = [0, 1, 3, 5, 7, 9, 11, 13, 15, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34, 36, 38, 39, 41, 43, 45, 47, 49, 51, 53, 55, 57, 59, 61, 62, 64, 66, 68, 70, 72, 74, 76, 78, 80, 82, 84, 85, 87, 89, 91, 93, 95, 97, 100]
for _i in range(53):
    _ini = _base_sem + timedelta(weeks=_i)
    _fin = _ini + timedelta(days=6)
    SEMANAS_CAL.append({'n': _i+1, 'ini': _ini, 'fin': _fin, 'pv': _PV_SEMS[_i]})

def semana_de(d):
    """Retorna número de semana del calendario fijo dom-sáb.
    Si la fecha está fuera del rango, calcula la semana extendida."""
    if not d: return None
    if isinstance(d, str):
        try: d = datetime.strptime(d[:10], '%Y-%m-%d').date()
        except: return None
    if isinstance(d, datetime): d = d.date()
    for s in SEMANAS_CAL:
        if s['ini'] <= d <= s['fin']:
            return s['n']
    # Fuera del rango: calcular semana relativa
    base = SEMANAS_CAL[0]['ini']  # 2025-12-28
    diff = (d - base).days
    return int(diff // 7) + 1


def es_kickoff(n): return any(kw in n.lower() for kw in KICKOFF_KW)
def es_due(n):     return any(kw in n.lower() for kw in DUE_KW)

def fmt(d):
    if isinstance(d,(datetime,date)): return d.strftime('%Y-%m-%d')
    return ''

def parse(s):
    if not s: return None
    try: return datetime.strptime(str(s)[:10],'%Y-%m-%d').date()
    except: return None

def dias_lab(d1,d2):
    if not d1 or not d2 or d1>d2: return 0
    n,c=0,d1
    while c<=d2:
        if c.weekday()<5: n+=1
        c+=timedelta(days=1)
    return n

def lunes(d): return d-timedelta(days=d.weekday())

def leer_excel(path):
    wb=load_workbook(path,read_only=True,data_only=True)
    ws=wb.active
    nombre=''
    for row in ws.iter_rows(min_row=1,max_row=1,values_only=True):
        nombre=str(row[0] or '').strip(); break
    headers={}
    for row in ws.iter_rows(min_row=3,max_row=3,values_only=True):
        for i,v in enumerate(row):
            if v: headers[str(v).strip()]=i
        break
    def col(n,d): return headers.get(n,d)
    ci_name=col('Name',4); ci_parent=col('Parent task',13)
    # Fechas: SOLO Start Date (col i=8) y Due Date (col j=9)
    ci_ini_asana = col('Start Date', 8)
    ci_fin_asana = col('Due Date', 9)
    ci_av=col('Avance Tarea',None) or col('% Avance proyecto',None) or col('Avance',None) or 18
    ci_comp=col('Completed At',2)

    tareas=[]; seccion=''; due_str=''
    for row in ws.iter_rows(min_row=4,values_only=True):
        name=str(row[ci_name] or '').strip()
        parent=str(row[ci_parent] or '').strip()
        if not name: continue
        if not parent:
            seccion=name.rstrip(':').strip(); continue
        # Usar Start Date y Due Date de Asana
        ini_raw = row[ci_ini_asana] if ci_ini_asana is not None else None
        fin_raw = row[ci_fin_asana] if ci_fin_asana is not None else None
        fin=fmt(fin_raw) if fin_raw else ''
        ini=fmt(ini_raw) if ini_raw else fin
        comp_raw=row[ci_comp]
        comp_str=fmt(comp_raw) if comp_raw else ''
        if not ini and not fin: continue
        if not ini: ini=fin
        if not fin: fin=ini
        # Si ini > fin (fechas invertidas), intercambiar
        if ini > fin:
            ini, fin = fin, ini
        # Avance: si tiene Completed At → 1.0
        # Si col av es numérico → usarlo
        # Si col av es texto (Alta/Baja/Media) → 0.0
        comp_check = row[ci_comp]
        if comp_check:
            av = 1.0
        else:
            try:
                av_raw2 = row[ci_av]
                if av_raw2 is None:
                    av = 0.0
                elif isinstance(av_raw2, (int, float)):
                    av = round(min(max(float(av_raw2), 0.0), 1.0), 4)
                elif isinstance(av_raw2, str) and av_raw2.replace('.','').replace(',','').isdigit():
                    av = round(min(max(float(av_raw2.replace(',','.')), 0.0), 1.0), 4)
                else:
                    av = 0.0  # texto como Alta/Baja/Media
            except:
                av = 0.0
        hoy_str=HOY.strftime('%Y-%m-%d')
        ini_ef,fin_ef=ini,fin
        if av>=1.0 and fin>hoy_str: fin_ef=hoy_str; ini_ef=min(ini,hoy_str)
        if es_due(name) and fin_ef:
            if not due_str or fin_ef>due_str: due_str=fin_ef
        sem_ini=semana_de(ini_ef)
        sem_fin=semana_de(fin_ef)
        tareas.append({'name':name,'section':seccion,'ini':ini_ef,'fin':fin_ef,
                       'av':av,'kickoff':es_kickoff(name),
                       'completed_at':comp_str,'semana_ini':sem_ini,'semana_fin':sem_fin})
    wb.close()
    return nombre,due_str,tareas

def extraer_pedido(filename):
    fn=filename.upper()
    m=re.search(r'5\d{7}',filename)
    p=m.group(0) if m else ''
    if p=='50001504':
        return p,'50001504-B' if ('CODESTABLE B' in fn or ' B.' in fn or '-B.' in fn) else '50001504'
    if p=='50001557':
        return p,'50001557' if 'OT4340' in fn else '50001557b'
    if p=='50001553':
        if 'OT4327' in fn: return p,'50001553-OT4327'
        if 'OT4324' in fn or 'OT4325' in fn: return p,'50001553-OT4324'
        return p,'50001553'
    return p,p

def calcular_proyecto(pedido,nombre_proy,due_str,tareas):
    kickoffs=[t for t in tareas if t['kickoff']]
    normales=[t for t in tareas if not t['kickoff']]
    n_kof=max(len(kickoffs),1)

    todas_ini=sorted([t['ini'] for t in tareas if t['ini']])
    todas_fin=sorted([t['fin'] for t in tareas if t['fin']])
    kickoff_str=todas_ini[0] if todas_ini else ''
    fin_real_str=todas_fin[-1] if todas_fin else ''
    kickoff_d=parse(kickoff_str)
    fin_real_d=parse(fin_real_str)

    # Buscar contractual (plazo contractual en tareas kickoff)
    contractual_str=None
    for t in kickoffs:
        if 'plazo contractual' in t['name'].lower():
            contractual_str=t['fin']
            break

    # total_pv se calculará después de distribuir PV por semanas

    # Usar calendario fijo SEMANAS_CAL (dom-sáb)
    if not kickoff_d: return None

    # Filtrar semanas relevantes: desde kickoff hasta fin_real + buffer
    end_date = fin_real_d or (HOY + timedelta(weeks=4))
    # Buffer de 4 semanas extra para asegurar que PV llega a 100%
    # Solo semanas desde la semana del kickoff en adelante (sin semanas previas)
    semanas_rel = [s for s in SEMANAS_CAL
                   if s['ini'] <= kickoff_d <= s['fin'] or  # semana del kickoff
                   (s['ini'] > kickoff_d and s['ini'] <= end_date + timedelta(weeks=4))]

    # Si el kickoff está antes del S1, agregar semanas extra hacia atrás
    if kickoff_d < SEMANAS_CAL[0]['ini']:
        extra = []
        cur = SEMANAS_CAL[0]['ini'] - timedelta(weeks=1)
        while cur + timedelta(days=6) >= kickoff_d:
            extra.insert(0, {'n': None, 'ini': cur, 'fin': cur + timedelta(days=6)})
            cur -= timedelta(weeks=1)
        semanas_rel = extra + semanas_rel

    # PV y EV por semana del calendario
    pv_sem = {}
    ev_sem = {}
    n_sem  = {s['ini']: 0 for s in semanas_rel}
    tareas_calc = []  # resultados de la 1ra pasada (PV), usados en la 2da pasada (EV)

    for t in tareas:
        ini_d = parse(t['ini']); fin_d = parse(t['fin'])
        if not ini_d: continue
        fin_d = fin_d or ini_d
        # Horas: kickoff=1.6 fijo, resto=días laborales × 8.3
        # Si completada antes de la fecha inicio planificada → duración real = 1 día
        comp_check = parse(t.get('completed_at',''))
        if t['kickoff']:
            h = 1.6
        elif comp_check and ini_d and comp_check < ini_d:
            h = HORAS_DIA  # completada antes del inicio → 1 día
        else:
            h = max(dias_lab(ini_d, fin_d), 1) * HORAS_DIA
        if h <= 0: continue

        # PV va TODO a la semana de fin_d (Due Date) — igual que tu Excel
        sem_pv = None
        for s in semanas_rel:
            if s['ini'] <= fin_d <= s['fin']:
                sem_pv = s['ini']
                break
        if sem_pv is None:
            # Buscar semana más cercana al fin_d
            dists = [(abs((s['ini'] - fin_d).days), s['ini']) for s in semanas_rel]
            sem_pv = min(dists, key=lambda x: x[0])[1]
        pv_sem[sem_pv] = pv_sem.get(sem_pv, 0.0) + h
        n_sem[sem_pv] = n_sem.get(sem_pv, 0) + 1

        tareas_calc.append({'h': h, 'sem_pv': sem_pv, 'comp_d': parse(t.get('completed_at', '')), 'av': t['av']})

    # Última semana con PV real (planificación). Si el trabajo real se extendió más allá
    # de lo planificado, todo el EV posterior se acumula en esta última semana en vez de
    # seguir generando semanas nuevas — el proyecto "cierra" visualmente en su horizonte planificado.
    last_pv_week = max(pv_sem.keys()) if pv_sem else None
    if last_pv_week is not None:
        semanas_rel = [s for s in semanas_rel if s['ini'] <= last_pv_week]

    for tc in tareas_calc:
        h, sem_pv, comp_d, av = tc['h'], tc['sem_pv'], tc['comp_d'], tc['av']
        # EV va a la semana de completed_at usando SEMANAS_CAL fijo
        if av >= 1.0 and comp_d:
            # Buscar en SEMANAS_CAL (calendario fijo dom-sáb)
            sem_ev = None
            for sc in SEMANAS_CAL:
                if sc['ini'] <= comp_d <= sc['fin']:
                    sem_ev = sc['ini']
                    break
            if sem_ev is None:
                # Fuera del rango del calendario
                diff = (comp_d - SEMANAS_CAL[0]['ini']).days
                n = int(diff // 7)
                sem_ev = SEMANAS_CAL[0]['ini'] + timedelta(weeks=n)
            # Clamp: si se completó después de la última semana planificada, se cuenta igual
            # pero se acumula en esa última semana (no se generan semanas nuevas)
            if last_pv_week is not None and sem_ev > last_pv_week:
                sem_ev = last_pv_week
            ev_sem[sem_ev] = ev_sem.get(sem_ev, 0.0) + h * av
        elif av > 0:
            # Tarea en progreso: EV va a semana de fin_d × avance
            ev_sem[sem_pv] = ev_sem.get(sem_pv, 0.0) + h * av

    # Rellenar semanas sin PV con 0
    for s in semanas_rel:
        if s['ini'] not in pv_sem: pv_sem[s['ini']] = 0.0
        if s['ini'] not in ev_sem: ev_sem[s['ini']] = 0.0
        if s['ini'] not in n_sem:  n_sem[s['ini']]  = 0

    # total_pv = suma de todo el PV
    total_pv = sum(pv_sem.values())
    if total_pv <= 0: return None

    # Construir rows
    pv_a = 0.0; ev_a = 0.0
    rows = []
    for idx_s, s in enumerate(semanas_rel):
        pv_a += pv_sem[s['ini']]
        # EV: sumar todo ev_sem cuya clave (fecha completada) <= HOY y <= fin de esta semana
        # EV: sumar claves de ev_sem dentro del rango de esta semana y <= HOY
        for ev_key, ev_val in ev_sem.items():
            if s['ini'] <= ev_key <= s['fin'] and ev_key <= HOY:
                ev_a += ev_val
        # pctPV: calculado por horas reales del proyecto
        pct_pv_cal = round(pv_a / total_pv * 100, 1)
        pct_ev = round(ev_a / total_pv * 100, 1)
        # EV semanal = suma de ev_sem cuyas claves caen en esta semana
        ev_sem_val = round(sum(v for k,v in ev_sem.items() if s['ini'] <= k <= s['fin'] and k <= HOY), 1)
        # Semana del calendario fijo
        cal_wk = s.get('n')
        rows.append({
            'idx': idx_s + 1,
            'ws':  s['ini'].strftime('%Y-%m-%d'),
            'we':  s['fin'].strftime('%Y-%m-%d'),
            'n':   n_sem[s['ini']],
            'pv':  round(pv_sem[s['ini']], 1),
            'ev':  ev_sem_val,
            'pvA': round(pv_a, 1),
            'evA': round(ev_a, 1),
            'pctPV': pct_pv_cal,
            'pctEV': pct_ev,
            'cal_week': cal_wk,
        })

    # Métricas finales
    # rows_pas: última fila cuyo inicio de semana ya llegó (incluye la semana en curso completa,
    # aunque aún no haya terminado) — así el KPI se actualiza desde el día 1 de cada semana
    rows_pas=[r for r in rows if parse(r['ws'])<=HOY]
    at_weeks=len(rows_pas)-1 if rows_pas else 0

    ultima = rows_pas[-1] if rows_pas else None
    pct_ev_f=ultima['pctEV'] if ultima else 0.0
    pct_pv_f=ultima['pctPV'] if ultima else 0.0
    semana_actual=ultima['cal_week'] if ultima else None
    spit=round(pct_ev_f/pct_pv_f,2) if pct_pv_f>0 else 0.0
    svt=round(pct_ev_f-pct_pv_f,1)

    # ES
    es_weeks=0.0
    if pct_ev_f>0:
        for i,r in enumerate(rows):
            if r['pctPV']>=pct_ev_f:
                if i==0: es_weeks=0.0
                else:
                    prev=rows[i-1]
                    frac=(pct_ev_f-prev['pctPV'])/(r['pctPV']-prev['pctPV']) if r['pctPV']!=prev['pctPV'] else 1.0
                    es_weeks=(i-1)+frac
                break
        else: es_weeks=len(rows)-1.0

    # EAC
    eac_date=None
    if spit>0 and pct_ev_f<100:
        sem_rest=(100-pct_ev_f)/max(pct_ev_f/max(at_weeks,1),0.01)
        eac_date=(HOY+timedelta(weeks=sem_rest)).strftime('%Y-%m-%d')
    elif pct_ev_f>=100:
        eac_date=fin_real_str

    # Estado
    due_d=parse(due_str)
    if pct_ev_f>=100: estado='Terminado'
    elif due_d and HOY>due_d and pct_ev_f<100: estado='Vencido'
    elif svt>=0: estado='Adelantado'
    else: estado='Atrasado'

    # Prob
    if pct_ev_f>=100: prob=100
    elif spit>=1.0: prob=min(90,int(spit*70))
    elif spit>=0.8: prob=35
    elif spit>=0.5: prob=15
    else: prob=5

    nombre_d=f"{pedido} - {nombre_proy}" if nombre_proy and pedido not in nombre_proy else (nombre_proy or pedido)

    tareas_out=[]
    for t in tareas:
        comp = t.get('completed_at','')
        sem_comp = semana_de(comp) if comp else None
        tareas_out.append({
            'name':t['name'],'section':t['section'],
            'ini':t['ini'],'fin':t['fin'],
            'av':t['av'],'kickoff':t['kickoff'],
            'completed_at':comp,
            'semana_ini':t.get('semana_ini'),
            'semana_fin':t.get('semana_fin'),
            'semana_comp':sem_comp
        })

    return {
        'nombre':nombre_d,
        'kickoff':kickoff_str,
        'contractual':contractual_str,
        'exw':due_str or None,
        'fin_real':fin_real_str,
        'eac_date':eac_date,
        'total_pv':round(total_pv,1),
        'pct_ev':pct_ev_f,
        'pct_pv':pct_pv_f,
        'at_weeks':at_weeks,
        'es_weeks':round(es_weeks,2),
        'es_iso':round(es_weeks,1),
        'spit':spit,
        'svt':svt,
        'semana_actual':semana_actual,
        'prob':prob,
        'estado':estado,
        'despachado':False,
        'rows':rows,
        'tareas':tareas_out
    }

def procesar(carpeta):
    archivos=sorted([f for f in os.listdir(carpeta) if f.lower().endswith('.xlsx')])
    proyectos=[]
    for archivo in archivos:
        print(f'[INFO] {archivo}')
        try:
            nombre,due,tareas=leer_excel(os.path.join(carpeta,archivo))
            pedido,key=extraer_pedido(archivo)
            if not pedido or not tareas: print('  [SKIP]'); continue
            p=calcular_proyecto(pedido,nombre,due,tareas)
            if p:
                proyectos.append(p)
                print(f'  → EV={p["pct_ev"]}% SPI={p["spit"]} Estado={p["estado"]}')
            else:
                print('  [SKIP] sin fechas')
        except Exception as e:
            print(f'  [ERROR] {e}')
            import traceback; traceback.print_exc()
    return proyectos

def inyectar(template,output,proyectos):
    with open(template,'r',encoding='utf-8') as f: html=f.read()

    nuevo_data='const DATA = '+json.dumps(proyectos,ensure_ascii=False,separators=(',',':'))+';'

    # Reemplazar const DATA = [...];
    idx=html.find('const DATA =')
    if idx>=0:
        # Buscar cierre del array
        depth=0; pos=idx; in_str=False; sc=None; found=-1
        while pos<len(html):
            c=html[pos]
            if in_str:
                if c==sc and html[pos-1]!='\\': in_str=False
            else:
                if c in('"',"'",'`'): in_str=True; sc=c
                elif c=='[': depth+=1
                elif c==']':
                    depth-=1
                    if depth==0:
                        j=pos+1
                        while j<len(html) and html[j] in(' ','\n','\r','\t'): j+=1
                        if j<len(html) and html[j]==';': found=j
                        else: found=pos
                        break
            pos+=1
        if found>=0:
            html=html[:idx]+nuevo_data+html[found+1:]
            print(f'[OK] DATA reemplazado ({len(proyectos)} proyectos)')
        else:
            print('[WARN] No se encontró cierre de DATA')
    else:
        print('[WARN] No se encontró const DATA — insertando')
        html=html.replace('<script>','<script>\n'+nuevo_data+'\n',1)

    # Timestamp
    # Guardar timestamp en UTC — el JS lo convierte a hora Chile
    ahora=datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    html=re.sub(r'<!-- updated:.*?-->',f'<!-- updated: {ahora} -->',html)
    if '<!-- updated:' not in html:
        html=html.replace('</head>',f'\n<!-- updated: {ahora} -->\n</head>',1)
    print(f'[OK] Timestamp: {ahora}')

    with open(output,'w',encoding='utf-8') as f: f.write(html)
    print(f'[OK] Guardado: {output}')

if __name__=='__main__':
    if len(sys.argv)<4:
        print('Uso: python3 generar_curvas.py <data> <template.html> <output.html>')
        sys.exit(1)
    print(f'\n{"="*60}')
    proyectos=procesar(sys.argv[1])
    if not proyectos: print('[ERROR] Sin proyectos'); sys.exit(1)
    inyectar(sys.argv[2],sys.argv[3],proyectos)
    print(f'\n✅ Listo. {len(proyectos)} proyectos.')
