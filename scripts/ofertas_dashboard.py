#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/ofertas_dashboard.py  — VERSIÓN CON SOPORTE ONEDRIVE
=============================================================

Lee el .xlsx de Ofertas y regenera ofertas2026.html reemplazando el
bloque `var D={...};` dentro del <script> del dashboard.

DETECTA AUTOMÁTICAMENTE DOS FORMATOS:

  FORMATO ASANA (formato anterior)
    Encabezados: Task ID | Name | Due Date | Completed At | Parent task | ...
    Estructura:  subtareas por oferta (Analisis Departamento ingenieria, etc.)

  FORMATO ONEDRIVE (formato nuevo — Excel de Ofertas Pendientes)
    Encabezados: Nº Oferta | Cliente | Descripcion | Status | OC recibida |
                 Economica | Tecnica | ENVIO | Fecha entrada |
                 Fecha limite presentacion | Fecha salida | Asignado
    Estructura:  una fila por oferta (tabla plana)

USO (sin cambios respecto a la versión anterior):
  python3 scripts/ofertas_dashboard.py <carpeta_data> ofertas2026.html [archivo.xlsx]
"""

import sys
import os
import re
import glob
import json
import datetime
from zoneinfo import ZoneInfo
import warnings
import openpyxl

warnings.filterwarnings("ignore")

# Constantes para formato Asana
SECCION_INGENIERIA = "analisis departamento ingenieria"
SECCION_PROYECTO = "analisis departamento proyecto"
SECCION_CONSOLIDADO_PREFIX = "consolidado ofert"

# País por prefijo del número de oferta
PAIS_MAP = {
    "CL": "Chile", "PE": "Peru", "MX": "Mexico", "CO": "Colombia",
    "AR": "Argentina", "EC": "Ecuador", "CR": "Costa Rica",
    "NI": "Nicaragua", "BO": "Bolivia", "VE": "Venezuela",
}

TZ = ZoneInfo("America/Santiago")


# ─────────────────────────────────────────────────────────────────
# UTILIDADES COMPARTIDAS
# ─────────────────────────────────────────────────────────────────

def normalize(s):
    s = (s or "").strip().lower()
    for a, b in {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n"}.items():
        s = s.replace(a, b)
    return s


def find_column(headers_norm, *substrings):
    subs = [normalize(s) for s in substrings]
    for h, hn in headers_norm.items():
        if all(s in hn for s in subs):
            return h
    return None


def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def fmt_date(d):
    return d.strftime("%d-%m-%Y") if d else ""


def str_val(v):
    if v is None:
        return ""
    return str(v).strip()


def pais_from_nro(nro):
    m = re.match(r"\d+([A-Z]{2})\d+", (nro or "").upper())
    return PAIS_MAP.get(m.group(1), "Sin pais") if m else "Sin pais"


# ─────────────────────────────────────────────────────────────────
# DETECCIÓN DE FORMATO
# ─────────────────────────────────────────────────────────────────

def detect_format(ws, max_scan=12):
    """
    Devuelve (formato, fila_encabezado, lista_headers).
    formato: "asana" | "onedrive" | "unknown"
    """
    for r in range(1, max_scan + 1):
        row = [c.value for c in ws[r]]
        if not any(v for v in row):
            continue
        row_str = " ".join(str(v or "") for v in row).lower()

        # Formato Asana: tiene Task ID y Name
        if "task id" in row_str and "name" in row_str:
            return "asana", r, row

        # Formato OneDrive: tiene ENVIO y Fecha salida (o similar)
        if ("envio" in row_str or "envío" in row_str) and (
            "salida" in row_str or "oferta" in row_str or "cliente" in row_str
        ):
            return "onedrive", r, row

        # También: tiene Nº Oferta
        if any(x in row_str for x in ["nº oferta", "n° oferta", "numero oferta"]):
            return "onedrive", r, row

    return "unknown", None, None


# ─────────────────────────────────────────────────────────────────
# LOCALIZAR ARCHIVO XLSX
# ─────────────────────────────────────────────────────────────────

def locate_xlsx(data_dir, explicit_name=None):
    if explicit_name:
        path = os.path.join(data_dir, explicit_name)
        if os.path.exists(path):
            return path
        sys.exit(f"ERROR: no existe {path}")

    candidates = sorted(glob.glob(os.path.join(data_dir, "*.xlsx")))
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]

    for c in candidates:
        n = normalize(os.path.basename(c))
        if ("zitron" in n and "2026" in n) or "oferta" in n:
            return c

    if candidates:
        return candidates[-1]  # último modificado

    sys.exit("ERROR: no se encontró ningún .xlsx en " + data_dir)


# ─────────────────────────────────────────────────────────────────
# PROCESAMIENTO — FORMATO ONEDRIVE (nuevo)
# ─────────────────────────────────────────────────────────────────

def build_offers_onedrive(ws, header_row, headers):
    """
    Una fila = una oferta.
    Columnas esperadas (acepta variaciones de nombre/tilde):
      Nº Oferta | Cliente | Descripcion | Status | OC recibida |
      Economica | Tecnica | ENVIO | Fecha entrada |
      Fecha limite presentacion | Fecha salida | Asignado
    """
    idx = {name: i for i, name in enumerate(headers) if name is not None}
    hn = {name: normalize(str(name)) for name in idx if name}

    # Mapeo de columnas
    col_nro   = (find_column(hn, "oferta") or find_column(hn, "nro")
                 or find_column(hn, "numero"))
    col_cli   = find_column(hn, "cliente")
    col_desc  = find_column(hn, "descripci") or col_cli
    col_status = find_column(hn, "status")
    col_oc    = find_column(hn, "oc", "recibida") or find_column(hn, "oc")
    col_econ  = find_column(hn, "econom")
    col_tec   = find_column(hn, "tecn")
    col_envio = find_column(hn, "envio")
    col_f_ent = find_column(hn, "entrada")
    col_f_lim = (find_column(hn, "limite") or find_column(hn, "presentaci")
                 or find_column(hn, "plazo"))
    col_f_sal = find_column(hn, "salida") or find_column(hn, "finaliz")
    col_asig  = find_column(hn, "asignado") or find_column(hn, "responsable")
    col_pais  = find_column(hn, "pais")
    col_notas = find_column(hn, "nota") or find_column(hn, "observaci")

    print("  [OneDrive] Columnas detectadas:")
    print(f"    nro={col_nro!r}  cliente={col_cli!r}  status={col_status!r}  envio={col_envio!r}")
    print(f"    f_ent={col_f_ent!r}  f_lim={col_f_lim!r}  f_sal={col_f_sal!r}  asig={col_asig!r}")

    if not col_nro:
        sys.exit("ERROR OneDrive: no se encontró columna 'Nº Oferta'.")

    def get(row, col):
        if not col:
            return ""
        v = row[idx[col]] if idx.get(col) is not None else None
        return str(v).strip() if v is not None else ""

    offers = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(c is None for c in row):
            continue

        nro = get(row, col_nro)
        if not nro:
            continue
        # Saltar filas que sean repetición del encabezado
        if normalize(nro) in ("n oferta", "nro", "numero oferta", "oferta"):
            continue

        cli    = get(row, col_cli)
        desc   = get(row, col_desc) or cli or nro
        status = get(row, col_status)
        envio  = get(row, col_envio)
        asig   = get(row, col_asig) or "Sin Asignar"
        notas  = get(row, col_notas) if col_notas else ""

        # País
        pais = get(row, col_pais) if col_pais else ""
        if not pais:
            pais = pais_from_nro(nro)

        # Fechas
        f_ent_d  = to_date(get(row, col_f_ent))  if col_f_ent  else None
        f_lim_d  = to_date(get(row, col_f_lim))  if col_f_lim  else None
        f_sal_d  = to_date(get(row, col_f_sal))  if col_f_sal  else None
        f_econ_d = to_date(get(row, col_econ))   if col_econ   else None
        f_tec_d  = to_date(get(row, col_tec))    if col_tec    else None

        # Estado unificado
        envio_n  = normalize(envio)
        status_n = normalize(status)

        if "enviada" in envio_n or "enviada" in status_n:
            estado        = "Enviada"
            ing = proy = cons = True
            comp, tot = 3, 3
        elif ("parcial" in status_n or "tecnica" in envio_n
              or ("tecn" in envio_n and "enviada" not in envio_n)):
            estado = "Parcial"
            ing = proy = True; cons = False
            comp, tot = 2, 3
        elif any(x in status_n for x in
                 ["progreso", "asignado", "revision", "aclaraci", "espera", "rev2",
                  "lista", "listo", "definici", "reenviada"]):
            estado = "En Progreso"
            ing = True; proy = cons = False
            comp, tot = 1, 3
        elif "urgente" in status_n:
            estado = "Urgente"
            ing = proy = cons = False
            comp, tot = 0, 3
        else:
            estado = "Pendiente"
            ing = proy = cons = False
            comp, tot = 0, 3

        # Notas enriquecidas con fechas técnica/económica si existen
        extras = []
        if status and normalize(status) not in ("enviada", "pendiente", ""):
            extras.append(status)
        if f_econ_d:
            extras.append(f"Eco: {fmt_date(f_econ_d)}")
        if f_tec_d:
            extras.append(f"Tec: {fmt_date(f_tec_d)}")
        if notas:
            extras.append(notas)
        notas_out = " | ".join(extras)

        offers.append({
            "nro":       nro,
            "desc":      desc,
            "pais":      pais,
            "cli":       cli,
            "estado":    estado,
            "status_xl": status,
            "f_ent":     fmt_date(f_ent_d),
            "f_lim":     fmt_date(f_lim_d),
            "f_sal":     fmt_date(f_sal_d),
            "asig":      asig,
            "notas":     notas_out,
            "prio":      0,
            "ing":       ing,
            "proy":      proy,
            "cons":      cons,
            "comp":      comp,
            "tot":       tot,
        })

    return offers


# ─────────────────────────────────────────────────────────────────
# PROCESAMIENTO — FORMATO ASANA (original, sin cambios)
# ─────────────────────────────────────────────────────────────────

def find_header_row(ws, max_scan=10):
    for r in range(1, max_scan + 1):
        values = [c.value for c in ws[r]]
        if "Task ID" in values and "Name" in values:
            return r, values
    return None, None


def is_completed_asana(row, idx, col_completed, col_completed_at):
    if col_completed is not None:
        v = row[idx[col_completed]]
        if v in (True, "true", "True", 1, "1", "Si", "Sí", "si", "sí"):
            return True
    if col_completed_at is not None:
        if to_date(row[idx[col_completed_at]]) is not None:
            return True
    return False


def build_offers_asana(ws, header_row, headers):
    """Formato Asana original — subtareas por oferta."""
    idx = {name: i for i, name in enumerate(headers) if name is not None}
    headers_norm = {name: normalize(name) for name in idx}

    col_name        = "Name" if "Name" in idx else None
    col_parent      = ("Parent task" if "Parent task" in idx
                       else find_column(headers_norm, "parent", "task"))
    col_section     = ("Section/Column" if "Section/Column" in idx
                       else find_column(headers_norm, "section"))
    col_assignee    = ("Assignee" if "Assignee" in idx
                       else find_column(headers_norm, "assignee"))
    col_due         = "Due Date" if "Due Date" in idx else find_column(headers_norm, "due", "date")
    col_start       = "Start Date" if "Start Date" in idx else find_column(headers_norm, "start", "date")
    col_completed   = "Completed" if "Completed" in idx else find_column(headers_norm, "completad")
    col_completed_at = ("Completed At" if "Completed At" in idx
                        else find_column(headers_norm, "completed", "at"))

    col_f_ent       = find_column(headers_norm, "fecha", "entrega") or col_due
    col_f_sal       = find_column(headers_norm, "fecha", "finaliza")
    col_cliente     = find_column(headers_norm, "cliente")
    col_pais        = find_column(headers_norm, "pais")
    col_responsable = find_column(headers_norm, "responsable")
    col_nro         = (find_column(headers_norm, "numero", "oferta")
                       or find_column(headers_norm, "numero", "ofe"))
    col_notas       = find_column(headers_norm, "observacio")
    col_estado      = find_column(headers_norm, "estado")

    print("  [Asana] Columnas detectadas:")
    print(f"    Name={col_name!r} Parent={col_parent!r} Section={col_section!r}")

    if not col_nro:
        sys.exit("ERROR Asana: no se encontró la columna 'Numero de oferta'.")

    def get(row, col):
        if not col or row is None:
            return None
        return row[idx[col]]

    grupos, orden, nro_display = {}, [], {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = str_val(get(row, col_name))
        if not name:
            continue
        nro_raw = str_val(get(row, col_nro))
        cli_raw = str_val(get(row, col_cliente))
        if not nro_raw:
            if not cli_raw:
                continue
            nro_raw = cli_raw
        key = nro_raw.strip().upper() + "|" + normalize(cli_raw)
        if key not in grupos:
            grupos[key] = []
            orden.append(key)
            nro_display[key] = nro_raw
        grupos[key].append(row)

    def first_nonempty(rows_, col):
        if not col:
            return ""
        for r in rows_:
            v = str_val(get(r, col))
            if v:
                return v
        return ""

    offers = []
    for nro in orden:
        rows_grp = grupos[nro]
        pais            = first_nonempty(rows_grp, col_pais) or "Sin pais"
        cliente         = first_nonempty(rows_grp, col_cliente) or "-"
        notas           = first_nonempty(rows_grp, col_notas)
        estado_field    = first_nonempty(rows_grp, col_estado)
        responsable_field = first_nonempty(rows_grp, col_responsable)

        f_ent_d = f_sal_d = None
        ing = proy = cons = None
        cons_responsable = any_responsable = ""

        for r in rows_grp:
            name_n    = normalize(str_val(get(r, col_name)))
            f_fin_d   = to_date(get(r, col_f_sal))
            completed = (is_completed_asana(r, idx, col_completed, col_completed_at)
                         or (f_fin_d is not None))
            asignee_r = str_val(get(r, col_assignee))
            if asignee_r and not any_responsable:
                any_responsable = asignee_r

            if "ingenier" in name_n:
                ing = bool(ing) or completed
                if f_ent_d is None:
                    f_ent_d = ((to_date(get(r, col_start)) if col_start else None)
                               or to_date(get(r, col_due)))
            elif "proyecto" in name_n:
                proy = bool(proy) or completed
                if f_ent_d is None:
                    f_ent_d = ((to_date(get(r, col_start)) if col_start else None)
                               or to_date(get(r, col_due)))
                if asignee_r:
                    cons_responsable = asignee_r
            elif "consolidado" in name_n or name_n in ("ofertas", "ofert", "oferta"):
                cons = bool(cons) or completed
                if f_fin_d:
                    f_sal_d = f_fin_d

        if f_ent_d is None:
            for r in rows_grp:
                d = to_date(get(r, col_f_ent))
                if d:
                    f_ent_d = d
                    break

        responsable = cons_responsable or responsable_field or any_responsable or "Sin Asignar"

        flags      = [ing, proy, cons]
        aplicables = [f for f in flags if f is not None]
        comp       = sum(1 for f in aplicables if f)
        tot        = len(aplicables) if aplicables else 3
        if not aplicables:
            ing = proy = cons = False

        if estado_field:
            estado = estado_field
        elif cons:
            estado = "Enviada"
        elif ing and proy:
            estado = "En Progreso"
        elif comp > 0:
            estado = "Parcial"
        else:
            estado = "Pendiente"

        offers.append({
            "nro": nro_display[nro], "desc": cliente, "pais": pais,
            "cli": cliente, "estado": estado, "status_xl": estado,
            "f_ent": fmt_date(f_ent_d), "f_lim": "", "f_sal": fmt_date(f_sal_d),
            "asig": responsable, "notas": notas, "prio": 0,
            "ing": ing, "proy": proy, "cons": cons, "comp": comp, "tot": tot,
        })

    return offers


# ─────────────────────────────────────────────────────────────────
# DISPATCHER — detecta formato y llama la función correcta
# ─────────────────────────────────────────────────────────────────

def build_offers(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    fmt, header_row, headers = detect_format(ws)
    print(f"  Formato detectado: {fmt.upper()}")

    if fmt == "onedrive":
        return build_offers_onedrive(ws, header_row, headers)
    elif fmt == "asana":
        return build_offers_asana(ws, header_row, headers)
    else:
        # Último recurso: probar Asana
        header_row, headers = find_header_row(ws)
        if header_row:
            print("  Formato detectado: ASANA (fallback)")
            return build_offers_asana(ws, header_row, headers)
        sys.exit(
            "ERROR: No se reconoce el formato del Excel.\n"
            "  Asegúrate de que tenga encabezados 'Task ID/Name' (Asana) "
            "o 'Nº Oferta/ENVIO' (OneDrive)."
        )


# ─────────────────────────────────────────────────────────────────
# AGREGACIONES (sin cambios)
# ─────────────────────────────────────────────────────────────────

def parse_ddmmyyyy(s):
    if not s:
        return None
    try:
        d, m, y = s.split("-")
        return datetime.date(int(y), int(m), int(d))
    except (ValueError, AttributeError):
        return None


def iso_week_label(date):
    return date.isocalendar()[1]


def build_aggregates(offers, today):
    eventos = {}
    for o in offers:
        fe = parse_ddmmyyyy(o["f_ent"])
        if fe:
            eventos.setdefault(fe, {"ent": 0, "sal": 0})
            eventos[fe]["ent"] += 1
        fs = parse_ddmmyyyy(o["f_sal"])
        if fs:
            eventos.setdefault(fs, {"ent": 0, "sal": 0})
            eventos[fs]["sal"] += 1

    if not eventos:
        return [], [], [], [], [], []

    fechas = sorted(eventos.keys())
    min_fecha, max_fecha = fechas[0], fechas[-1]
    DOW = ["Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo"]

    day_rows = []
    cur = min_fecha
    while cur <= max_fecha:
        if cur in eventos:
            ev = eventos[cur]
            day_rows.append({
                "date":  cur.isoformat(),
                "label": cur.strftime("%d-%m"),
                "dow":   DOW[cur.weekday()],
                "ent":   ev["ent"],
                "sal":   ev["sal"],
                "week":  iso_week_label(cur),
            })
        cur += datetime.timedelta(days=1)

    semanas = {}
    for r in day_rows:
        w = r["week"]
        semanas.setdefault(w, {"ent": 0, "sal": 0})
        semanas[w]["ent"] += r["ent"]
        semanas[w]["sal"] += r["sal"]
    week_rows = [{"week": w, "label": f"Semana {w}", "ent": v["ent"], "sal": v["sal"]}
                 for w, v in sorted(semanas.items())]

    pais_acc, asig_acc = {}, {}
    for o in offers:
        p = pais_acc.setdefault(o["pais"], {"total": 0, "env": 0})
        p["total"] += 1
        if o["estado"] == "Enviada":
            p["env"] += 1
        a = asig_acc.setdefault(o["asig"] or "Sin Asignar", {"total": 0, "env": 0})
        a["total"] += 1
        if o["estado"] == "Enviada":
            a["env"] += 1

    pais_rows = sorted(pais_acc.items(), key=lambda kv: kv[1]["total"], reverse=True)
    asig_rows = sorted(asig_acc.items(), key=lambda kv: kv[1]["total"], reverse=True)

    return day_rows, day_rows, week_rows, week_rows, pais_rows, asig_rows


def build_stats(offers):
    enviadas   = sum(1 for o in offers if o["estado"] == "Enviada")
    parciales  = sum(1 for o in offers if o["estado"] == "Parcial")
    en_progreso= sum(1 for o in offers if normalize(o["estado"]) in
                     ("en progreso", "en revision", "en aclaraciones", "en espera"))
    pendientes = sum(1 for o in offers if normalize(o["estado"]) == "pendiente")
    urgentes   = sum(1 for o in offers if "urgente" in normalize(o["estado"]))
    return {
        "entradas":   len(offers),
        "enviadas":   enviadas,
        "parciales":  parciales,
        "en_progreso":en_progreso,
        "pendientes": pendientes,
        "urgentes":   urgentes,
    }


# ─────────────────────────────────────────────────────────────────
# ESCRITURA DEL HTML (sin cambios)
# ─────────────────────────────────────────────────────────────────

D_RE           = re.compile(r"var D=\{.*?\};\nvar OFFERS", re.S)
TODAY_RE       = re.compile(r"var TODAY='[^']*';")
ACTUALIZADO_RE = re.compile(r"(Actualizado: )\d{2}/\d{2}/\d{4}(?: \d{2}:\d{2})?")
ASANA_FECHA_RE = re.compile(r"(Asana \+ Excel . )\d{2}/\d{2}/\d{4}(?: \d{2}:\d{2})?")
HBADGE_RE      = re.compile(r"(<span class=\"hbadge\">)\d+( entradas . )\d+( enviadas</span>)")
KPI_ENTRADAS_RE= re.compile(r"(Entradas</div><div class=\"kpi-val\")(>)\d+(</div>)")


def replace_data(html, D, today_str_iso, today_str_disp, hora_str):
    new_d = "var D=" + json.dumps(D, ensure_ascii=False) + ";\nvar OFFERS"
    html = D_RE.sub(new_d, html, count=1)
    html = TODAY_RE.sub(f"var TODAY='{today_str_iso}';", html)
    fecha_hora = f"{today_str_disp} {hora_str}"
    html = ACTUALIZADO_RE.sub(rf"\g<1>{fecha_hora}", html)
    html = ASANA_FECHA_RE.sub(rf"\g<1>{fecha_hora}", html)
    html = HBADGE_RE.sub(
        rf"\g<1>{D['stats']['entradas']}\g<2>{D['stats']['enviadas']}\g<3>", html
    )
    html = KPI_ENTRADAS_RE.sub(
        rf"\g<1>\g<2>{D['stats']['entradas']}\g<3>", html
    )
    return html


# ─────────────────────────────────────────────────────────────────
# MAIN (sin cambios)
# ─────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        sys.exit(
            "USO: python3 scripts/ofertas_dashboard.py "
            "<carpeta_data> ofertas2026.html [archivo.xlsx]"
        )

    data_dir  = sys.argv[1]
    html_path = sys.argv[2]
    explicit  = sys.argv[3] if len(sys.argv) > 3 else None

    if not os.path.exists(html_path):
        sys.exit(f"ERROR: no se encontró el archivo {html_path}")

    xlsx_path = locate_xlsx(data_dir, explicit)
    print(f"Usando archivo: {xlsx_path}")

    now   = datetime.datetime.now(TZ)
    today = now.date()

    offers = build_offers(xlsx_path)
    print(f"  {len(offers)} ofertas detectadas")

    day_rows, day_rows_chart, week_rows, week_rows_chart, pais_rows, asig_rows = \
        build_aggregates(offers, today)

    D = {
        "offers":          offers,
        "day_rows":        day_rows,
        "week_rows":       week_rows,
        "pais_rows":       pais_rows,
        "asig_rows":       asig_rows,
        "stats":           build_stats(offers),
        "day_rows_chart":  day_rows_chart,
        "week_rows_chart": week_rows_chart,
        "pre_sal_offset":  0,
        "pre_ent_offset":  0,
    }

    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()

    today_iso  = today.isoformat()
    today_disp = today.strftime("%d/%m/%Y")
    hora_str   = now.strftime("%H:%M")
    html_out   = replace_data(html, D, today_iso, today_disp, hora_str)

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_out)

    s = D["stats"]
    print(
        f"OK -> {html_path} actualizado "
        f"({s['entradas']} ofertas, {s['enviadas']} enviadas, "
        f"{s['pendientes']} pendientes)"
    )


if __name__ == "__main__":
    main()
