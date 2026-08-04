#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ENRIQUECER - Fecha de Despacho + Pronostico de termino
======================================================

Lee los XLSX que deja asana_exporter.py en /data y para cada proyecto calcula:

  1. despacho : la fecha de la tarea "Despacho" que cuelga de la fase Logistica.
  2. pron     : pronostico de fecha de termino del proyecto, en base a la
                velocidad real con la que se han ido cerrando subtareas.
  3. atraso   : dias de diferencia entre el pronostico y la fecha de entrega
                comprometida (negativo = va adelantado).
  4. conf     : confianza del pronostico (alta / media / baja).

Uso
---
    python enriquecer.py                          # solo genera pronosticos.json
    python enriquecer.py --html ../index.html     # ademas parchea el dashboard

El parche del HTML es idempotente: se puede correr las veces que se quiera.
"""

import argparse
import json
import re
import sys
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

import calendario_cl as cal

BASE = Path(__file__).resolve().parent
DATA = BASE / "data"

# ── Parametros del modelo de pronostico ──────────────────────────────────────
# Todo el modelo trabaja en DIAS HABILES DE CHILE (lunes a viernes sin feriados).
VENTANA_HABILES  = 22      # habiles hacia atras para medir la velocidad reciente
PESO_RECIENTE    = 0.70    # cuanto pesa la velocidad reciente vs la historica
MIN_CIERRES_ALTA = 12      # cierres en la ventana para considerar confianza alta
MIN_CIERRES_MEDIA = 5
TOPE_HABILES     = 500     # si el pronostico se dispara mas alla, se marca asi

# Palabras que identifican la tarea de despacho dentro de la fase Logistica.
# Si en Asana se llama distinto, agregar la palabra aca (en minusculas, sin tildes).
PALABRAS_DESPACHO = ("despach", "embarq", "envio", "entrega a obra", "shipping")

# Palabras que identifican la fase de logistica
PALABRAS_LOGISTICA = ("logist", "despacho")

# ── Etiquetas de los tags en la tarjeta ──────────────────────────────────────
# Editar aca si se quieren nombres mas cortos (las tarjetas se hacen angostas
# cuando los tres tags son largos).
ETIQUETA_ENTREGA    = "Fecha entrega cliente"   # viene del portafolio de Asana
ETIQUETA_DESPACHO   = "Fecha despacho Asana"    # tarea Despacho de la fase Logistica
ETIQUETA_PRONOSTICO = "Pronostico de entrega"   # calculado por velocidad


# ── Utilidades ───────────────────────────────────────────────────────────────
def norm(s) -> str:
    """minusculas, sin tildes, sin espacios sobrantes."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


def a_fecha(v):
    """Convierte lo que venga del XLSX a date, o None."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    return None


def ddmmyyyy(d):
    return d.strftime("%d/%m/%Y") if d else ""


# Asana exporta los encabezados en el idioma de la cuenta: aceptamos ambos.
COLUMNAS = {
    "id":         ["task id", "id de la tarea", "id"],
    "nombre":     ["name", "nombre", "nombre de la tarea"],
    "creado":     ["created at", "fecha de creacion", "creado el"],
    "completado": ["completed at", "fecha de finalizacion", "completado el",
                   "fecha de realizacion"],
    "vence":      ["due date", "fecha de vencimiento", "fecha limite",
                   "fecha de entrega"],
    "inicio":     ["start date", "fecha de inicio"],
    "padre":      ["parent task", "tarea principal", "tarea padre",
                   "parent task id"],
    "seccion":    ["section/column", "seccion/columna", "seccion", "columna"],
}


def mapear_columnas(encabezados):
    """{clave_interna: indice_de_columna}"""
    m = {}
    normalizados = [norm(h) for h in encabezados]
    for clave, alias in COLUMNAS.items():
        for i, h in enumerate(normalizados):
            if h in alias:
                m[clave] = i
                break
        else:
            for i, h in enumerate(normalizados):
                if h and any(a in h for a in alias):
                    m[clave] = i
                    break
    return m


def buscar_fila_encabezado(ws, max_scan=10):
    """
    Misma logica que generar_portafolio.py: el encabezado es la primera fila
    (dentro de las primeras 10) que tenga 'Task ID' y 'Name'.
    """
    for r in range(1, max_scan + 1):
        valores = [c.value for c in ws[r]]
        if "Task ID" in valores and "Name" in valores:
            return r, valores
    # Respaldo por si el export viniera localizado al espaniol
    for r in range(1, max_scan + 1):
        valores = [c.value for c in ws[r]]
        vn = [norm(v) for v in valores]
        if any(v in ("name", "nombre") for v in vn):
            return r, valores
    return None, None


def leer_export(ruta: Path):
    """
    Devuelve (lista_de_tareas, nombre_proyecto).

    IMPORTANTE: se abre SIN read_only. Los XLSX de Asana declaran mal el rango
    de dimensiones y en modo read_only openpyxl devuelve solo el encabezado.
    Es el mismo criterio que usa generar_portafolio.py.
    """
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active

    fila_enc, encabezados = buscar_fila_encabezado(ws)
    if fila_enc is None:
        wb.close()
        return [], ruta.stem

    col = mapear_columnas(encabezados)
    tareas = []
    for fila in ws.iter_rows(min_row=fila_enc + 1, values_only=True):
        def g(clave):
            i = col.get(clave)
            return fila[i] if i is not None and i < len(fila) else None

        nombre = g("nombre")
        if nombre in (None, ""):
            continue

        tareas.append({
            "id":         str(g("id") or "").strip(),
            "nombre":     str(nombre).strip(),
            "creado":     a_fecha(g("creado")),
            "completado": a_fecha(g("completado")),
            "vence":      a_fecha(g("vence")),
            "inicio":     a_fecha(g("inicio")),
            "padre":      str(g("padre") or "").strip(),
            "seccion":    str(g("seccion") or "").strip(),
        })
    wb.close()
    return tareas, ruta.stem


# ── 1. Fecha de despacho ─────────────────────────────────────────────────────
def cadena_padres(tarea, por_id, por_nombre, tope=6):
    """Nombres de todos los ancestros de una tarea, de abajo hacia arriba."""
    cadena, actual, visto = [], tarea, set()
    for _ in range(tope):
        ref = (actual.get("padre") or "").strip()
        if not ref or ref in visto:
            break
        visto.add(ref)
        padre = por_id.get(ref) or por_nombre.get(norm(ref))
        if padre is None:
            cadena.append(ref)
            break
        cadena.append(padre["nombre"])
        actual = padre
    return cadena


def buscar_despacho(tareas):
    """
    Fecha de la tarea de despacho (la que cuelga de la fase Logistica).

    En Asana la estructura es:
        Logistica:            <- fase nivel 1 (cabecera, sin fecha)
          Despacho            <- nivel 2, ESTA es la que buscamos
            ...5 subtareas    <- nivel 3, pueden traer la fecha real

    Prioridad de fecha: vencimiento propio -> vencimiento mas tardio de sus
    subtareas -> completada propia -> completada mas tardia de sus subtareas
    -> inicio. Devuelve (fecha, origen).
    """
    por_id = {t["id"]: t for t in tareas if t["id"]}
    por_nombre = {norm(t["nombre"]): t for t in tareas if t["nombre"]}

    # hijos por nombre de padre, para poder mirar el nivel 3
    hijos = {}
    for t in tareas:
        if t["padre"]:
            hijos.setdefault(norm(t["padre"]), []).append(t)

    candidatas = []
    for t in tareas:
        n = norm(t["nombre"])
        if not any(pal in n for pal in PALABRAS_DESPACHO):
            continue
        # las fases de nivel 1 son cabeceras, nunca llevan la fecha
        if not t["padre"]:
            continue
        ancestros = [norm(a) for a in cadena_padres(t, por_id, por_nombre)]
        contexto = " ".join(ancestros + [norm(t["seccion"])])
        bajo_logistica = any(pal in contexto for pal in PALABRAS_LOGISTICA)
        candidatas.append((bajo_logistica, t))

    if not candidatas:
        return None, ""

    candidatas.sort(key=lambda c: not c[0])   # primero las de Logistica
    bajo_logistica, t = candidatas[0]
    origen = "Logistica" if bajo_logistica else "tarea suelta"

    subtareas = hijos.get(norm(t["nombre"]), [])

    if t["vence"]:
        return t["vence"], f"{origen} · vencimiento"

    vencimientos = [h["vence"] for h in subtareas if h["vence"]]
    if vencimientos:
        return max(vencimientos), f"{origen} · vencimiento de subtareas"

    if t["completado"]:
        return t["completado"], f"{origen} · completada"

    completadas = [h["completado"] for h in subtareas if h["completado"]]
    if completadas:
        return max(completadas), f"{origen} · subtareas completadas"

    if t["inicio"]:
        return t["inicio"], f"{origen} · inicio"

    return None, f"{origen} · sin fecha cargada"


def hijos_de_logistica(tareas):
    """Nombres de las tareas que cuelgan de la fase Logistica (solo para el log)."""
    por_id = {t["id"]: t for t in tareas if t["id"]}
    por_nombre = {norm(t["nombre"]): t for t in tareas if t["nombre"]}
    out = []
    for t in tareas:
        ancestros = [norm(a) for a in cadena_padres(t, por_id, por_nombre)]
        if any(pal in " ".join(ancestros) for pal in PALABRAS_LOGISTICA):
            out.append(t["nombre"])
    return out


# ── 2. Pronostico de termino ─────────────────────────────────────────────────
def calcular_pronostico(tareas, hoy):
    """
    Modelo de velocidad medido en DIAS HABILES DE CHILE: cuantas subtareas se
    cierran por dia habil, y cuantos dias habiles faltan para cerrar el resto.
    La fecha final se obtiene sumando esos habiles a hoy, saltando fines de
    semana y feriados, asi que nunca cae en sabado, domingo ni feriado.

    Se cuentan solo las subtareas (las que tienen tarea principal), que es el
    mismo universo que usa el dashboard para el % de avance.
    """
    subtareas = [t for t in tareas if t["padre"]]
    if not subtareas:
        subtareas = tareas

    total = len(subtareas)
    cerradas = [t for t in subtareas if t["completado"]]
    hechas = len(cerradas)
    faltan = total - hechas

    res = {
        "total": total, "hechas": hechas, "faltan": faltan,
        "fecha": None, "conf": "baja", "nota": "", "vel": 0.0, "habiles": None,
    }

    if total == 0:
        res["nota"] = "sin subtareas en el export"
        return res

    if faltan <= 0:
        ultima = max(t["completado"] for t in cerradas)
        res.update(fecha=ultima, conf="alta", nota="proyecto cerrado", habiles=0)
        return res

    fechas_cierre = [t["completado"] for t in cerradas]
    if not fechas_cierre:
        res["nota"] = "sin subtareas cerradas: no hay velocidad medible"
        return res

    # velocidad reciente, sobre los ultimos VENTANA_HABILES dias habiles
    corte = cal.restar_habiles(hoy, VENTANA_HABILES)
    recientes = [f for f in fechas_cierre if corte <= f <= hoy]
    vel_reciente = len(recientes) / VENTANA_HABILES

    # velocidad historica desde el primer movimiento del proyecto
    arranques = [t["creado"] for t in subtareas if t["creado"]] + fechas_cierre
    inicio = min(arranques)
    habiles_vividos = max(cal.dias_habiles(inicio, hoy), 1)
    vel_historica = hechas / habiles_vividos

    if vel_reciente > 0:
        vel = PESO_RECIENTE * vel_reciente + (1 - PESO_RECIENTE) * vel_historica
        n = len(recientes)
        conf = "alta" if n >= MIN_CIERRES_ALTA else ("media" if n >= MIN_CIERRES_MEDIA else "baja")
        nota = f"{n} cierres en {VENTANA_HABILES} dias habiles"
    else:
        vel = vel_historica
        conf = "baja"
        nota = f"sin cierres hace {VENTANA_HABILES}+ habiles, se usa promedio historico"

    if vel <= 0:
        res.update(nota="velocidad cero: proyecto detenido")
        return res

    habiles = faltan / vel
    if habiles > TOPE_HABILES:
        res.update(vel=round(vel, 3), conf="baja",
                   nota=f"al ritmo actual supera {TOPE_HABILES} dias habiles: revisar")
        return res

    habiles = int(round(habiles))
    res.update(fecha=cal.sumar_habiles(hoy, habiles), conf=conf,
               nota=f"{nota}; faltan ~{habiles} dias habiles", vel=round(vel, 3),
               habiles=habiles)
    return res


# ── 3. Recorrer los exports ──────────────────────────────────────────────────
def procesar(carpeta: Path, hoy: date):
    resultados = {}
    sin_despacho_mostrado = [False]
    archivos = sorted(list(carpeta.glob("*.xlsx")) + list(carpeta.glob("*.csv")))
    archivos = [a for a in archivos if not a.name.startswith("~$")
                and "PORTAFOLIO" not in a.name.upper()]

    for ruta in archivos:
        if ruta.suffix.lower() != ".xlsx":
            print(f"  [omitido] {ruta.name} (solo se leen .xlsx)")
            continue
        try:
            tareas, nombre = leer_export(ruta)
        except Exception as e:
            print(f"  [ERROR] {ruta.name}: {e}")
            continue

        if not tareas:
            print(f"  [AVISO] {ruta.name}: 0 tareas leidas (revisar encabezados)")

        despacho, origen = buscar_despacho(tareas)
        pron = calcular_pronostico(tareas, hoy)

        resultados[nombre] = {
            "proyecto":   nombre,
            "despacho":   ddmmyyyy(despacho),
            "despacho_origen": origen,
            "pron":       ddmmyyyy(pron["fecha"]),
            "conf":       pron["conf"],
            "nota":       pron["nota"],
            "vel_dia":    pron["vel"],
            "faltan":     pron["faltan"],
            "habiles":    pron["habiles"],
            "total":      pron["total"],
        }
        print(f"  {nombre[:44]:<46} tareas={len(tareas):<4}"
              f" despacho={resultados[nombre]['despacho'] or '-':<11}"
              f" pron={resultados[nombre]['pron'] or '-':<11} ({pron['conf']})")
        if pron["nota"]:
            print(f"       {pron['nota']}")
        if origen and not despacho:
            print(f"       [!] Tarea de despacho encontrada pero {origen}")

        # Si no se encontro la tarea de despacho, mostrar que cuelga de
        # Logistica para poder ajustar PALABRAS_DESPACHO.
        if not despacho and not sin_despacho_mostrado[0]:
            hijos = hijos_de_logistica(tareas)
            if hijos:
                sin_despacho_mostrado[0] = True
                print(f"       [?] No hay tarea de despacho. Bajo Logistica hay: "
                      f"{hijos[:12]}")
                print(f"       Si alguna es el despacho, agregar su palabra a "
                      f"PALABRAS_DESPACHO en enriquecer.py")

    return resultados


# ── 4. Parche del HTML ───────────────────────────────────────────────────────
MARCA = "/* despacho-pronostico */"

CSS_EXTRA = """
.desp-tag{font-size:10px;color:#5B21B6;padding:2px 7px;border-radius:8px;background:#F5F3FF;border:1px solid #DDD6FE;white-space:nowrap;font-weight:600;}
.pron-tag{font-size:10px;padding:2px 7px;border-radius:8px;white-space:nowrap;font-weight:600;background:#F1F5F9;border:1px solid #CBD5E1;color:#334155;}
.pron-tag.ok{background:#ECFDF5;border-color:#A7F3D0;color:#065F46;}
.pron-tag.riesgo{background:#FEF3C7;border-color:#FDE68A;color:#92400E;}
.pron-tag.critico{background:#FEE2E2;border-color:#FECACA;color:#991B1B;}
.ptop{gap:6px 8px;align-items:center;}
.pname{flex:1 1 100%;min-width:0;line-height:1.35;margin-bottom:2px;}
.pcard .ppct{margin-left:auto;}
@media(max-width:820px){
  .desp-tag,.pron-tag,.entrega-tag,.exw-tag{font-size:9px;padding:2px 6px;}
  .pcnt{min-width:0;}
}
"""

JS_TAGS = """
    // Fecha de despacho tomada de la fase Logistica en Asana
    var despTag='';
    if(x.despacho){
      var dDesp=parseExw(x.despacho);
      var despVenc=dDesp&&dDesp<TODAY&&!desp;
      despTag='<span class="desp-tag"'+(despVenc?' style="background:#FEE2E2;border-color:#FECACA;color:#991B1B"':'')
        +' title="Tarea Despacho (Logistica)">\\uD83D\\uDE9A __ETQ_DESPACHO__ '+x.despacho+'</span>';
    }
    // Pronostico de termino segun velocidad real de cierre de subtareas
    var pronTag='';
    if(x.pron&&!desp){
      var dPron=parseExw(x.pron), dEnt=parseExw(x.entrega);
      var clase='', extra='';
      if(dPron&&dEnt){
        var difDias=Math.round((dPron-dEnt)/86400000);
        clase=difDias<=0?'ok':(difDias<=15?'riesgo':'critico');
        extra=difDias>0?(' +'+difDias+'d'):(difDias<0?(' '+difDias+'d'):' en fecha');
      }
      pronTag='<span class="pron-tag '+clase+'" title="Pronostico por velocidad. Confianza: '
        +(x.conf||'baja')+'. '+(x.nota||'')+'">\\uD83D\\uDD2E __ETQ_PRONOSTICO__ '+x.pron+extra+'</span>';
    }
"""


def parchear_html(ruta_html: Path, mapa: dict):
    html = ruta_html.read_text(encoding="utf-8")

    # 4.1 datos: inyectar despacho / pron / conf / nota dentro de var P = [...]
    m = re.search(r"var\s+P\s*=\s*", html)
    if not m or "[" not in html[m.end():m.end() + 5]:
        print("  [ERROR] no se encontro el bloque 'var P = [...]' en el HTML")
        return False

    ini = html.index("[", m.end())
    proyectos, largo = json.JSONDecoder().raw_decode(html, ini)
    fin = largo
    while fin < len(html) and html[fin] in " ;":
        fin += 1
    indice = {norm(k): v for k, v in mapa.items()}

    aciertos = 0
    for p in proyectos:
        clave = norm(p.get("n", ""))
        dato = indice.get(clave)
        if dato is None:
            # match tolerante por el numero de pedido (500015xx)
            num = re.match(r"\s*(\d{6,})", p.get("n", ""))
            if num:
                for k, v in indice.items():
                    if k.startswith(num.group(1)):
                        dato = v
                        break
        if dato:
            aciertos += 1
            p["despacho"] = dato["despacho"]
            p["pron"] = dato["pron"]
            p["conf"] = dato["conf"]
            p["nota"] = dato["nota"]
        else:
            p.setdefault("despacho", "")
            p.setdefault("pron", "")
            p.setdefault("conf", "")
            p.setdefault("nota", "")

    nuevo = "var P = " + json.dumps(proyectos, ensure_ascii=False) + ";"
    html = html[:m.start()] + nuevo + html[fin:]
    print(f"  datos inyectados en {aciertos}/{len(proyectos)} proyectos")

    # 4.2 renombrar el tag de Entrega que viene del TEMPLATE de
    # generar_portafolio.py. Se hace siempre porque ese script regenera el
    # HTML en cada corrida. El ancla no incluye emojis a proposito, para no
    # depender de como quedaron codificados.
    if "Entrega '+(entregaVencida" in html:
        html = html.replace("Entrega '+(entregaVencida",
                            ETIQUETA_ENTREGA + " '+(entregaVencida", 1)
        print(f"  tag de entrega renombrado a '{ETIQUETA_ENTREGA}'")

    # 4.3 estilos y render (solo la primera vez)
    if MARCA not in html:
        html = html.replace(
            ".entrega-tag.vencida{",
            MARCA + CSS_EXTRA + ".entrega-tag.vencida{", 1)

        tags_js = (JS_TAGS
                   .replace("__ETQ_DESPACHO__", ETIQUETA_DESPACHO)
                   .replace("__ETQ_PRONOSTICO__", ETIQUETA_PRONOSTICO))
        ancla = "    return '<div class=\"pcard\" id=\"c'+i+'\""
        if ancla in html:
            html = html.replace(ancla, tags_js + ancla, 1)
        else:
            print("  [AVISO] no se encontro el ancla del render de tarjetas")

        html = html.replace("+entregaTag\n", "+entregaTag\n      +despTag\n      +pronTag\n", 1)

        # orden por pronostico
        html = html.replace(
            '<option value="exwz">EXW mas reciente primero</option>',
            '<option value="exwz">EXW mas reciente primero</option>\n'
            '      <option value="prona">Pronostico mas cercano</option>\n'
            '      <option value="atraso">Mayor atraso proyectado</option>', 1)

        html = html.replace(
            '  } else d.sort(function(a,b){return a.n.localeCompare(b.n);});',
            '''  } else if(s==="prona"){
    d.sort(function(a,b){
      var da=parseExw(a.pron),db=parseExw(b.pron);
      if(!da&&!db)return 0; if(!da)return 1; if(!db)return -1; return da-db;
    });
  } else if(s==="atraso"){
    var atr=function(x){
      var dp=parseExw(x.pron),de=parseExw(x.entrega);
      if(!dp||!de)return -99999;
      return Math.round((dp-de)/86400000);
    };
    d.sort(function(a,b){return atr(b)-atr(a);});
  } else d.sort(function(a,b){return a.n.localeCompare(b.n);});''', 1)

        # KPI de proyectos con pronostico posterior a la entrega comprometida
        html = html.replace(
            '    {l:"EXW vencida",v:vencidos,c:"#E24B4A"},',
            '''    {l:"EXW vencida",v:vencidos,c:"#E24B4A"},
    {l:"Termino en riesgo",v:P.filter(function(x){
        if(isDespachado(x))return false;
        var dp=parseExw(x.pron),de=parseExw(x.entrega);
        return dp&&de&&dp>de;
      }).length,c:"#E24B4A"},''', 1)
        print("  render y estilos parcheados")
    else:
        print("  render ya estaba parcheado, solo se refrescaron los datos")

    ruta_html.write_text(html, encoding="utf-8")
    return True


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", default=str(DATA), help="carpeta con los XLSX")
    ap.add_argument("--html", default=None, help="dashboard a parchear")
    ap.add_argument("--hoy", default=None, help="fecha de calculo YYYY-MM-DD")
    ap.add_argument("--out", default=None, help="ruta del JSON de salida")
    args = ap.parse_args()

    hoy = a_fecha(args.hoy) or date.today()
    carpeta = Path(args.data)
    if not carpeta.exists():
        sys.exit(f"No existe la carpeta {carpeta}")

    print("=" * 70)
    print(f"  Despacho + pronostico · calculado al {hoy.strftime('%d/%m/%Y')}")
    print("=" * 70)

    mapa = procesar(carpeta, hoy)

    salida = Path(args.out) if args.out else carpeta / "pronosticos.json"
    salida.write_text(json.dumps(mapa, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n  JSON -> {salida}")

    if args.html:
        ruta_html = Path(args.html)
        if not ruta_html.exists():
            sys.exit(f"No existe el HTML {ruta_html}")
        print(f"\n  Parcheando {ruta_html.name}")
        parchear_html(ruta_html, mapa)

    print("\n  Listo.")


if __name__ == "__main__":
    main()
